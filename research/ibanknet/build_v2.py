#!/usr/bin/env python3
"""IBNET-ASSEMBLER-2: build iBankNet_CA_Deposits_v2.xlsx (+ CSVs) from v1 + deep-dive results.
ADD-ONLY vs v1. No row inserts on existing tabs. New tabs 5-8."""
import csv, re, shutil, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = '/home/user/awake/research/ibanknet'
V1 = f'{BASE}/deliverables/iBankNet_CA_Deposits_v1.xlsx'
V2 = f'{BASE}/deliverables/iBankNet_CA_Deposits_v2.xlsx'
DD = ' ‖ DD 2026-06-12: '

# ---------- name canonicalization ----------
def canon(s):
    s = ' '.join(str(s).upper().split())
    s = re.sub(r'\s*\([^)]*\)\s*$', '', s).strip()
    s = re.sub(r',?\s*N\.A\.$', ', NATIONAL ASSOCIATION', s)
    s = s.rstrip('.').strip()
    return s

def first_url(s):
    m = re.search(r'https?://[^\s|;,)\]]+', str(s))
    return m.group(0) if m else None

# ---------- load inputs ----------
def load_csv(paths):
    rows = []
    for p in paths:
        with open(p, encoding='utf-8-sig') as fh:
            rs = list(csv.DictReader(fh))
        for r in rs:
            for k in list(r):
                if k.startswith('Evidence_grade') and k != 'Evidence_grade':
                    r['Evidence_grade'] = r.pop(k)
            r['_src'] = p
        rows += rs
    return rows

bureau = load_csv([f'{BASE}/out/deep_bureau_results_{i}.csv' for i in (1, 2, 3)])
chex   = load_csv([f'{BASE}/out/deep_chex_results_{i}.csv' for i in (1, 2)])
nodoc  = load_csv([f'{BASE}/out/nodoc_dps_{i}.csv' for i in (1, 2)])
extras = load_csv([f'{BASE}/out/extras_cfpb.csv'])
assert len(bureau) == 101 and len(chex) == 101 and len(extras) == 101 and len(nodoc) == 149, \
    (len(bureau), len(chex), len(extras), len(nodoc))

# ---------- workbook ----------
shutil.copy(V1, V2)
wb = openpyxl.load_workbook(V2)
ws1 = wb['1. CA INSTITUTIONS BY DEPOSITS']
tab1_row = {}
for r in range(2, ws1.max_row + 1):
    n = ws1.cell(r, 2).value
    if n:
        tab1_row[canon(n)] = r

def resolve(name):
    return tab1_row.get(canon(name))

mismatch = []
for tag, rows in [('bureau', bureau), ('chex', chex), ('extras', extras)]:
    for r in rows:
        if resolve(r['Institution']) is None:
            mismatch.append((tag, r['Institution']))
if mismatch:
    print('NAME MISMATCHES (deep tabs):', mismatch)
    sys.exit(1)
# dupe check within each deep dataset
for tag, rows in [('bureau', bureau), ('chex', chex), ('extras', extras)]:
    cs = [canon(r['Institution']) for r in rows]
    assert len(set(cs)) == 101, (tag, 'dupes', [c for c in set(cs) if cs.count(c) > 1])

nodoc_unmatched = sorted({r['Institution'] for r in nodoc if resolve(r['Institution']) is None})
print(f'nodoc rows not mapping to a single Tab-1 row (OFF-LIST/program-level, Tab 7 only): {len(nodoc_unmatched)}')

bureau_by = {canon(r['Institution']): r for r in bureau}
chex_by   = {canon(r['Institution']): r for r in chex}
extras_by = {canon(r['Institution']): r for r in extras}

# ---------- helpers ----------
def trunc(s, n=200):
    s = ' '.join(str(s).split())
    return s if len(s) <= n else s[:n - 1] + '…'

def append_cell(ws, row, col, text, url=None):
    c = ws.cell(row, col)
    old = c.value
    c.value = (str(old) if old not in (None, '') else '') + text
    if url and not c.hyperlink:
        c.hyperlink = url

# best no-doc DP per institution
def amount_val(a):
    best = 0
    for m in re.finditer(r'\$?\s?([\d,]+(?:\.\d+)?)\s*([KkMm]?)', str(a)):
        try:
            v = float(m.group(1).replace(',', ''))
        except ValueError:
            continue
        if m.group(2).upper() == 'K':
            v *= 1000
        elif m.group(2).upper() == 'M':
            v *= 1e6
        best = max(best, v)
    return best

def dp_priority(r):
    g, t = r['Evidence_grade'], r['DP_type']
    if 'VERIFIED-own-page' in g:
        p = 0
    elif t == 'approval':
        p = 1
    elif t == 'published-requirements':
        p = 2
    else:
        p = 3
    return (p, -amount_val(r['Amount']))

def dp_line(r):
    t = r['DP_type']
    if t == 'none-found':
        return f"no no-doc DP found ({trunc(r['Notes'], 110)}) [none-found]"
    amt = r['Amount'] if r['Amount'] not in ('-', '') else ''
    docs = r['Docs_required'] if r['Docs_required'] not in ('-', '') else 'docs n/s'
    head = 'DP' if t == 'approval' else 'Published reqs'
    return f"{head}: {trunc(r['Product'], 60)}{' ' + trunc(amt, 40) if amt else ''} — {trunc(docs, 110)} [{trunc(r['Evidence_grade'], 60)}]"

nodoc_best = {}
for r in nodoc:
    k = resolve(r['Institution'])
    if k is None:
        continue
    cn = canon(r['Institution'])
    if cn not in nodoc_best or dp_priority(r) < dp_priority(nodoc_best[cn]):
        nodoc_best[cn] = r

MAT_SKIP = re.compile(r'^(No change|No new|NOTHING NEW|UNKNOWN|TWIN-TRAP)', re.I)
def material(s):
    return bool(s.strip()) and not MAT_SKIP.match(s.strip())
def leniency_informative(s):
    return s.strip() not in ('', '—', '-', 'Unknown', 'Unknown leniency')

# ---------- TASK 1: enrich Tab 1 ----------
cnt = dict(bureau=0, nodoc=0, chex=0, ews=0, flags=0)
for cn, br in bureau_by.items():
    row = tab1_row[cn]
    append_cell(ws1, row, 16, f"{DD}{trunc(br['Bureau_finding'], 220)} [{trunc(br['Evidence_grade'], 90)}]",
                first_url(br['Source_links']))
    cnt['bureau'] += 1
    nb = nodoc_best.get(cn)
    if nb:
        append_cell(ws1, row, 17, DD + dp_line(nb), first_url(nb['Source_link']))
        cnt['nodoc'] += 1
    cx = chex_by[cn]
    if material(cx['New_Chex_finding']) or leniency_informative(cx['Leniency_note']):
        parts = []
        if material(cx['New_Chex_finding']):
            parts.append(trunc(cx['New_Chex_finding'], 260))
        if leniency_informative(cx['Leniency_note']):
            parts.append('Leniency: ' + trunc(cx['Leniency_note'], 160))
        append_cell(ws1, row, 18, DD + ' | '.join(parts), first_url(cx['Source_links']))
        cnt['chex'] += 1
    if material(cx['New_EWS_finding']):
        append_cell(ws1, row, 19, DD + trunc(cx['New_EWS_finding'], 260), first_url(cx['Source_links']))
        cnt['ews'] += 1
    ex = extras_by[cn]
    if ex['Material_flags'].strip():
        append_cell(ws1, row, 22, DD + trunc(ex['Material_flags'], 300), first_url(ex['Material_flags']) or first_url(ex['Source_links']))
        cnt['flags'] += 1
print('Tab 1 enrichment counts:', cnt)

# ---------- styles ----------
GREEN  = PatternFill('solid', fgColor='C6EFCE')
LGREEN = PatternFill('solid', fgColor='E2EFDA')
AMBER  = PatternFill('solid', fgColor='FFE699')
GRAY   = PatternFill('solid', fgColor='D9D9D9')
RED    = PatternFill('solid', fgColor='FFC7CE')
GOLD   = PatternFill('solid', fgColor='FFD966')
HDR    = Font(bold=True)
NOTEF  = Font(bold=True, size=9)
DIM    = Font(color='808080', size=9)
WRAP   = Alignment(wrap_text=True, vertical='top')

def new_tab(title, note, headers, widths):
    ws = wb.create_sheet(title)
    ws.cell(1, 1, note).font = NOTEF
    ws.cell(1, 1).alignment = WRAP
    ws.row_dimensions[1].height = 86
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h).font = HDR
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'
    return ws

def set_row(ws, r, vals, link_col=None, link=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v)
        c.alignment = WRAP
    if link_col and link:
        ws.cell(r, link_col).hyperlink = link

# ---------- TASK 2: Tab 5 BUREAU DEEP DIVE ----------
note5 = ('CROSS-CUTTING RULES (apply across rows): (1) Elan same-day applications = ONE combined hard pull '
         '[myFICO thread 6308920]. (2) Elan ROTATES bureaus — no single-bureau guarantee on Elan rails. '
         '(3) Pull-reuse windows exist at Logix / Kinecta / SDCCU / Stanford FCU but are NOT SAFE to rely on. '
         '(4) U.S. Bank also pulls SageStream (IDA) + ARS on applications — freeze strategy documented at DoC; '
         'US Bank approves even if frozen. (5) myFICO-sourced DPs partly read via Google snippets (forum 403s '
         'direct fetch) — snippet-graded, treat amounts/details as community anecdote, not policy. '
         'Grades: VERIFIED-disclosure (green) / COMMUNITY-DP (light green) / CONSUMER-PROXY & RAIL-INFERENCE (amber) / UNKNOWN (gray).')
h5 = ['Institution', 'Rank', 'Underwriter_rail', 'Bureau_finding', 'Hard_or_soft', 'Product_specificity',
      'Secondary_pulls', 'DP_details', 'Evidence_grade', 'DP_dates', 'Source_links', 'Notes']
ws5 = new_tab('5. BUREAU DEEP DIVE', note5, h5, [34, 6, 14, 40, 22, 30, 30, 50, 26, 18, 50, 50])
def grade_fill(g):
    gu = str(g).upper()
    if 'VERIFIED' in gu:
        return GREEN
    if 'COMMUNITY-DP' in gu:
        return LGREEN
    if 'CONSUMER-PROXY' in gu or 'RAIL-INFERENCE' in gu:
        return AMBER
    return GRAY
r5 = 3
for br in sorted(bureau, key=lambda r: int(r['Rank'])):
    vals = [br[h] for h in h5]
    set_row(ws5, r5, vals, link_col=11, link=first_url(br['Source_links']))
    f = grade_fill(br['Evidence_grade'])
    for c in range(1, 13):
        ws5.cell(r5, c).fill = f
    r5 += 1

# ---------- TASK 3: Tab 6 CHEX-EWS DEEP DIVE ----------
note6 = ('GUARDRAILS: (1) In-branch MAY bypass an automatic online ChexSystems run (verified DP: Bank of the '
         'Sierra — LetMeBank). (2) "Uses-but-LENIENT" ≠ "no Chex" — a lenient institution still pulls and records '
         'the inquiry. (3) Crediful-conflict guardrail: Crediful no-Chex lists conflict with own-doc disclosures '
         'at several institutions — own-document evidence ALWAYS outranks aggregator lists. '
         '(4) QualiFile (Chex product) embeds LexisNexis — a frozen LN file can auto-decline (see Tab 0 standing warning). '
         'Leniency colors: lenient/second-chance = green, strict/sensitive = red, unknown = gray.')
h6 = ['Institution', 'Rank', 'Prior_Chex', 'New_Chex_finding', 'Chex_business_specific', 'Prior_EWS',
      'New_EWS_finding', 'Leniency_note', 'Source_links', 'Notes']
ws6 = new_tab('6. CHEX-EWS DEEP DIVE', note6, h6, [34, 6, 30, 50, 36, 22, 36, 36, 50, 46])
REDPAT = re.compile(r'not lenient|strict|sensitiv|denial|denies|auto-deny|disqualif', re.I)
GRNPAT = re.compile(r'lenient|second.chance|gentle|bypass|tolerant|soft-pull|fresh start', re.I)
r6 = 3
for cx in sorted(chex, key=lambda r: int(r['Rank'])):
    set_row(ws6, r6, [cx[h] for h in h6], link_col=9, link=first_url(cx['Source_links']))
    ln = cx['Leniency_note']
    cell = ws6.cell(r6, 8)
    if REDPAT.search(ln):
        cell.fill = RED
    elif GRNPAT.search(ln):
        cell.fill = GREEN
    else:
        cell.fill = GRAY
    r6 += 1

# ---------- TASK 4: Tab 7 NO-DOC DP LOG ----------
note7 = ('NO-DOC DP LOG — one data point per row (149 rows; 2026-06-12 deep dive). Sort: VERIFIED-own-page first, '
         'then community approval DPs by amount DESC, then published-requirements, then the none-found block '
         '(gray, bottom). OFF-LIST rows (institutions outside the CA roster, incl. Elan program-level rows) are '
         'labeled in the Institution cell. ⭐ GOLD rows = top verified no-doc paths. myFICO rows partly '
         'snippet-graded (forum 403s direct fetch).')
h7 = ['Institution', 'Rank', 'DP_type', 'Product', 'Amount', 'Profile', 'Docs_required', 'Bureau_pulled',
      'DP_date', 'Evidence_grade', 'Source_link', 'Notes']
ws7 = new_tab('7. NO-DOC DP LOG', note7, h7, [40, 6, 20, 36, 22, 30, 36, 18, 14, 22, 50, 50])

def sort_key7(r):
    g, t = r['Evidence_grade'], r['DP_type']
    if 'VERIFIED-own-page' in g:
        p = 0
    elif t == 'approval':
        p = 1
    elif t == 'published-requirements':
        p = 2
    else:
        p = 3
    rank = int(re.sub(r'\D', '', r['Rank']) or 9999)
    return (p, -amount_val(r['Amount']), rank)

# the 7 intended gold rows: Sierra QuickBiz, PNC BLOC soft-EQ, Patelco $40K, UMB checklist,
# Altura $25K, WF $50K BLOC DP, Stanford box
def is_star(r):
    inst, prod, t = r['Institution'].upper(), r['Product'].upper(), r['DP_type']
    if inst.startswith('OFF-LIST'):
        return False
    if 'BANK OF THE SIERRA' in inst and 'QUICKBIZ' in prod:
        return True
    if 'PNC' in inst and 'BUSINESS LINE OF CREDIT' in prod:
        return True
    if 'PATELCO' in inst and t == 'published-requirements':
        return True
    if inst.startswith('UMB ') and t == 'published-requirements':
        return True
    if 'ALTURA' in inst and t == 'published-requirements':
        return True
    if 'WELLS FARGO' in inst and 'BUSINESSLINE' in prod:
        return True
    if 'STANFORD' in inst and t == 'published-requirements':
        return True
    return False

starred = []
r7 = 3
for rr in sorted(nodoc, key=sort_key7):
    nf = rr['DP_type'] == 'none-found'
    vals = [rr[h] for h in h7]
    star = (not nf) and is_star(rr)
    if star:
        vals[3] = '⭐ ' + vals[3]
        starred.append(f"{rr['Institution']} | {rr['Product']} | {rr['Amount']}")
    set_row(ws7, r7, vals, link_col=11, link=first_url(rr['Source_link']))
    if star:
        for c in range(1, 13):
            ws7.cell(r7, c).fill = GOLD
    elif nf:
        for c in range(1, 13):
            ws7.cell(r7, c).fill = GRAY
            ws7.cell(r7, c).font = DIM
    elif 'VERIFIED-own-page' in rr['Evidence_grade']:
        for c in range(1, 13):
            ws7.cell(r7, c).fill = LGREEN
    r7 += 1
print(f'starred rows ({len(starred)}):')
for s in starred:
    print('  ⭐', s)

# ---------- TASK 5: Tab 8 PG-CLI-CFPB & FLAGS ----------
note8 = ('CFPB Credit Card Agreement DB match + PG/CLI policies (101 institutions, 2026-06-12). SCOPE CAVEAT: the '
         'CFPB DB holds CONSUMER agreements (CARD Act; <10K-account de-minimis exemption) — ABSENT never '
         'contradicts a SELF business-card call; PRESENT for an Elan agent bank is a tension, not an auto-error. '
         'Result: 43 PRESENT / 58 ABSENT; creditor-name audit of all PRESENT agreements found ZERO third-party '
         'creditor hits (4 scanned-image PDFs unauditable, flagged). Material flags list below the table.')
h8 = ['Institution', 'CFPB_status', 'Corroboration_note', 'PG_policy(+source)', 'CLI_policy(+source)',
      'Material_flags', 'Source_links', 'Date']
ws8 = new_tab('8. PG-CLI-CFPB & FLAGS', note8, h8, [34, 12, 44, 50, 50, 50, 50, 12])
r8 = 3
rank_of = {canon(r['Institution']): int(r['Rank']) for r in bureau}
for ex in sorted(extras, key=lambda r: rank_of.get(canon(r['Institution']), 9999)):
    set_row(ws8, r8, [ex[h] for h in h8], link_col=7, link=first_url(ex['Source_links']))
    if ex['Material_flags'].strip():
        ws8.cell(r8, 6).fill = AMBER
    if ex['CFPB_status'].strip().upper() == 'PRESENT':
        ws8.cell(r8, 2).fill = LGREEN
    r8 += 1
# material flags from extras_report.md
md = open(f'{BASE}/out/extras_report.md', encoding='utf-8').read()
job3 = md.split('## JOB 3')[1].split('## Method notes')[0]
flags = re.findall(r'^(\d+)\.\s+(.*?)(?=^\d+\.\s|\Z)', job3, re.M | re.S)
assert len(flags) == 14, len(flags)
r8 += 1
ws8.cell(r8, 1, 'MATERIAL FLAGS — extras_report.md JOB 3 (dated, sourced; 2026-06-12)').font = HDR
r8 += 1
for num, body in flags:
    text = ' '.join(body.split()).replace('**', '')
    url = first_url(text)
    if not url and 'Stockton' in text:  # flag 12 cites JOB 1 tension #1; carry its URL
        url = 'https://www.consumerfinance.gov/credit-cards/agreements/issuer/bank-of-stockton/'
        text += ' Source (JOB 1 tension #1): ' + url
    c1 = ws8.cell(r8, 1, f'FLAG {num}')
    c1.font = HDR
    c2 = ws8.cell(r8, 2, text)
    c2.alignment = WRAP
    if url:
        ws8.cell(r8, 7, url).hyperlink = url
    r8 += 1
flag_rows = len(flags)

# ---------- TASK 6: Tab 0 append + Tab 4 enrich ----------
ws0 = wb['0. GUIDE & LEGEND']
r0 = ws0.max_row + 2
def w0(a, b=None, bold=False):
    global r0
    c = ws0.cell(r0, 1, a)
    if bold:
        c.font = HDR
    if b:
        ws0.cell(r0, 2, b).alignment = WRAP
    r0 += 1
w0('DEEP-DIVE 2026-06-12', 'Add-only enrichment pass over v1. Nothing was deleted; every enriched cell APPENDS after the marker "‖ DD 2026-06-12:". New tabs 5-8 added; v2 CSV of Tab 1 + new NoDoc DP Log CSV exported.', bold=True)
w0('WHAT WAS ADDED', f"Tab 1: {cnt['bureau']} Bureau(quick), {cnt['nodoc']} No-doc(quick), {cnt['chex']} ChexSystems, {cnt['ews']} EWS and {cnt['flags']} Live-check-note cells enriched for the 101 deep-dived institutions. Tab 4: same bureau/no-doc data appended for its rows. NEW Tab 5 = 101-row bureau deep dive; NEW Tab 6 = 101-row Chex/EWS deep dive; NEW Tab 7 = 149-row no-doc DP log (one DP per row); NEW Tab 8 = CFPB-agreement match + PG/CLI policies + 14 material flags.")
w0('NEW EVIDENCE GRADES (Tabs 5-8)', 'VERIFIED-disclosure = institution\'s own document names the bureau/vendor. COMMUNITY-DP = individual community data point (myFICO/DoC/Reddit). CONSUMER-PROXY = consumer-card pull data used as proxy for the business product. RAIL-INFERENCE = inferred from the underwriting rail (e.g. Elan program behavior), not institution-specific evidence. UNKNOWN = searched, nothing found. DP-log grades: VERIFIED-own-page / COMMUNITY-DP / UNVERIFIED-RUMOR / none-found.')
w0('CROSS-CUTTING STRATEGY FACTS', 'Elan same-day applications = ONE combined hard pull [myFICO 6308920]; Elan rotates bureaus (no single-bureau guarantee). Pull-reuse windows exist at Logix / Kinecta / SDCCU / Stanford FCU but are NOT SAFE to rely on. U.S. Bank pulls SageStream (IDA) + ARS on top of the main bureau (freeze strategy per DoC; approves even if frozen). REMINDER: QualiFile (ChexSystems product) embeds LexisNexis — frozen LN file reads as abnormal Chex score and can AUTO-DECLINE (standing warning (b) above). myFICO caveat: several DPs read via Google snippets (forum 403s direct fetch) — snippet-graded.')
SRC_COVERAGE_ROW = r0  # placeholder, fill after coverage computed
w0('SOURCE COVERAGE', 'PLACEHOLDER')

ws4 = wb['4. BEST OF LIST']
t4b = t4n = 0
for r in range(2, ws4.max_row + 1):
    n = ws4.cell(r, 2).value
    if not n:
        continue
    cn = canon(n)
    if cn in bureau_by:
        br = bureau_by[cn]
        append_cell(ws4, r, 13, f"{DD}Bureau: {trunc(br['Bureau_finding'], 200)} [{trunc(br['Evidence_grade'], 90)}]",
                    first_url(br['Source_links']))
        t4b += 1
    nb = nodoc_best.get(cn)
    if nb:
        append_cell(ws4, r, 10, DD + dp_line(nb), first_url(nb['Source_link']))
        t4n += 1
print(f'Tab 4 enriched: bureau(Why-on-list col) {t4b}, no-doc col {t4n}')

# ---------- coverage ----------
def coverage(ws, start, end, cols):
    tot = have = 0
    for r in range(start, end):
        if all(ws.cell(r, c).value in (None, '') for c in range(1, 4)):
            continue
        tot += 1
        if any(ws.cell(r, c).hyperlink for c in cols):
            have += 1
    return have, tot

cov5 = coverage(ws5, 3, r5, [11])
cov6 = coverage(ws6, 3, r6, [9])
cov7 = coverage(ws7, 3, r7, [11])
cov8 = coverage(ws8, 3, r8, [7])
# claim-bearing for tab 7 = non-none-found rows + none-found rows that cite a URL
cb7_tot = cb7_have = 0
for rr in nodoc:
    if rr['DP_type'] != 'none-found' or 'http' in rr['Source_link']:
        cb7_tot += 1
        if 'http' in rr['Source_link']:
            cb7_have += 1
pct = lambda h, t: f'{h}/{t} ({100*h/t:.1f}%)'
cov_text = (f"Hyperlinked-source coverage — Tab 5: {pct(*cov5)} rows; Tab 6: {pct(*cov6)} rows; "
            f"Tab 7: {pct(*cov7)} of all 149 DP rows, {pct(cb7_have, cb7_tot)} of claim-bearing rows "
            f"(the remainder are none-found search-null rows with no citable page); Tab 8 table: {pct(*cov8)} rows "
            f"+ all {flag_rows} material flags sourced. Tab 1/4 enriched cells inherit or gain a hyperlink where the deep dive supplied one.")
ws0.cell(SRC_COVERAGE_ROW, 2, cov_text).alignment = WRAP
print(cov_text)

wb.save(V2)
print('saved', V2)

# ---------- TASK 7: CSV exports ----------
def render(cell):
    v = cell.value
    v = '' if v is None else str(v)
    if cell.hyperlink and cell.hyperlink.target:
        t = cell.hyperlink.target
        if v == '':
            return t
        if v != t:
            return f'{v} ({t})'
    return v

wb2 = openpyxl.load_workbook(V2)  # reload to verify post-save state
def export(ws, path, ncols):
    with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
        w = csv.writer(fh)
        nrows = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ncols):
            vals = [render(c) for c in row]
            if all(v == '' for v in vals):
                continue
            w.writerow(vals)
            nrows += 1
    print('wrote', path, nrows, 'rows')

export(wb2['1. CA INSTITUTIONS BY DEPOSITS'], f'{BASE}/deliverables/iBankNet_CA_Deposits_v2.csv', 24)
export(wb2['7. NO-DOC DP LOG'], f'{BASE}/deliverables/iBankNet_NoDoc_DP_Log_v2.csv', 12)
print('BUILD OK')
