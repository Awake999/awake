#!/usr/bin/env python3
"""IBNET-ASSEMBLER: build iBankNet_CA_Deposits_v1.xlsx + .csv from the out/ inputs."""
import csv, re, os, sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/home/user/awake/research/ibanknet/out'
DLV = '/home/user/awake/research/ibanknet/deliverables'
os.makedirs(DLV, exist_ok=True)

# ---------- helpers ----------
def norm(s):
    return re.sub(r'[^A-Z0-9]', '', s.upper())

SUFFIX = re.compile(r'\b(NATIONAL ASSOCIATION|N\.?A\.?|F\.?S\.?B\.?|FEDERAL SAVINGS BANK|FEDERAL CREDIT UNION|CREDIT UNION|FCU|CU|LTD\.?|COMPANY)\b', re.I)
def loose(s):
    s = re.sub(r'\(.*?\)', '', s.upper())
    s = s.replace('&', ' AND ')
    s = SUFFIX.sub('', s)
    return re.sub(r'[^A-Z0-9]', '', s)

def load_csv(path):
    rows = []
    with open(path, newline='') as fh:
        rd = csv.reader(fh)
        hdr = [re.sub(r'\(.*\)', '', h).strip() for h in next(rd)]
        for line in rd:
            rows.append(dict(zip(hdr, line)))
    return rows

# ---------- load roster ----------
roster = load_csv(f'{OUT}/roster_ranked.csv')
assert len(roster) == 451, len(roster)
# key: norm(name); POLAM twins disambiguated by city
roster_by_key = {}
for r in roster:
    roster_by_key.setdefault(norm(r['Institution']), []).append(r)

# ---------- load batches ----------
batch = {}   # (normname, normcity-or-'') -> rowdict with _src/_kind
batch_list = []
for i in range(1, 8):
    for f, kind in ((f'screened_batch_{i}.csv', 'screened'), (f'verified_batch_{i}.csv', 'verified')):
        for row in load_csv(f'{OUT}/{f}'):
            row['_src'] = f; row['_kind'] = kind
            batch_list.append(row)
print(f'batch rows loaded: {len(batch_list)} (screened+verified)')

# match batch -> roster
ALIAS_BATCH = {'WASHINGTONFEDERALBANKWAFD': 'WASHINGTONFEDERALBANK'}
consumed = set()           # id(roster row)
roster_data = {}           # id(roster row) -> batch row
unmatched_batch = []
for b in batch_list:
    n = norm(b['Name'])
    n = ALIAS_BATCH.get(n, n)
    if n not in roster_by_key:
        n2 = norm(re.sub(r'\(.*?\)', '', b['Name']))
        n = ALIAS_BATCH.get(n2, n2)
    cands = roster_by_key.get(n, [])
    if len(cands) > 1:  # POLAM: disambiguate by city
        cands = [c for c in cands if norm(c['City']) == norm(b['City'])]
    if len(cands) != 1:
        unmatched_batch.append(b['Name']); continue
    r = cands[0]
    if id(r) in roster_data:
        print(f'!! double batch coverage for {r["Institution"]}'); sys.exit(1)
    roster_data[id(r)] = b
    consumed.add(id(r))

# ---------- class rows ----------
class_rows = {norm(r['Name']): r for r in load_csv(f'{OUT}/class_rows.csv')}
no_coverage = []
for r in roster:
    if id(r) in roster_data:
        continue
    if norm(r['Institution']) in class_rows or r['Type'].startswith('Foreign Banking') or r['Type'] == 'Non-Depository Trust Bank':
        continue
    no_coverage.append((r['Rank'], r['Institution'], r['Type']))

# ---------- secondary bureaus ----------
sec_rows = list(csv.DictReader(open(f'{OUT}/secondary_bureaus.csv')))
ALIAS_SEC = {
    'California Bank & Trust': 'Zions Bancorporation, N.A.',   # CB&T is a Zions division
    'Chase (JPMorgan Chase)': 'JPMorgan Chase Bank, National Association',
    'Citi (Citibank)': 'Citibank, National Association',
    'Credit Union of Southern California (CU SoCal)': 'CREDIT UNION OF SOUTHERN CA, A',
    'ICBC USA, N.A.': 'Industrial and Commercial Bank of China USA, National Association',
    'First Bank (MO)': 'First Bank',
    'First Citizens Bank': 'First-Citizens Bank & Trust Company',
    'First Federal S&L of San Rafael': 'FIRST FEDERAL SAVINGS AND LOAN ASSOCIATION OF SAN RAFAEL',
    'First Tech Federal Credit Union': 'FIRST TECHNOLOGY FEDERAL CREDIT UNION',
    'Golden 1 Credit Union': 'THE GOLDEN 1 CREDIT UNION',
    'U.S. Bank (direct)': 'U.S. Bank National Association',
    'WaFd Bank': 'Washington Federal Bank',
    'Wells Fargo': 'Wells Fargo Bank, National Association',
    'Wescom Credit Union': 'WESCOM CENTRAL CREDIT UNION',
    'TIB, N.A. (The Independent BankersBank)': 'TIB National Association',
}
# DELIBERATE non-matches (name twins / off-roster entities) - never alias:
TWIN_GUARD = {'America First Credit Union'}  # Utah CU; roster AMERICAN FIRST CREDIT UNION (Brea CA) is a different institution

rloose = {}
for r in roster:
    rloose.setdefault(loose(r['Institution']), []).append(r)
rloose = {k: v[0] for k, v in rloose.items() if len(v) == 1}

sec_by_roster = {}   # id(roster row) -> sec row
sec_unmatched = []
for s in sec_rows:
    name = s['Institution']
    target = None
    if name in TWIN_GUARD:
        sec_unmatched.append((name, 'DELIBERATE twin-guard (different institution than similarly named roster row)'))
        continue
    if name in ALIAS_SEC:
        cands = roster_by_key.get(norm(ALIAS_SEC[name]), [])
        target = cands[0] if len(cands) == 1 else None
    if target is None:
        cands = roster_by_key.get(norm(name), [])
        if len(cands) == 1: target = cands[0]
    if target is None:
        target = rloose.get(loose(name))
    if target is None:
        sec_unmatched.append((name, 'no roster match (off-roster context/fintech/out-of-state row; retained in Tab 3)'))
        continue
    if id(target) in sec_by_roster:
        sec_unmatched.append((name, f'roster row already matched by another secondary row ({sec_by_roster[id(target)]["Institution"]})'))
        continue
    sec_by_roster[id(target)] = s
    s['_roster_match'] = target['Institution']

# ---------- field logic ----------
def canon_uw(raw, biz):
    u = (raw or '').strip(); up = u.upper()
    if biz == 'NO': return '—'
    if up.startswith('SELF'): return 'SELF'
    if 'ELAN' in up and not up.startswith('UNKNOWN'): return 'Elan'
    if 'TCM' in up: return 'TCM'
    if re.search(r'\bTIB\b', up): return 'TIB'
    if 'FNBO' in up: return 'FNBO'
    if 'CORSERV' in up: return 'CorServ'
    if re.search(r'\bUMB\b', up): return 'UMB'
    if 'SERVISFIRST' in up: return 'ServisFirst'
    if 'TORPAGO' in up: return 'Torpago'
    if up.startswith('UNKNOWN') or not u or up.startswith(('N/A', 'N/M', 'NONE', '(MERGED')): return 'UNKNOWN'
    return 'UNKNOWN'

def intro_star(v):
    if not v: return False
    u = v.strip().upper()
    if u.startswith(('NO', 'NOT', 'UNCLEAR', 'N/A', 'N/M', '—')): return False
    return u.startswith('YES') or '0%' in u

def terms_star(t, biz):
    if biz != 'YES' or not t: return False
    for m in re.finditer(r'\$\s?(\d{1,3})[Kk]\b|\$\s?(\d{2,3}),(\d{3})', t):
        val = int(m.group(1)) * 1000 if m.group(1) else int(m.group(2) + m.group(3))
        if val >= 20000:
            ctx = t[max(0, m.start() - 45):m.end() + 15].lower()
            if re.search(r'limit|line|to \$|up to|max', ctx): return True
    return False

NODOC_POS = re.compile(r'no.?doc|low.?doc|minimal paperwork|no tax.?returns?|without tax|soft.?pull|flex on docs|no financials|ein.?only|no tax-return requirement', re.I)
NODOC_NEG = re.compile(r'not no.?doc|full.?doc|tightens prior no.?doc', re.I)
def nodoc_star(v):
    if not v: return False
    return not NODOC_NEG.search(v) and bool(NODOC_POS.search(v))

def class_note(rtype):
    if rtype.startswith('Foreign Banking'):
        return 'No retail business cards — FFIEC-002 foreign bank office (class rule)'
    if rtype == 'Non-Depository Trust Bank':
        return 'Non-depository trust charter — no retail products (class rule)'
    return None

def sec_cell(srow, val_col, link_col):
    if srow is None: return ('UNKNOWN (no sweep row)', None)
    v = srow[val_col].strip() or 'UNKNOWN'
    l = srow[link_col].strip()
    return (v, l if l.startswith('http') else None)

# ---------- assemble tab-1 records ----------
HEAD = ['Rank', 'Institution', 'Type', 'City', 'Website', 'Deposits', 'Assets', 'Biz card?',
        'Card link', '0% intro', 'Terms', 'Underwriter', 'UW evidence', 'Apply', 'Apply link',
        'Bureau (quick)', 'No-doc (quick)', 'ChexSystems', 'EWS', 'LexisNexis/SageStream',
        'Other secondary', 'Live-check note', 'Sources', 'Screen date']

records = []   # dicts: values + flags
for r in roster:
    b = roster_data.get(id(r))
    s = sec_by_roster.get(id(r))
    rec = {'rank': int(r['Rank']), 'inst': r['Institution'], 'type': r['Type'], 'city': r['City']}
    dep = r['Deposits']
    rec['deposits'] = int(dep) if dep not in ('', 'N/A') else 'N/A'
    rec['assets'] = int(r['Assets']) if r['Assets'] else ''
    if b:
        biz = b['Biz_card'].strip().upper() or 'UNCLEAR'
        biz = {'YES': 'YES', 'NO': 'NO'}.get(biz, 'UNCLEAR')
        rec.update(website=b['Website'].strip(), biz=biz, card_link=b['Card_link'].strip(),
                   intro=b['Intro_0pct'].strip() or '—', terms=b['Terms_note'].strip(),
                   uw=canon_uw(b['Underwriter'], biz),
                   uw_ev=(b['UW_evidence'].strip() + (f" [raw: {b['Underwriter'].strip()}]" if b['Underwriter'].strip() and canon_uw(b['Underwriter'], biz) not in b['Underwriter'] else '')).strip(),
                   apply=b['Apply'].strip() or '—', apply_link=b['Apply_link'].strip(),
                   bureau_q=b['Bureau_quick'].strip() or '—', nodoc_q=b['NoDoc_quick'].strip() or '—',
                   sources=b['Source_links'].strip(), sdate=b['Screen_date'].strip())
        if b['_kind'] == 'verified':
            note = b['Agreement_vs_prior'].strip()
            tw = b.get('Twin_check', '').strip()
            if 'FAIL' in tw.upper(): note += ' | TWIN: ' + tw
            rec['live'] = note
        else:
            tw = b.get('Twin_check', '').strip()
            rec['live'] = 'Fresh screen (single pass; not in live re-verify set)' + (f' | TWIN: {tw}' if 'FAIL' in tw.upper() else '')
        rec['coverage'] = b['_kind']
    else:
        cn = class_note(r['Type'])
        if cn is None:
            cn = 'UNCLEAR-unscreened — no batch coverage and no class rule'
            biz = 'UNCLEAR'
        else:
            biz = 'NO'
        rec.update(website='', biz=biz, card_link='', intro='—', terms=cn, uw='—', uw_ev='—',
                   apply='—', apply_link='', bureau_q='—', nodoc_q='—',
                   live='Class rule applied — not individually screened',
                   sources=r['ibanknet_URL'], sdate='2026-06-10 (cache sweep)')
        rec['coverage'] = 'class-rule'
    for col, (v, l) in zip(('chex', 'ews', 'ln', 'other'),
                           (sec_cell(s, 'ChexSystems', 'Chex_link'), sec_cell(s, 'EWS', 'EWS_link'),
                            sec_cell(s, 'LexisNexis_SageStream', 'LN_link'), sec_cell(s, 'Other_secondary', 'Other_link'))):
        rec[col], rec[col + '_link'] = v, l
    rec['star_intro'] = intro_star(rec['intro'])
    rec['star_terms'] = terms_star(rec['terms'], rec['biz'])
    rec['star_nodoc'] = nodoc_star(rec['nodoc_q'])
    records.append(rec)
records.sort(key=lambda x: x['rank'])

# ---------- workbook ----------
wb = Workbook()
BLUE = Font(color='0563C1', underline='single')
HDR_FILL = PatternFill('solid', fgColor='1F4E78'); HDR_FONT = Font(color='FFFFFF', bold=True)
GREEN = PatternFill('solid', fgColor='C6EFCE'); GREEN_F = Font(color='006100', bold=True)
ORANGE = PatternFill('solid', fgColor='F4B084'); ORANGE_F = Font(color='833C00', bold=True)
AMBER = PatternFill('solid', fgColor='FFE699'); AMBER_F = Font(color='7F6000')
GOLD = PatternFill('solid', fgColor='FFD966'); GOLD_F = Font(bold=True)
DIV_FILL = PatternFill('solid', fgColor='808080'); DIV_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')
hyperlink_count = 0

def set_link(cell, url, text=None):
    global hyperlink_count
    if url and url.startswith('http'):
        cell.hyperlink = url; cell.font = BLUE; hyperlink_count += 1
    if text is not None: cell.value = text

# ===== TAB 0 =====
ws0 = wb.active; ws0.title = '0. GUIDE & LEGEND'
ws0.column_dimensions['A'].width = 34; ws0.column_dimensions['B'].width = 130
ws0.sheet_view.showGridLines = False
g = [
    ('iBankNet CALIFORNIA — BUSINESS CREDIT CARD WORKBOOK v1', ''),
    ('Built', '2026-06-12 | 451 CA institutions (ibanknet 2026-06-10 sweep) x 419 individual screens (216 fresh + 203 live-verified) x 230-row secondary-bureau sweep'),
    ('', ''),
    ('HOW TO READ', 'Tab 1 is THE list: every CA institution, rank order by CA deposits (ranks 1-422), then the labeled [N/A-DEPOSITS — FBO OFFICES] block (423-451, by office assets; FFIEC-002 offices publish no deposit figure). One row per institution. Color tells you the play at a glance; stars flag 0% intro / $20K+ stated limit / low-or-no-doc language. Tab 2 = every place the 2026-06-11/12 live re-check CONTRADICTED the prior screen (read before trusting any old note). Tab 3 = full secondary-bureau sweep dump. Tab 4 = the skim shortlist (SELF-underwritten YES cards + every starred row).'),
    ('', ''),
    ('COLOR KEY (Institution / Biz card / Underwriter cells)', ''),
    ('  GREEN', 'Card YES + SELF-underwritten — the bank itself is the credit decision; best odds of relationship/manual underwriting.'),
    ('  ORANGE-RED', 'Card YES + third-party underwriter (Elan / TCM / TIB / FNBO / CorServ / UMB / ServisFirst / Torpago) — application is decisioned by the program bank, not the institution on the row.'),
    ('  AMBER', 'Card YES with UNKNOWN underwriter, and ALL UNCLEAR rows — needs a phone call before relying on it.'),
    ('  NO FILL', 'No business credit card (incl. class-rule rows).'),
    ('  GOLD CELL (star)', 'Accent on: 0% intro cell when an intro offer exists; Terms cell when a $20K+ credit limit is stated; No-doc cell when low/no-doc language exists. Hyperlinks are blue+underlined.'),
    ('', ''),
    ('EVIDENCE GRADES (secondary-bureau cells)', 'VERIFIED-own-doc = institution\'s own disclosure names the bureau. DOC-LIST = DoctorOfCredit community-maintained list. COMMUNITY-DP = individual community data points (forums/secondary sites). Generic-CRA-clause = deposit agreement reserves the right to use "consumer reporting agencies" without naming one. UNKNOWN = no data — never guess; verify with banker at account opening.'),
    ('', ''),
    ('STANDING WARNING (a) — ROSTER DATE', 'ibanknet.com was DOWN (total origin outage, HTTP 503 on every endpoint/retry) on 2026-06-12. The roster in Tab 1 is the complete 2026-06-10 live sweep (451 rows, zero dropped), with FDIC/NCUA asset gap-fills. A live re-pull is queued for when the site recovers; two days of churn risk is near zero. Known staleness retained with notes: Comerica (merged into Fifth Third 2026-02-01), Pacific Premier (merged into Columbia Bank 2025-09-01).'),
    ('STANDING WARNING (b) — QUALIFILE / LEXISNEXIS', 'QualiFile (a ChexSystems product) EMBEDS LexisNexis data. If your LexisNexis file is FROZEN, it reads back as an abnormal/incomplete Chex score and AUTO-DECLINES at QualiFile banks (e.g., Banner). A freeze strategy is therefore BANK-SPECIFIC: thaw LN before applying anywhere flagged QualiFile/Chex in Tab 1 cols R-U.'),
    ('NOTE — cardaccount.net RAIL', 'The cardaccount.net servicing/apply rail is shared by BOTH TCM Bank N.A. AND TIB N.A. (proof: Community West Bank\'s live application names TIB N.A. verbatim on a cardaccount.net rail that v6 had attributed to TCM). NEVER infer TCM from that domain alone — only a verbatim issuer line settles it.'),
    ('', ''),
    ('UNDERWRITER COLUMN VALUES', 'SELF / Elan / TCM / TIB / FNBO / CorServ / UMB / ServisFirst / Torpago / UNKNOWN / — (no card). Raw screener wording preserved in UW evidence as [raw: ...] where it differed.'),
    ('COVERAGE', '216 institutions fresh-screened + 203 live-verified (2026-06-11/12) = 419 individual screens; 32 remaining rows (29 FBO offices + 3 non-depository trusts) carry class rules. All 3 ILCs were individually screened.'),
    ('UNITS', 'Deposits/Assets in $ thousands. SOD rows: Deposits = CA-branch deposits only (FDIC SOD 2025); Assets = total consolidated (FDIC 2026-03-31).'),
]
for i, (a, bv) in enumerate(g, 1):
    ca = ws0.cell(row=i, column=1, value=a); cb = ws0.cell(row=i, column=2, value=bv)
    ca.alignment = WRAP; cb.alignment = WRAP
    if i == 1: ca.font = Font(bold=True, size=14)
    elif a and not a.startswith('  '): ca.font = Font(bold=True)
ws0.cell(row=7, column=1).fill = GREEN; ws0.cell(row=7, column=1).font = GREEN_F
ws0.cell(row=8, column=1).fill = ORANGE; ws0.cell(row=8, column=1).font = ORANGE_F
ws0.cell(row=9, column=1).fill = AMBER; ws0.cell(row=9, column=1).font = AMBER_F
ws0.cell(row=11, column=1).fill = GOLD; ws0.cell(row=11, column=1).font = GOLD_F

# ===== TAB 1 =====
ws1 = wb.create_sheet('1. CA INSTITUTIONS BY DEPOSITS')
for c, h in enumerate(HEAD, 1):
    cell = ws1.cell(row=1, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='center')
widths = [6, 38, 26, 16, 30, 13, 13, 10, 32, 22, 48, 12, 40, 12, 32, 22, 26, 28, 24, 22, 22, 50, 40, 16]
for c, w in enumerate(widths, 1): ws1.column_dimensions[get_column_letter(c)].width = w

rowi = 2
fbo_divider_done = False
for rec in records:
    if rec['rank'] == 423 and not fbo_divider_done:
        ws1.cell(row=rowi, column=2, value='[N/A-DEPOSITS — FBO OFFICES]  ranks 423-451, sorted by office assets (FFIEC-002 offices; deposits not published)')
        for c in range(1, 25):
            ws1.cell(row=rowi, column=c).fill = DIV_FILL
        ws1.cell(row=rowi, column=2).font = DIV_FONT
        fbo_divider_done = True; rowi += 1
    vals = [rec['rank'], rec['inst'], rec['type'], rec['city'], rec['website'],
            rec['deposits'], rec['assets'], rec['biz'], rec['card_link'],
            ('⭐ ' if rec['star_intro'] else '') + rec['intro'],
            ('⭐ ' if rec['star_terms'] else '') + rec['terms'],
            rec['uw'], rec['uw_ev'], rec['apply'], rec['apply_link'],
            rec['bureau_q'], ('⭐ ' if rec['star_nodoc'] else '') + rec['nodoc_q'],
            rec['chex'], rec['ews'], rec['ln'], rec['other'],
            rec['live'], rec['sources'], rec['sdate']]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=rowi, column=c, value=v)
        if c in (2, 3, 11, 13, 16, 17, 18, 19, 20, 21, 22, 23): cell.alignment = WRAP
        if c in (6, 7) and isinstance(v, int): cell.number_format = '#,##0'
    set_link(ws1.cell(row=rowi, column=5), rec['website'])
    set_link(ws1.cell(row=rowi, column=9), rec['card_link'])
    set_link(ws1.cell(row=rowi, column=15), rec['apply_link'])
    for c, lk in ((18, rec['chex_link']), (19, rec['ews_link']), (20, rec['ln_link']), (21, rec['other_link'])):
        if lk: set_link(ws1.cell(row=rowi, column=c), lk)
    # color code (Institution=2, Biz card=8, Underwriter=12)
    fill = font = None
    if rec['biz'] == 'YES' and rec['uw'] == 'SELF': fill, font = GREEN, GREEN_F
    elif rec['biz'] == 'YES' and rec['uw'] in ('Elan', 'TCM', 'TIB', 'FNBO', 'CorServ', 'UMB', 'ServisFirst', 'Torpago'): fill, font = ORANGE, ORANGE_F
    elif (rec['biz'] == 'YES' and rec['uw'] == 'UNKNOWN') or rec['biz'] == 'UNCLEAR': fill, font = AMBER, AMBER_F
    if fill:
        for c in (2, 8, 12):
            ws1.cell(row=rowi, column=c).fill = fill
            ws1.cell(row=rowi, column=c).font = font
    if rec['star_intro']: ws1.cell(row=rowi, column=10).fill = GOLD; ws1.cell(row=rowi, column=10).font = GOLD_F
    if rec['star_terms']: ws1.cell(row=rowi, column=11).fill = GOLD; ws1.cell(row=rowi, column=11).font = GOLD_F
    if rec['star_nodoc']: ws1.cell(row=rowi, column=17).fill = GOLD; ws1.cell(row=rowi, column=17).font = GOLD_F
    rec['_xlrow'] = rowi
    rowi += 1
ws1.freeze_panes = 'C2'
ws1.auto_filter.ref = f'A1:X{rowi-1}'

# ===== TAB 2 =====
ws2 = wb.create_sheet('2. LIVE-CHECK CORRECTIONS')
H2 = ['Institution', 'Rank', 'Batch', 'Category', 'What the live check found (Agreement_vs_prior, verbatim)', 'Twin_check', 'Source links', 'Screen date']
for c, h in enumerate(H2, 1):
    cell = ws2.cell(row=1, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
for c, w in enumerate([36, 7, 18, 30, 95, 50, 50, 13], 1): ws2.column_dimensions[get_column_letter(c)].width = w

def categorize(name, txt):
    t = txt.upper()
    if 'TWIN' in t or 'WRONG ENTITY' in t or 'DIFFERENT INSTITUTION' in t.replace('A DIFFERENT', 'DIFFERENT'): return 'TWIN-FIX (wrong institution in prior)'
    if 'UNDERWRITER CHANGED' in t: return 'UNDERWRITER CORRECTED'
    if 'TIB' in t and 'TCM' in t: return 'RAIL/ISSUER CORRECTED (TIB not TCM)'
    if 'WITHDRAWN' in t or 'STRIPPED' in t: return 'PRODUCT WITHDRAWN'
    if 'CLOSED' in t and ('DEAL' in t or 'ACQUI' in t or 'GONE' in t): return 'INSTITUTION GONE (acquisition closed)'
    if 'LINEUP' in t: return 'CARD LINEUP CHANGED'
    if re.search(r'\bSUB\b|BONUS', t): return 'SIGN-UP BONUS CHANGED'
    if re.search(r'APPLY[:=]| APPLY=|PRIOR APPLY', t): return 'APPLY CHANNEL CHANGED'
    if 'NO INTRO' in t or re.search(r'(?<![\d.])0%', t): return 'INTRO/TERMS CHANGED'
    if 'DOWNGRADED' in t or 'UNCLEAR' in t: return 'STATUS DOWNGRADED'
    return 'CORRECTION'

corrections = []
for rec in records:
    b = roster_data.get(next(id(r) for r in roster if r['Institution'] == rec['inst'] and int(r['Rank']) == rec['rank']))
    if b and b['_kind'] == 'verified' and 'LIVE-CONTRADICTS' in b['Agreement_vs_prior'].upper():
        corrections.append((rec, b))
ri = 2
for rec, b in corrections:
    vals = [rec['inst'], rec['rank'], b['_src'], categorize(rec['inst'], b['Agreement_vs_prior']),
            b['Agreement_vs_prior'], b.get('Twin_check', ''), b['Source_links'], b['Screen_date']]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=c, value=v); cell.alignment = WRAP
    ws2.cell(row=ri, column=1).fill = ORANGE; ws2.cell(row=ri, column=1).font = ORANGE_F
    ri += 1
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:H{ri-1}'

# ===== TAB 3 =====
ws3 = wb.create_sheet('3. SECONDARY BUREAUS (full)')
H3 = list(sec_rows[0].keys()) + ['Roster_match']
H3 = [h for h in H3 if h != '_roster_match']
for c, h in enumerate(H3, 1):
    cell = ws3.cell(row=1, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
for c, w in enumerate([38, 30, 36, 30, 36, 28, 36, 24, 36, 60, 38], 1): ws3.column_dimensions[get_column_letter(c)].width = w
for ri, s in enumerate(sec_rows, 2):
    for c, h in enumerate(H3[:-1], 1):
        cell = ws3.cell(row=ri, column=c, value=s[h]); cell.alignment = WRAP
        if h.endswith('_link') and s[h].startswith('http'): set_link(cell, s[h])
    ws3.cell(row=ri, column=len(H3), value=s.get('_roster_match', '— (no roster match; context row)')).alignment = WRAP
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f'A1:{get_column_letter(len(H3))}{len(sec_rows)+1}'

# ===== TAB 4 =====
ws4 = wb.create_sheet('4. BEST OF LIST')
H4 = ['Rank', 'Institution', 'Type', 'City', 'Deposits', 'Biz card?', 'Underwriter', '0% intro', 'Terms', 'No-doc (quick)', 'Apply', 'Card link', 'Why on list']
for c, h in enumerate(H4, 1):
    cell = ws4.cell(row=1, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT
for c, w in enumerate([6, 38, 24, 16, 13, 10, 12, 26, 60, 30, 12, 36, 34], 1): ws4.column_dimensions[get_column_letter(c)].width = w
ri = 2
best_count = 0
for rec in records:
    why = []
    if rec['biz'] == 'YES' and rec['uw'] == 'SELF': why.append('SELF-underwritten')
    if rec['star_intro']: why.append('0% intro')
    if rec['star_terms']: why.append('$20K+ stated limit')
    if rec['star_nodoc']: why.append('low/no-doc language')
    if not why: continue
    best_count += 1
    vals = [rec['rank'], rec['inst'], rec['type'], rec['city'], rec['deposits'], rec['biz'], rec['uw'],
            ('⭐ ' if rec['star_intro'] else '') + rec['intro'],
            ('⭐ ' if rec['star_terms'] else '') + rec['terms'],
            ('⭐ ' if rec['star_nodoc'] else '') + rec['nodoc_q'],
            rec['apply'], rec['card_link'], ' + '.join(why)]
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(row=ri, column=c, value=v)
        if c in (2, 8, 9, 10, 13): cell.alignment = WRAP
        if c == 5 and isinstance(v, int): cell.number_format = '#,##0'
    set_link(ws4.cell(row=ri, column=12), rec['card_link'])
    fill, font = (GREEN, GREEN_F) if (rec['biz'] == 'YES' and rec['uw'] == 'SELF') else \
                 ((ORANGE, ORANGE_F) if rec['uw'] in ('Elan', 'TCM', 'TIB', 'FNBO', 'CorServ', 'UMB', 'ServisFirst', 'Torpago') and rec['biz'] == 'YES' else (AMBER, AMBER_F))
    ws4.cell(row=ri, column=2).fill = fill; ws4.cell(row=ri, column=2).font = font
    ws4.cell(row=ri, column=7).fill = fill; ws4.cell(row=ri, column=7).font = font
    ri += 1
ws4.freeze_panes = 'C2'
ws4.auto_filter.ref = f'A1:M{ri-1}'

xlsx_path = f'{DLV}/iBankNet_CA_Deposits_v1.xlsx'
wb.save(xlsx_path)

# ---------- CSV export (tab 1 full) ----------
csv_path = f'{DLV}/iBankNet_CA_Deposits_v1.csv'
with open(csv_path, 'w', newline='', encoding='utf-8-sig') as fh:
    w = csv.writer(fh)
    w.writerow(HEAD)
    for rec in records:
        def lc(v, l): return f'{v} ({l})' if l else v
        w.writerow([rec['rank'], rec['inst'], rec['type'], rec['city'], rec['website'],
                    rec['deposits'], rec['assets'], rec['biz'], rec['card_link'],
                    ('⭐ ' if rec['star_intro'] else '') + rec['intro'],
                    ('⭐ ' if rec['star_terms'] else '') + rec['terms'],
                    rec['uw'], rec['uw_ev'], rec['apply'], rec['apply_link'],
                    rec['bureau_q'], ('⭐ ' if rec['star_nodoc'] else '') + rec['nodoc_q'],
                    lc(rec['chex'], rec['chex_link']), lc(rec['ews'], rec['ews_link']),
                    lc(rec['ln'], rec['ln_link']), lc(rec['other'], rec['other_link']),
                    rec['live'], rec['sources'], rec['sdate']])

# ---------- SELF-CHECK ----------
print('=' * 70)
print('SELF-CHECK')
print('=' * 70)
print(f'Roster rows in tab 1            : {len(records)} / 451 (each exactly once: {len(records) == len(roster) == 451})')
print(f'Batch rows loaded               : {len(batch_list)} (216 screened + 203 verified = {len(batch_list) == 419})')
print(f'Batch rows matched to roster    : {len(roster_data)}   unmatched batch rows: {len(unmatched_batch)} {unmatched_batch}')
cov = {'screened': 0, 'verified': 0, 'class-rule': 0}
for rec in records: cov[rec['coverage']] += 1
print(f'Coverage                        : {cov}')
print(f'Roster rows w/o batch or class  : {len(no_coverage)} {no_coverage}')
ilcs = [(rec['inst'], rec['coverage']) for rec in records if 'Industrial Loan' in rec['type']]
print(f'ILC coverage                    : {ilcs}')
yes = sum(1 for r in records if r['biz'] == 'YES')
unclear = sum(1 for r in records if r['biz'] == 'UNCLEAR')
uwc = {}
for r in records:
    if r['biz'] == 'YES': uwc[r['uw']] = uwc.get(r['uw'], 0) + 1
print(f'Biz card YES                    : {yes}   NO: {sum(1 for r in records if r["biz"]=="NO")}   UNCLEAR: {unclear}')
print(f'Underwriters (YES rows)         : {dict(sorted(uwc.items(), key=lambda x: -x[1]))}')
print(f'0% intro stars                  : {sum(1 for r in records if r["star_intro"])}')
print(f'$20K+ limit stars               : {sum(1 for r in records if r["star_terms"])}')
print(f'No/low-doc stars                : {sum(1 for r in records if r["star_nodoc"])}')
known = lambda col: sum(1 for r in records if not r[col].upper().startswith(('UNKNOWN', '—', 'N/M')))
print(f'Secondary known (Chex/EWS/LN/Other): {known("chex")}/{known("ews")}/{known("ln")}/{known("other")} of 451')
print(f'Secondary rows matched to roster: {len(sec_by_roster)} / {len(sec_rows)}')
print(f'Secondary rows UNMATCHED ({len(sec_unmatched)}):')
for n, why in sec_unmatched: print(f'   - {n}: {why}')
print(f'Live-check corrections (tab 2)  : {len(corrections)}')
for rec, b in corrections: print(f'   - {rec["inst"]} [{categorize(rec["inst"], b["Agreement_vs_prior"])}]')
print(f'Best-of rows (tab 4)            : {best_count}')
print(f'Hyperlink cells                 : {hyperlink_count}')
print(f'Files: {xlsx_path}  |  {csv_path}')

# spot-check 8 rows verbatim
print('\nSPOT-CHECK (workbook vs source CSV, verbatim):')
import random
random.seed(7)
spot_names = ['Bank of America, National Association', 'SESLOC CREDIT UNION', 'MISSION FEDERAL CREDIT UNION',
              'PATELCO CREDIT UNION', 'HCN BANK', 'WOORI BK LA BR',
              'BLACKROCK INSTITUTIONAL TRUST COMPANY, NATIONAL ASSOCIATION', 'POLAM FEDERAL CREDIT UNION']
for nm in spot_names:
    matches = [rec for rec in records if rec['inst'] == nm]
    for rec in matches:
        rrow = next(r for r in roster if r['Institution'] == nm and int(r['Rank']) == rec['rank'])
        b = roster_data.get(id(rrow))
        if b:
            ok = (rec['biz'] == ({'YES': 'YES', 'NO': 'NO'}.get(b['Biz_card'].strip().upper() or 'UNCLEAR', 'UNCLEAR'))
                  and rec['terms'] == b['Terms_note'].strip() and rec['sdate'] == b['Screen_date'].strip()
                  and rec['card_link'] == b['Card_link'].strip())
            print(f'  [{ "OK" if ok else "MISMATCH" }] {nm} (rank {rec["rank"]}, {b["_src"]}): biz={rec["biz"]} uw={rec["uw"]} | terms[:60]={rec["terms"][:60]!r}')
        else:
            print(f'  [OK-class] {nm} (rank {rec["rank"]}): {rec["terms"]}')
