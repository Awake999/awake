#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the SoCal Self-UW Tiered (no-Chex) workbook from build_tiered records."""
import build_tiered as bt
import csv, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DELV = bt.DELV
OUTX = DELV + '/SoCal_SelfUW_Tiered_NoChex_v1.xlsx'
OUTC = DELV + '/SoCal_SelfUW_Tiered_NoChex_v1.csv'
BUILD_DATE = '2026-06-12'

# ---- colors ----
T1_FILL = PatternFill('solid', fgColor='C6EFCE')   # green
T2_FILL = PatternFill('solid', fgColor='FFEB9C')   # yellow
T3_FILL = PatternFill('solid', fgColor='FFC7CE')   # red
GREEN   = PatternFill('solid', fgColor='C6EFCE')
RED     = PatternFill('solid', fgColor='FFC7CE')
GRAY    = PatternFill('solid', fgColor='D9D9D9')
GOLD    = PatternFill('solid', fgColor='FFE699')   # star/no-doc
HDR     = PatternFill('solid', fgColor='1F4E78')   # dark blue band
T1_BAND = PatternFill('solid', fgColor='548235')   # dark green band
T2_BAND = PatternFill('solid', fgColor='BF8F00')   # dark gold band
T3_BAND = PatternFill('solid', fgColor='C00000')   # dark red band
CTY_BAND= PatternFill('solid', fgColor='305496')
LINK    = Font(color='0563C1', underline='single')
WHITE_B = Font(color='FFFFFF', bold=True, size=12)
HDR_F   = Font(color='FFFFFF', bold=True, size=10)
BOLD    = Font(bold=True)
THIN    = Side(style='thin', color='BFBFBF')
BORDER  = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WRAP    = Alignment(vertical='top', wrap_text=True)
CTR     = Alignment(horizontal='center', vertical='center', wrap_text=True)

def setcell(ws, r, c, val, font=None, fill=None, align=WRAP, link=None, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    if link:
        cell.hyperlink = link
        cell.font = font or LINK
    elif font:
        cell.font = font
    if fill: cell.fill = fill
    cell.alignment = align
    if border: cell.border = BORDER
    return cell

def first_url(s):
    m = re.search(r'https?://[^\s;,)]+', s or '')
    return m.group(0) if m else None

def chex_disp(status, grade):
    em = {'NO':'NO','YES':'YES','UNKNOWN':'UNK'}.get(status, status)
    return f"{em} ({grade})" if grade and grade!='UNKNOWN' else em

def cell_fill_for(status):
    if status=='NO': return GREEN
    if status=='YES': return RED
    return GRAY

wb = Workbook()

# ===================================================================
# TAB 0 — GUIDE & LEGEND
# ===================================================================
ws = wb.active
ws.title = '0. GUIDE & LEGEND'
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 120
def gline(r, text, font=None, fill=None):
    setcell(ws, r, 2, text, font=font or Font(size=10), fill=fill,
            align=Alignment(vertical='top', wrap_text=True), border=False)
gr = 1
setcell(ws, gr,2,'SoCal Self-Underwritten, Tiered by No-Chex / No-EWS funding  (v1)', Font(size=16,bold=True,color='1F4E78'), border=False); gr+=2
gline(gr,'HOW TO READ: The MAIN list (Tab 1) contains ONLY institutions that are SELF-UNDERWRITTEN + offer a BUSINESS CREDIT CARD + are SoCal-CONFIRMED. Within that set, the tier is set purely by ChexSystems/EWS screening behavior. BEST = the no-Chex / no-EWS funding tier.', Font(size=10,bold=True)); gr+=2
setcell(ws,gr,2,'TIER DEFINITION (the spine — applied exactly)', Font(size=12,bold=True,color='1F4E78'), border=False); gr+=1
for txt,fill in [
 ('🟩 TIER 1 (green, BEST): Chex = NO (verified/own-doc/strong) AND EWS = NO AND no-doc = YES/partial. (If no-doc is unknown but Chex+EWS both clean, still Tier 1 but mark the no-doc gap.)', T1_FILL),
 ('🟨 TIER 2 (yellow, UNDETERMINED): self-UW + card + SoCal, but ChexSystems OR EWS is UNKNOWN (not a known puller).', T2_FILL),
 ('🟥 TIER 3 (red, SCREENS): pulls ChexSystems AND/OR EWS (confirmed). If a cell is unknown, it is Tier 2, never green.', T3_FILL)]:
    gline(gr,txt,Font(size=10),fill); gr+=1
gr+=1
setcell(ws,gr,2,'COLOR KEY', Font(size=12,bold=True,color='1F4E78'), border=False); gr+=1
for txt,fill in [('🟩 Tier 1 row / NO-screen cell (green)',GREEN),('🟨 Tier 2 row / UNKNOWN cell (yellow / gray)',T2_FILL),
                 ('🟥 Tier 3 row / YES-screen cell (red)',RED),('⭐ No-doc = YES (gold star cell)',GOLD),
                 ('💙 Blue underlined text = hyperlink to website / source',None)]:
    c=gline(gr,txt,Font(size=10),fill)
    if 'hyperlink' in txt: ws.cell(row=gr,column=2).font=LINK
    gr+=1
gr+=1
setcell(ws,gr,2,'EVIDENCE-GRADE DEFINITIONS', Font(size=12,bold=True,color='1F4E78'), border=False); gr+=1
for txt in ['VERIFIED-own-doc : read from the institution’s own deposit/card agreement or disclosure (absence or presence of Chex/EWS language).',
 'COMMUNITY-DP : data point(s) from DoctorOfCredit / myFICO / CreditBoards community reports.',
 'CONSUMER-PROXY : inferred from consumer-account behavior where business-specific evidence is absent.',
 'RAIL-INFERENCE : inferred from the card-issuing rail / processor, not the institution directly.',
 'UNKNOWN : no reliable signal located — defaults the row to Tier 2 (never green).']:
    gline(gr,'• '+txt,Font(size=10)); gr+=1
gr+=1
setcell(ws,gr,2,'SOCAL COUNTY DEFINITION', Font(size=12,bold=True,color='1F4E78'), border=False); gr+=1
gline(gr,'SoCal = YES if the institution has ≥1 confirmed branch (FDIC SOD / NCUA census) in LA, San Bernardino, Riverside, Orange, Ventura, or San Diego — OR its HQ city sits in a SoCal county.',Font(size=10)); gr+=1
gline(gr,'PRIMARY counties: Los Angeles, Orange, San Bernardino, Riverside.   SECONDARY counties: Ventura, San Diego, Santa Barbara, Imperial.',Font(size=10)); gr+=2
setcell(ws,gr,2,'HONESTY NOTE', Font(size=12,bold=True,color='C00000'), border=False); gr+=1
gline(gr,'Tier 1 is intentionally sparse — almost no traditional SoCal self-UW bank is verified no-Chex AND no-EWS. The only Tier-1 candidates are own-document “likely-no” cases: City National Bank (own docs name no Chex/EWS; offers business cards), Beneficial State Bank (both agreements zero Chex/EWS), and Commercial Bank of California (own business sections have no consumer-report clause — but it has NO issued business-card product, so it is a DEPOSIT-ONLY no-Chex/EWS option, flagged in Tab 1). The verified-clean banking layer is fintech (Mercury / Bluevine / Found / Novo / Lili / Relay), listed in Tab 4 — pair one of those as your banking layer with a Tier-1/2 card.',Font(size=10,color='C00000')); gr+=2
gline(gr,f'Build date {BUILD_DATE}.  Sources: iBankNet_CA_Deposits_v2, FDIC SOD 6-county census 2025, NCUA 6-county census, deep_chex/secondary_bureaus, deep+rerun bureau, nodoc_dps, chex_socal_cus.',Font(size=9,italic=True)); gr+=1

# ===================================================================
# TAB 1 — MAIN TIERED
# ===================================================================
def add_main(ws, recs):
    cols = ['Tier','Institution','SoCal?','County','Cities','Self-UW','Biz card','Card link',
            'No-doc (+grade)','ChexSystems','EWS','Bureau (hard pull)','Est. limit','Apply','Website','Sources']
    widths=[8,30,8,22,26,9,9,14,26,22,22,26,12,12,30,30]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    # header
    for i,h in enumerate(cols,1):
        setcell(ws,1,i,h,HDR_F,HDR,CTR)
    r=2
    tier_meta={1:('🟩 TIER 1 — BEST (no Chex / no EWS, verified-absence own-doc)',T1_BAND,T1_FILL),
               2:('🟨 TIER 2 — UNDETERMINED (Chex or EWS unknown)',T2_BAND,T2_FILL),
               3:('🟥 TIER 3 — SCREENS (pulls ChexSystems and/or EWS)',T3_BAND,T3_FILL)}
    def sortkey(x):
        prim = 0 if (set(x['counties']) & bt.PRIMARY) else 1
        dep = -bt.n_int(x['deposits'])
        return (prim, dep)
    for tier in (1,2,3):
        grp=sorted([x for x in recs if x['tier']==tier], key=sortkey)
        if not grp: continue
        label,band,rowfill=tier_meta[tier]
        setcell(ws,r,1,label,WHITE_B,band,Alignment(horizontal='left',vertical='center'))
        for c in range(2,len(cols)+1):
            ws.cell(row=r,column=c).fill=band
        ws.row_dimensions[r].height=20
        r+=1
        for x in grp:
            tname={1:'🟩 1',2:'🟨 2',3:'🟥 3'}[tier]
            setcell(ws,r,1,tname,BOLD,rowfill,CTR)
            setcell(ws,r,2,x['institution'],Font(bold=True,color='0563C1',underline='single') if x['website'].startswith('http') else BOLD, rowfill, WRAP, link=x['website'] if x['website'].startswith('http') else None)
            setcell(ws,r,3,'YES',Font(bold=True,color='548235'),GREEN,CTR)
            setcell(ws,r,4,'; '.join(x['counties']) or '—',None,None,WRAP)
            setcell(ws,r,5,', '.join(x['cities']) or '—',None,None,WRAP)
            setcell(ws,r,6,x['self_uw'],None,None,CTR)
            setcell(ws,r,7,x['card'],None,None,CTR)
            cl=x['card_link']
            setcell(ws,r,8,'link' if cl.startswith('http') else '—',None,None,CTR, link=cl if cl.startswith('http') else None)
            # no-doc
            ndtxt=f"{x['nodoc']} ({x['nodoc_grade']})"
            ndfill=GOLD if x['nodoc']=='YES' else None
            ndlink=first_url(x['nodoc_src']) or first_url(x['nodoc_note'])
            setcell(ws,r,9,('⭐ ' if x['nodoc']=='YES' else '')+ndtxt,None,ndfill,WRAP,link=ndlink)
            # chex
            setcell(ws,r,10,chex_disp(x['chex'],x['chex_grade']),None,cell_fill_for(x['chex']),WRAP,link=first_url(x['chex_src'])or first_url(x['chex_note']))
            setcell(ws,r,11,chex_disp(x['ews'],x['ews_grade']),None,cell_fill_for(x['ews']),WRAP,link=first_url(x['ews_note']))
            # bureau
            btxt=(x['bureau'][:60] or 'unknown')+(f" ({x['bureau_grade']})" if x['bureau_grade'] else '')
            setcell(ws,r,12,btxt,None,None,WRAP,link=first_url(x['bureau_src']))
            setcell(ws,r,13,x['est_limit'] or '—',None,None,CTR)
            setcell(ws,r,14,x['apply'] or '—',None,None,WRAP,link=x['apply_link'] if x['apply_link'].startswith('http') else None)
            setcell(ws,r,15,x['website'],LINK if x['website'].startswith('http') else None,None,WRAP,link=x['website'] if x['website'].startswith('http') else None)
            srctxt = first_url(x['sources']) or (x['sources'][:50] if x['sources'] else '—')
            if x.get('note'):
                srctxt = 'NOTE: '+x['note']+('  | '+(first_url(x['sources']) or '') if first_url(x['sources']) else '')
            setcell(ws,r,16,srctxt,None,GOLD if x.get('note') else None,WRAP,link=first_url(x['sources']))
            ws.row_dimensions[r].height=42
            r+=1
        r+=1
    ws.freeze_panes='A2'
    ws.auto_filter.ref=f'A1:{get_column_letter(len(cols))}1'

ws1 = wb.create_sheet('1. SOCAL SELF-UW TIERED')
add_main(ws1, bt.socal_recs)

# ===================================================================
# TAB 2 — BY COUNTY  (Tier 1 & 2 regrouped)
# ===================================================================
ws2 = wb.create_sheet('2. BY COUNTY')
cols2=['Institution','Tier','Self-UW','Card','Chex','EWS','No-doc','Website']
w2=[34,8,9,8,16,16,16,34]
for i,w in enumerate(w2,1): ws2.column_dimensions[get_column_letter(i)].width=w
COUNTY_ORDER=['Los Angeles','Orange','San Bernardino','Riverside','Ventura','San Diego']
r=1
t12=[x for x in bt.socal_recs if x['tier'] in (1,2)]
for cty in COUNTY_ORDER:
    grp=[x for x in t12 if cty in x['counties']]
    if not grp: continue
    grp.sort(key=lambda x:(x['tier'], -bt.n_int(x['deposits'])))
    setcell(ws2,r,1,cty.upper(),WHITE_B,CTY_BAND,Alignment(horizontal='left',vertical='center'))
    for c in range(2,len(cols2)+1): ws2.cell(row=r,column=c).fill=CTY_BAND
    ws2.row_dimensions[r].height=18; r+=1
    for i,h in enumerate(cols2,1): setcell(ws2,r,i,h,HDR_F,HDR,CTR)
    r+=1
    for x in grp:
        rowfill={1:T1_FILL,2:T2_FILL}[x['tier']]
        setcell(ws2,r,1,x['institution'],Font(bold=True,color='0563C1',underline='single') if x['website'].startswith('http') else BOLD,rowfill,WRAP,link=x['website'] if x['website'].startswith('http') else None)
        setcell(ws2,r,2,{1:'🟩 1',2:'🟨 2'}[x['tier']],BOLD,rowfill,CTR)
        setcell(ws2,r,3,x['self_uw'],None,None,CTR)
        setcell(ws2,r,4,x['card'],None,None,CTR)
        setcell(ws2,r,5,chex_disp(x['chex'],x['chex_grade']),None,cell_fill_for(x['chex']),WRAP)
        setcell(ws2,r,6,chex_disp(x['ews'],x['ews_grade']),None,cell_fill_for(x['ews']),WRAP)
        setcell(ws2,r,7,('⭐ ' if x['nodoc']=='YES' else '')+x['nodoc'],None,GOLD if x['nodoc']=='YES' else None,CTR)
        setcell(ws2,r,8,x['website'],LINK if x['website'].startswith('http') else None,None,WRAP,link=x['website'] if x['website'].startswith('http') else None)
        ws2.row_dimensions[r].height=30; r+=1
    r+=1
ws2.freeze_panes='A1'

# ===================================================================
# TAB 3 — NOT SOCAL-CONFIRMED
# ===================================================================
ws3 = wb.create_sheet('3. SELF-UW NOT SOCAL')
cols3=['Institution','Why here','HQ City','HQ State','Self-UW','Card','Chex','EWS','No-doc','Website']
w3=[32,30,16,8,9,8,16,16,14,32]
for i,w in enumerate(w3,1): ws3.column_dimensions[get_column_letter(i)].width=w
setcell(ws3,1,1,'Self-underwritten + business card, but NO confirmed SoCal branch (reference only — not in the main tiered list).',Font(bold=True,color='C00000'),None,Alignment(wrap_text=True))
ws3.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols3))
for i,h in enumerate(cols3,1): setcell(ws3,2,i,h,HDR_F,HDR,CTR)
OOS_STATE={'providence':'RI','kansas city':'MO','new york':'NY','hagatna':'GU'}
NORCAL={'santa rosa','chico','dublin','fremont','vacaville','roseville','folsom',
        'palo alto','bakersfield','san jose','eureka','redwood city','san francisco',
        'livermore','woodland','sacramento','san luis obispo','clayton','st. louis'}
r=3
for x in sorted(bt.notsocal_recs, key=lambda x:-bt.n_int(x['deposits'])):
    hqc_l=(x['hq_city'] or x['city']).lower()
    st = x['hq_state'] or OOS_STATE.get(hqc_l,'') or ('CA' if hqc_l in NORCAL else '')
    x['hq_state']=st
    if x['online_national']:
        why='Online-national (no CA branches)'
    elif st and st!='CA':
        why=f"Out-of-state HQ ({x['hq_city']}, {st}) — no confirmed SoCal branch in census"
    elif hqc_l in NORCAL:
        why=f"NorCal HQ ({x['hq_city']}) — no confirmed SoCal branch in census"
    else:
        why=f"HQ {x['hq_city']} — no confirmed SoCal branch in census"
    setcell(ws3,r,1,x['institution'],Font(bold=True,color='0563C1',underline='single') if x['website'].startswith('http') else BOLD,None,WRAP,link=x['website'] if x['website'].startswith('http') else None)
    setcell(ws3,r,2,why,None,None,WRAP)
    setcell(ws3,r,3,x['hq_city'] or x['city'],None,None,CTR)
    setcell(ws3,r,4,x['hq_state'] or '—',None,None,CTR)
    setcell(ws3,r,5,x['self_uw'],None,None,CTR)
    setcell(ws3,r,6,x['card'],None,None,CTR)
    setcell(ws3,r,7,chex_disp(x['chex'],x['chex_grade']),None,cell_fill_for(x['chex']),WRAP)
    setcell(ws3,r,8,chex_disp(x['ews'],x['ews_grade']),None,cell_fill_for(x['ews']),WRAP)
    setcell(ws3,r,9,('⭐ ' if x['nodoc']=='YES' else '')+x['nodoc'],None,GOLD if x['nodoc']=='YES' else None,CTR)
    setcell(ws3,r,10,x['website'],LINK if x['website'].startswith('http') else None,None,WRAP,link=x['website'] if x['website'].startswith('http') else None)
    ws3.row_dimensions[r].height=30; r+=1
ws3.freeze_panes='A3'

# ===================================================================
# TAB 4 — VERIFIED NO-CHEX BANKING LAYER (fintechs)
# ===================================================================
ws4 = wb.create_sheet('4. NO-CHEX BANKING LAYER')
fintechs = [r for r in bt.rd(bt.V7+'/chex_socal_cus.csv') if 'fintech' in r['Institution'].lower()]
cols4=['Fintech','Partner bank(s)','Chex','EWS','How to open','Verdict','Sources']
w4=[20,34,12,12,40,30,34]
for i,w in enumerate(w4,1): ws4.column_dimensions[get_column_letter(i)].width=w
setcell(ws4,1,1,'VERIFIED no-Chex banking layer — the ONLY verified no-Chex options. These are the deposit/banking layer; pair one with a Tier-1/2 business card from Tab 1.',Font(bold=True,color='548235'),None,Alignment(wrap_text=True))
ws4.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols4))
for i,h in enumerate(cols4,1): setcell(ws4,2,i,h,HDR_F,HDR,CTR)
r=3
for f in fintechs:
    nm=f['Institution'].replace(' (fintech)','')
    src=first_url(f.get('Source_links',''))
    setcell(ws4,r,1,nm,BOLD,GREEN,CTR)
    setcell(ws4,r,2,f.get('Charter',''),None,None,WRAP)
    cs,_=bt.classify_chex(f.get('Chex_membership',''))
    es,_=bt.classify_chex(f.get('EWS_membership',''))
    setcell(ws4,r,3,'NO' if 'no' in f.get('Chex_membership','').lower()[:12] else cs,None,GREEN,CTR)
    setcell(ws4,r,4,'NO' if 'no' in f.get('EWS_membership','').lower()[:12] else es,None,GREEN if 'no' in f.get('EWS_membership','').lower()[:12] else GRAY,CTR)
    setcell(ws4,r,5,f.get('Join_path_from_Fontana_LB','')[:160],None,None,WRAP)
    setcell(ws4,r,6,f.get('Verdict_tier',''),None,None,WRAP)
    setcell(ws4,r,7,src or '—',None,None,WRAP,link=src)
    ws4.row_dimensions[r].height=46; r+=1
ws4.freeze_panes='A3'

# ===================================================================
# TAB 5 — CHANGE / COHERENCE LOG
# ===================================================================
ws5 = wb.create_sheet('5. CHANGE-COHERENCE LOG')
ws5.column_dimensions['A'].width=4
ws5.column_dimensions['B'].width=120
def l5(r,t,font=None,fill=None):
    setcell(ws5,r,2,t,font or Font(size=10),fill,Alignment(vertical='top',wrap_text=True),border=False)
r=1
setcell(ws5,r,2,'CHANGE / COHERENCE LOG — what this rebuild fixed', Font(size=14,bold=True,color='1F4E78'),border=False); r+=2
from collections import Counter
tc=Counter(x['tier'] for x in bt.socal_recs)
ctyc=Counter()
for x in bt.socal_recs:
    if x['tier'] in (1,2):
        for c in x['counties']: ctyc[c]+=1
defects=[
 ('D1','OLD: rows colored green by “self-underwriter + has card.”  FIX: green (Tier 1) now means verified no-Chex AND no-EWS only. Having a card no longer implies “best.”'),
 ('D2','OLD: Chex/EWS pullers could appear in the “good” color. FIX: any confirmed Chex/EWS pull → Tier 3 (red). Zero green rows carry Chex=YES or EWS=YES (asserted in self-check).'),
 ('D3','OLD: unknowns were treated optimistically. FIX: any UNKNOWN in Chex or EWS → Tier 2 (yellow), never green.'),
 ('D4','OLD: no SoCal gating — online-national / NorCal banks sat in the main list. FIX: main tab requires a SoCal-confirmed branch (FDIC SOD / NCUA census) or SoCal HQ; the rest moved to Tab 3.'),
 ('D5','OLD: no own-document layer for the genuinely-clean banks. FIX: Commercial Bank of California, Beneficial State Bank, City National graded NO from their own deposit/card agreements (verified-absence).'),
 ('D6','OLD: verified no-Chex fintech layer absent. FIX: Mercury/Bluevine/Found/Novo/Lili/Relay surfaced in Tab 4 as the only VERIFIED no-Chex banking layer.'),
]
for code,txt in defects:
    setcell(ws5,r,2,f'{code}: {txt}',Font(size=10),None,Alignment(vertical='top',wrap_text=True),border=False); r+=1
r+=1
l5(r,'TIER COUNTS (SoCal main list): '+f"Tier 1 = {tc.get(1,0)}, Tier 2 = {tc.get(2,0)}, Tier 3 = {tc.get(3,0)}.  Not-SoCal (Tab 3) = {len(bt.notsocal_recs)}.",Font(size=10,bold=True)); r+=1
l5(r,'COUNTY PRESENCE (Tier 1+2): '+', '.join(f'{k}={v}' for k,v in sorted(ctyc.items(), key=lambda x:-x[1])),Font(size=10)); r+=1
t1names=', '.join(x['institution'] for x in bt.socal_recs if x['tier']==1)
l5(r,'TIER-1 MEMBERS: '+t1names,Font(size=10,bold=True,color='548235')); r+=2
l5(r,'DATA SOURCES: iBankNet_CA_Deposits_v2.csv (451 rows); FDIC SOD 6-county census 2025; NCUA 6-county census; secondary_bureaus.csv + deep_chex_results_1/2; deep_bureau_results_1-3 + rerun_bureau_results_1-5 (Upgrade=YES preferred); nodoc_dps_1/2; chex_socal_cus.csv (fintech layer); geo_longbeach/geo_fontana (branch cities).',Font(size=9)); r+=1
l5(r,f'BUILD DATE: {BUILD_DATE}.',Font(size=9,italic=True)); r+=1

wb.save(OUTX)
print('wrote', OUTX)

# ===================================================================
# CSV export of Tab 1  ('text (URL)' pattern)
# ===================================================================
def tu(text, url):
    text=(text or '').strip()
    if url and url.startswith('http'):
        return f'{text} ({url})' if text else url
    return text

with open(OUTC,'w',newline='',encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['Tier','Institution','SoCal?','County','Cities','Self-UW','Biz card',
                'Card link','No-doc (+grade)','ChexSystems','EWS','Bureau (hard pull)',
                'Est. limit','Apply','Website','Sources'])
    def sk(x):
        return (0 if set(x['counties'])&bt.PRIMARY else 1, -bt.n_int(x['deposits']))
    for tier in (1,2,3):
        for x in sorted([y for y in bt.socal_recs if y['tier']==tier], key=sk):
            w.writerow([
                {1:'TIER 1',2:'TIER 2',3:'TIER 3'}[tier],
                tu(x['institution'], x['website']),
                'YES',
                '; '.join(x['counties']),
                ', '.join(x['cities']),
                x['self_uw'], x['card'],
                tu('link', x['card_link']),
                tu(('⭐ ' if x['nodoc']=='YES' else '')+f"{x['nodoc']} ({x['nodoc_grade']})", first_url(x['nodoc_src'])),
                chex_disp(x['chex'],x['chex_grade']),
                chex_disp(x['ews'],x['ews_grade']),
                tu((x['bureau'][:80]+(f" ({x['bureau_grade']})" if x['bureau_grade'] else '')), first_url(x['bureau_src'])),
                x['est_limit'],
                tu(x['apply'], x['apply_link']),
                x['website'],
                ('NOTE: '+x['note']+' | '+(first_url(x['sources']) or '')) if x.get('note') else (first_url(x['sources']) or ''),
            ])
print('wrote', OUTC)
