#!/usr/bin/env python3
"""Build ca_business_credit_funding_v7.xlsx — ADD-ONLY on top of v6."""
import csv, os, shutil, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

V6 = '/home/user/awake/research/v6/deliverables/ca_business_credit_funding_v6.xlsx'
OUT_DIR = '/home/user/awake/research/v7/deliverables'
V7 = os.path.join(OUT_DIR, 'ca_business_credit_funding_v7.xlsx')
V7OUT = '/home/user/awake/research/v7/out'
TODAY = '2026-06-11'
MARK = f' ‖ v7 {TODAY}: '   # ' ‖ v7 2026-06-11: '
os.makedirs(OUT_DIR, exist_ok=True)
shutil.copy(V6, V7)
wb = openpyxl.load_workbook(V7)

# ---------- styles ----------
LINKFONT = Font(color='0563C1', underline='single', size=10)
HDRFONT = Font(bold=True, color='FFFFFF', size=10)
HDRFILL = PatternFill('solid', fgColor='1F4E78')
GREEN = PatternFill('solid', fgColor='C6EFCE')
AMBER = PatternFill('solid', fgColor='FFE699')
RED = PatternFill('solid', fgColor='FFC7CE')
BLUEHL = PatternFill('solid', fgColor='DDEBF7')
SMALL = Font(size=10)
BOLD = Font(bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical='top')

def link(ws, row, col, text, url, fill=None):
    c = ws.cell(row, col, text)
    if url:
        c.hyperlink = url
        c.font = LINKFONT
    else:
        c.font = SMALL
    c.alignment = WRAP
    if fill: c.fill = fill
    return c

def put(ws, row, col, text, fill=None, font=None):
    c = ws.cell(row, col, text)
    c.font = font or SMALL
    c.alignment = WRAP
    if fill: c.fill = fill
    return c

def header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h); c.font = HDRFONT; c.fill = HDRFILL
        c.alignment = Alignment(wrap_text=True, vertical='center')

def firsturl(s):
    if not s: return None
    m = re.search(r'https?://[^\s;|,)‖"]+', s)
    if m: return m.group(0)
    m = re.search(r'(?:www\.|banks\.data\.|api\.fdic\.|directory\.|ficoforums\.)[^\s;|,)"]+', s)
    return ('https://' + m.group(0)) if m else None

changelog = []  # (location, change, source)
def log(loc, change, src=''):
    changelog.append((loc, change, src))

# ---------- load inputs ----------
def loadcsv(p): return list(csv.DictReader(open(p, encoding='utf-8-sig')))
geo_lb = loadcsv(f'{V7OUT}/geo_longbeach.csv')
geo_fon = loadcsv(f'{V7OUT}/geo_fontana.csv')
icba = loadcsv(f'{V7OUT}/icba_community_banks.csv')
bureau = loadcsv(f'{V7OUT}/bureau_pulls.csv')
afill = loadcsv(f'{V7OUT}/assets_fill.csv')
w2g = loadcsv(f'{V7OUT}/w2_geo_screen.csv')
ewsmap = loadcsv('/home/user/awake/research/v6/out/ews_chex_mapping.csv')
bureau_by = {r['Institution']: r for r in bureau}
ews_by = {r['Institution']: r for r in ewsmap}
w2_by = {r['Institution']: r for r in w2g}

# =====================================================================
# RENAME carried tabs
# =====================================================================
renames = {
    'Funding Table': '3. FUNDING TABLE MASTER',
    '💎 Diamond-Platinum Tiers': '4. DIAMOND-PLATINUM',
    'EWS-Chex Map v6': '9. EWS-CHEX MAP',
    'FDIC SOD 6cty 2025': '10. FDIC SOD 6CTY',
    'NCUA CU 6cty 2026Q1': '11. NCUA 6CTY',
    'ibanknet CA 2026Q1': '12. IBANKNET CA',
    'v6 Change Log': '13. CHANGE LOG',
    'Legend': '14. LEGEND (v5)',
    'FDIC SOD Census (4cty)': '15. CENSUS 4CTY FDIC (v5)',
    'NCUA CU Census (4cty)': '16. CENSUS 4CTY NCUA (v5)',
}
for old, new in renames.items():
    wb[old].title = new
log('All tabs', 'RENAMED to numbered v7 scheme (content unchanged): ' + '; '.join(f'"{o}"→"{n}"' for o, n in renames.items()), 'user request: numbered tabs')

ft = wb['3. FUNDING TABLE MASTER']
bank_row = {ft.cell(r, 3).value: r for r in range(2, ft.max_row + 1)}

# =====================================================================
# TAB 3 ENRICHMENT (a) HP Bureau from bureau_pulls
# =====================================================================
bp_to_bank = {
    'PNC': 'PNC Bank', 'Tustin Community Bank': 'Tustin Community Bank',
    'Sunwest Bank': 'Sunwest Bank', 'BMO': 'BMO Bank',
    'First Citizens Bank': 'First Citizens Bank', 'KeyBank': 'KeyBank',
    'California Credit Union': 'California Credit Union',
    'First Western Trust Bank': 'First Western Trust Bank',
    'Valley Bank (Valley National)': 'Valley National Bank',
    'Banner Bank': 'Banner Bank', 'Western Alliance Bank': 'Western Alliance Bank',
    'WaFd Bank': 'WaFd Bank', 'Los Angeles FCU': 'Los Angeles Federal Credit Union (LAFCU)',
    'California Bank & Trust': 'California Bank & Trust',
    'Chase Ink': 'Chase (JPMorgan Chase)', 'Amex (business)': 'American Express',
    'Bank of America (business)': 'Bank of America', 'Wells Fargo (business)': 'Wells Fargo',
    'U.S. Bank (business)': 'U.S. Bank (direct)',
    'FNBO Business Edition': 'FNBO (First National Bank of Omaha)',
    'Kinecta': 'Kinecta Federal Credit Union',
    'SDCCU': 'San Diego County Credit Union (SDCCU)',
    'Premier America CU': 'Premier America Credit Union',
    'Enterprise Bank & Trust': 'Enterprise Bank & Trust', 'UMB': 'UMB Bank N.A.',
    'Comerica': 'Comerica Bank', 'Cathay Bank': 'Cathay Bank',
    'East West Bank': 'East West Bank', 'Hanmi Bank': 'Hanmi Bank',
    'Bank of Hope': 'Bank of Hope', 'Stanford FCU': 'Stanford Federal Credit Union',
    'Valley Strong CU': 'Valley Strong Credit Union', 'Wescom CU': 'Wescom Credit Union',
    'BCU': 'Baxter Credit Union (BCU)', 'Skyla CU': 'Skyla Credit Union',
}
BLANKISH = (None, '', '—', 'N/M', 'N/A', '-')
def append_cell(ws, r, c, addition):
    """add-only: append with v7 marker, or fill blank/N-M with dated value"""
    old = ws.cell(r, c).value
    if old in BLANKISH or (isinstance(old, str) and old.strip() in ('', '—', 'N/M', 'N/A', '-')):
        ws.cell(r, c).value = f'v7 {TODAY}: {addition}'
    else:
        ws.cell(r, c).value = f'{old}{MARK}{addition}'

hp_count = 0
for bp_name, bank in bp_to_bank.items():
    rec = bureau_by[bp_name]; r = bank_row[bank]
    ev = rec['New_evidence'].strip()
    if len(ev) > 230: ev = ev[:227] + '...'
    add = (f"bureau verdict [{rec['Evidence_grade']}] vs mentor '{rec['Mentor_claim']}' = "
           f"{rec['Mentor_verdict']}. {ev} (DPs {rec['DP_dates']}). Action: {rec['Action']}")
    append_cell(ft, r, 23, add)
    append_cell(ft, r, 36, f"bureau: {rec['Source_links']}")
    hp_count += 1
log('3. FUNDING TABLE MASTER col 23 (HP Bureau) + col 36 (Sources)',
    f'ENRICHED (appended) {hp_count} rows with v7 bureau-pull verdicts + grades + source links from bureau_pulls.csv',
    'research/v7/out/bureau_pulls.csv + 5 bureau findings MDs')

# (b) Assets / Deposits fill
ad_count = 0
for rec in afill:
    if rec['Sheet'] != 'funding_table': continue
    r = bank_row[rec['Row_key']]
    src_short = rec['Source'].replace('https://', '')
    for col, val in ((12, rec['Assets_value']), (13, rec['Deposits_value'])):
        if not val: continue
        append_cell(ft, r, col, f"{val} ({rec['As_of_date']}, {src_short})")
    ad_count += 1
log('3. FUNDING TABLE MASTER cols 12-13 (Assets/Deposits)',
    f'FILLED/ENRICHED {ad_count} rows with FDIC/NCUA 2026-03-31 totals (blank/N-M filled; existing values kept + appended)',
    'research/v7/out/assets_fill.csv (174 funding_table rows)')

# (c) merger flags
mergers = [
    ('Pacific Premier Bank', 'MERGER: acquired by Columbia Banking System (completed Sept 2025; ppbi.com redirects; personal-acct conversion done 2026-01-26). Treat PPBI as defunct for new apps — research Columbia Bank/Elan instead.', 'bureau_ca_regionals_findings.md'),
    ('Comerica Bank', 'MERGER: Comerica CA branches transferred to Fifth Third (2026-02). For branch apps at former Comerica CA locations, expect Fifth Third onboarding.', 'geo_fontana.csv row 14 (Fifth Third Bank, N.A. (ex-Comerica))'),
    ('First Foundation Bank', 'MERGER: now Sunflower Bank N.A. dba First Foundation (FirstSun merger, eff. 2026-04-01). Legacy Elan/Card Assets DPs may not predict current behavior.', 'bureau_ca_regionals_findings.md'),
    ('Community West Bank', 'MERGER NOTE: United Security Bank merged into Community West Bank — route any United Security leads here.', 'v7 merger sweep 2026-06-11'),
]
for bank, note, src in mergers:
    r = bank_row[bank]
    append_cell(ft, r, 3, '⚠ ' + note.split(':')[0])
    append_cell(ft, r, 35, note)
    append_cell(ft, r, 36, f'merger: {src}')
    log(f'3. FUNDING TABLE MASTER row {r} ({bank})', f'MERGER FLAG appended: {note}', src)

# (d) corrections
corr = [
    ('Wescom Credit Union', 17, 'CORRECTION: NO business credit card — consumer lineup only (business question moot). Bureau (consumer) = Experian (myFICO 3380449/6145700/6152311; WalletHub).', 'bureau_cus_findings.md'),
    ('SchoolsFirst Federal Credit Union', 17, 'CORRECTION: NO business card product (consumer only). EX proxies myFICO 1110928/6539392; CLI hard pull reusable 60d.', 'bureau_cus_findings.md'),
    ('WaFd Bank', 11, 'real business Visa DP: $15,000 @ 13.99% approved 2023-05-06 (myFICO thread 6674742; bureau unreadable without login — HIGH-VALUE to re-read logged-in).', 'bureau_mentor_extu_findings.md / myFICO 6674742'),
]
for bank, col, note, src in corr:
    r = bank_row[bank]
    append_cell(ft, r, col, note)
    append_cell(ft, r, 36, f'v7 correction: {src}')
    log(f'3. FUNDING TABLE MASTER row {r} ({bank})', f'CORRECTION appended: {note}', src)
# American Plus underwriter update (TCM not TIB)
r = bank_row['American Plus Bank, N.A.']
append_cell(ft, r, 18, 'UPDATE: cards = TCM Bank, N.A. (ICBA agent) — bankaplus.com card apply routes to mycommunitycc.com (TCM platform); supersedes earlier TIB inference. Commercial loans = self-UW.')
append_cell(ft, r, 36, 'v7 underwriter update: https://bankaplus.com/Credit-Cards; https://www.mycommunitycc.com/')
log(f'3. FUNDING TABLE MASTER row {r} (American Plus Bank)', 'UNDERWRITER UPDATE appended: cards = TCM Bank (mycommunitycc.com portal), supersedes TIB inference', 'research/v7/out/w2_geo_screen.csv')

# (e) NEW ROWS — Rec Rank 237+
def w2row(name): return w2_by[name]
NEWPFX = f'NEW v7 {TODAY} — '
new_rows = []
def mk_new(rank, bank, typ, chex, ews, limit, assets, deposits, ca_pres, selfuw, card, uw,
           apply_, how, site, cardlink, hp, nodoc, pg, bestfor, products, region, hq, conf, sources, geo):
    return {1: str(rank), 2: str(rank), 3: bank, 4: typ, 5: geo, 6: 'see Confidence col',
            7: chex, 8: ews, 9: None, 10: None, 11: limit, 12: assets, 13: deposits,
            14: ca_pres, 15: ca_pres, 16: selfuw, 17: card, 18: uw, 19: apply_, 20: how,
            21: site, 22: cardlink, 23: hp, 24: 'N/M', 25: limit, 26: nodoc, 27: pg,
            28: 'N/M', 29: 'N/M', 30: bestfor, 31: products, 32: region, 33: hq,
            34: 'n/a', 35: NEWPFX + conf, 36: sources}

w = w2row('Rize Credit Union (ex-SCE FCU)')
new_rows.append(mk_new(237, 'Rize Credit Union (ex-SCE FCU)', 'CU',
    'UNKNOWN — verify at join', 'UNKNOWN',
    '$5,000–$50,000 published range on Business Visa Platinum (rizecu.com/business-visa 2026-06-11)',
    '$1.3B (v7 screen)', 'N/M',
    'Ontario CA HQ; 975 N Haven Ave, Ontario 91764 = 7.88 mi from Fontana 92336; joinable from BOTH addresses via free American Consumer Council membership (only ID/UT/LA/TX excluded)',
    'Likely self (no Elan/myaccountaccess/TCM/cardaccount.net tells; cards serviced in own digital banking banking.rizecu.com) — pending issuer confirm',
    '✅ Business Visa Platinum: $0 AF, employee cards w/ custom limits, card controls',
    'Self (likely — confirm with loan consultant)', 'Phone/branch',
    'Business banking team via inquiry form or 800-866-6474; rates/terms not published',
    'https://rizecu.com', 'https://rizecu.com/business-visa/', 'UNKNOWN — ask at app',
    '3', 'UNKNOWN — "speak with a loan consultant"; business inquiry form',
    'PLATINUM CANDIDATE — published $50K biz card ceiling + ACC-joinable from Fontana AND Long Beach',
    'Business Visa Platinum; business checking; business loans', 'Inland Empire (Ontario)', 'Ontario, CA',
    'Medium-High (w2_geo_screen). TARGET. Pending: issuer confirm + Chex check.', w['Source_links'],
    'SoCal (Ontario, 7.88 mi from Fontana)'))

w = w2row('Southland Credit Union')
new_rows.append(mk_new(238, 'Southland Credit Union', 'CU',
    'Yes — pulls (CONFIRMED own FAQ: "All new accounts are verified through ChexSystems and may be subject to credit approval")', 'UNKNOWN',
    '$25,000 max total unsecured (published consumer card cap); commercial loan limits unpublished',
    '$1.29B (v7 screen)', 'N/M',
    '1050 Linden Ave, Long Beach 90813 = 1.14 mi from 90802. Joinable from Long Beach (live/work/worship/school LA or Orange Co.); NOT from Fontana (San Bernardino Co.)',
    'Self — "issued by Southland Credit Union pursuant to a license from Visa USA Inc."; no Elan/TCM tells',
    '⚠️ PARTIAL — business accounts + commercial loans confirmed; dedicated biz CC/LOC unverified (phone-confirm)',
    'Self', 'Branch', 'Branch or Business Account Application PDF; site partially blocks bots (403) — phone likely',
    'https://www.southlandcu.org', 'https://www.southlandcu.org/borrow/credit-cards/', 'UNKNOWN',
    '2', 'Business Account Application PDF published (sole prop/LLC/corp/nonprofit/HOA); loan docs unpublished',
    'Long Beach-only community CU 1.1 mi out; biz banking real',
    'Business checking (2 tiers), business savings, commercial loans; consumer Enjoy! Rewards Visa 0% 12mo', 'Long Beach / LA-OC', 'Los Alamitos, CA',
    'Medium (w2_geo_screen). TARGET (LB only). Chex = BLOCKER flag.', w['Source_links'],
    'SoCal (Long Beach 1.14 mi)'))

w = w2row('American Plus Bank, N.A. (dba International City Bank, Long Beach)')
new_rows.append(mk_new(239, 'American Plus Bank, N.A. dba International City Bank (Long Beach) [cross-ref rank 34]', 'Bank',
    'UNKNOWN', 'UNKNOWN',
    'UNKNOWN (TCM program; limits not published)',
    '$891M (v7 screen; CERT 58469)', 'N/M',
    '249 E Ocean Blvd, Long Beach 90802 = 0.56 mi from 1 World Trade Ctr; LB branch acquired from United Fidelity/ICB May 2025; open to public',
    'No — cards = TCM Bank, N.A. (ICBA agent): bankaplus.com card apply routes to mycommunitycc.com = TCM platform. Commercial loans = self-UW',
    '✅ Visa Charge Company Rewards (business charge card) — TCM CAVEAT: credit decision is TCM\'s, not APB\'s',
    'TCM Bank (cards); Self (commercial loans)', 'Branch/Online',
    'Branch 562-436-9800 or online card app via mycommunitycc.com',
    'https://bankaplus.com', 'https://bankaplus.com/Credit-Cards', 'UNKNOWN (TCM pull not published; TCM ≈ Experian per stacking.capital)',
    '2', 'UNKNOWN — contact branch; standard bank commercial underwriting expected',
    'Closest real business bank to LB address (0.56 mi) w/ a business charge card',
    'Visa Charge Company Rewards (biz); Visa Platinum Cash Rewards (consumer); commercial/small-biz/CRE/construction loans', 'Long Beach / SGV', 'Arcadia, CA',
    'Medium-High (w2_geo_screen). Reference row for the ICB Long Beach branch angle; master row = rank 34.', w['Source_links'],
    'SoCal (Long Beach 0.56 mi)'))

w = w2row('Alta Vista Credit Union')
new_rows.append(mk_new(240, 'Alta Vista Credit Union', 'CU',
    'UNKNOWN', 'UNKNOWN', 'UNKNOWN — commercial lending limits unpublished',
    '$212M (v7 screen)', 'N/M',
    '2025 N Riverside Ave, Rialto 92377 = 4.91 mi from Fontana 92336; community charter (San Bernardino Co. live/work/worship + Riverside Co.); NOT joinable from Long Beach',
    'UNKNOWN — no Elan/TCM tells; consumer cards appear in-house',
    '⚠️ PARTIAL — small business checking + commercial lending; NO biz credit card found',
    'Unknown', 'Branch', 'Branch (Redlands HQ / Rialto) 909-809-3838',
    'https://altavistacu.org', 'https://altavistacu.org/', 'UNKNOWN', '1',
    'UNKNOWN', 'REFERENCE — Fontana-joinable commercial lender, no biz card surfaced',
    'Small Business Checking, Commercial Lending, Small Business Digital Services; consumer credit cards', 'Inland Empire (Rialto/Redlands)', 'Redlands, CA',
    'Medium (w2_geo_screen). REFERENCE/secondary target.', w['Source_links'],
    'SoCal (Rialto, 4.91 mi from Fontana)'))

w = w2row('Thinkwise Credit Union')
new_rows.append(mk_new(241, 'Thinkwise Credit Union', 'CU',
    'UNKNOWN', 'UNKNOWN', 'UNKNOWN',
    '$121M (v7 screen)', 'N/M',
    '1270 Renaissance Pkwy, Rialto 92376 = 3.91 mi from Fontana 92336. FOM GATE: school employees / SB-Orange-Riverside Co. member-openers OR business HQ in 3 counties demonstrably supporting the educational sector — plain Fontana residence does NOT qualify',
    'UNKNOWN — no third-party tells; small ($121M) so agent program possible but unverified',
    '✅ Business credit cards AND business LOC (restricted industries: cannabis, gambling, MSB, adult)',
    'Unknown — verify', 'Email/phone', 'Pre-review via mybusiness@thinkwisecu.org or 909-474-8448, or online interest form',
    'https://www.thinkwisecu.org', 'https://www.thinkwisecu.org/business-membership', 'UNKNOWN', '2',
    'Best-documented in cohort: govt ID for signers/BOs (25%+), EIN/SSN, FinCEN BO cert, formation docs, business license/Home Occupation Permit, $25 deposit',
    'REFERENCE (conditional) — full biz CC+LOC menu, education-sector FOM gate',
    'Business credit cards, business LOC, business checking/savings', 'Inland Empire (Rialto/San Bernardino)', 'San Bernardino, CA',
    'Medium-High (w2_geo_screen). Not previously in table → ADDED.', w['Source_links'],
    'SoCal (Rialto, 3.91 mi from Fontana)'))

w = w2row('Navy Federal Credit Union')
new_rows.append(mk_new(242, 'Navy Federal Credit Union', 'CU',
    'UNKNOWN (community reports say Chex used for consumer accounts — UNVERIFIED)', 'UNKNOWN',
    'UNKNOWN published max; ~750 score expectation cited; 1 yr individual membership or ~2 yrs in business cited',
    '$203.6B (largest US CU)', 'N/M',
    '4190 E 4th St Ste A, Ontario 91764 = 7.18 mi from Fontana 92336. MILITARY GATE: active duty/retiree/veteran, immediate family/household, or DoD civilian; ALL business owners must be NFCU members',
    'Self — Navy Federal issues and services its own cards (navyfederal.org program description nfcu_825cc.pdf)',
    '✅ GO BIZ Rewards (Visa or Mastercard)',
    'Self', 'Paper+upload', 'Paper app + beneficial owner form, wet signature, scanned + submitted via NFCU business account',
    'https://www.navyfederal.org/services/business/credit-cards.html', 'https://www.navyfederal.org/services/business/credit-cards.html',
    'Community-reported TransUnion and/or Equifax (DoC/myFICO communities; not official)', '2',
    'Application + BO form, wet signature; income/credit review; supporting docs on request',
    'REFERENCE — military gate; only actionable if a principal has military/DoD affiliation',
    'GO BIZ Rewards (Visa/MC), business checking/savings, business loans/LOC', 'National (Ontario CA branch)', 'Vienna, VA',
    'High (w2_geo_screen). REFERENCE row (military affiliation required).', w['Source_links'],
    'National (Ontario branch 7.18 mi from Fontana)'))

new_rows.append(mk_new(243, 'Chino Commercial Bank, N.A. — LOC reference [cross-ref rank 188]', 'Bank',
    'N/M (per master row)', 'N/M',
    'Unsecured business LOC offered (limits unpublished); NO business credit card (consumer third-party card only)',
    '$0.47B', 'N/M',
    '8229 Rochester Ave, Rancho Cucamonga 91730 = 5.50 mi from Fontana 92336 (geo_fontana #16)',
    'Self (commercial lending in-house); no business card',
    '❌ NO business card — LOC path only (master row rank 188 re-verified v6 2026-05-29)',
    'Self (LOC)', 'Branch', 'Branch — Rancho Cucamonga office; relationship commercial bank',
    'https://www.chinocommercialbank.com', None, 'N/M', '1',
    'Standard commercial LOC underwriting expected (full-doc likely)',
    'REFERENCE — nearby Fontana LOC option, not a card play',
    'Business LOC, commercial lending; consumer third-party card only', 'Inland Empire (Chino/RC)', 'Chino, CA',
    'Reference row added for LOC-no-card clarity near Fontana; master row = rank 188.',
    'https://www.chinocommercialbank.com; research/v7/out/geo_fontana.csv; v6 Funding Table rank 188',
    'SoCal (Rancho Cucamonga 5.50 mi from Fontana)'))

start = ft.max_row + 1
for i, nr in enumerate(new_rows):
    r = start + i
    for c, v in nr.items():
        if v is not None:
            put(ft, r, c, v)
    log(f'3. FUNDING TABLE MASTER row {r} (Rec Rank {nr[1]})',
        f'NEW ROW ADDED: {nr[3]} — {nr[30]}', nr[36][:200])
n_new = len(new_rows)

# =====================================================================
# TAB 4 DIAMOND-PLATINUM — appends
# =====================================================================
dp = wb['4. DIAMOND-PLATINUM']
# locate D1 (FNBO) and P1 (Stanford) rows
d1 = p1 = None
for r in range(1, dp.max_row + 1):
    v = str(dp.cell(r, 1).value)
    if v == 'D1': d1 = r
    if v == 'P1': p1 = r
append_cell(dp, d1, 8, 'bureau RESOLVED = EXPERIAN, business-specific DP (Business Edition pre-qual "based on your personal credit Experian file", myFICO 6348432 ~2022-23; DoC master page "usually pull Experian"). Pre-qual = SOFT, full app = hard.')
append_cell(dp, d1, 11, f'v7 {TODAY} bureau: bureau_mentor_eq_findings.md; ficoforums.myfico.com thread 6348432')
append_cell(dp, p1, 8, 'membership pull = EXPERIAN hard (reusable ~30-60d; myFICO 6726623/6723581/6036159). CONSUMER-PROXY for the biz card.')
append_cell(dp, p1, 11, f'v7 {TODAY} bureau: bureau_cus_findings.md; myFICO 6726623/6723581')
nr_dp = dp.max_row + 2
put(dp, nr_dp, 1, f'v7 {TODAY} ADDITIONS:', font=BOLD)
vals = ['P5', 'Rize Credit Union (ex-SCE FCU)', '🥈 Platinum candidate (NEW v7)',
        '✅ likely self-issued — no Elan/TCM/myaccountaccess tells; cards serviced in own digital banking (pending issuer confirm)',
        '✅ $5,000–$50,000 PUBLISHED range on Business Visa Platinum (rizecu.com/business-visa)',
        '❓ UNKNOWN — Chex check pending',
        '✅ joinable from Fontana AND Long Beach via free American Consumer Council membership; Ontario branch 7.88 mi from Fontana',
        'Issuer confirm (loan consultant) + Chex behavior at join',
        'Call 800-866-6474: confirm self-issue, Chex at join, doc stack for $50K',
        '$5K–$50K published', f'rizecu.com/business-visa (v7 {TODAY}); w2_geo_screen.csv']
for c, v in enumerate(vals, 1):
    put(dp, nr_dp + 1, c, v, fill=GREEN if c == 3 else None)
log('4. DIAMOND-PLATINUM', 'APPENDED: FNBO D1 bureau resolved = Experian (biz-specific DP); Stanford P1 membership pull = Experian hard; NEW P5 row Rize CU Platinum candidate ($50K published, ACC-joinable; pending issuer confirm + Chex)', 'bureau_mentor_eq_findings.md; bureau_cus_findings.md; w2_geo_screen.csv')

# =====================================================================
# helper data for TAB 1
# =====================================================================
def ews_links(name_frag):
    for k, v in ews_by.items():
        if name_frag.lower() in k.lower():
            return firsturl(v['Source_links'])
    return None
DOC_CHEX = 'https://www.doctorofcredit.com/banks-credit-unions-dodont-pull-chexsystems/'
DOC_EWS = 'https://www.doctorofcredit.com/banks-credit-unions-dodont-pull-early-warning-services/'
def bp_url(inst):
    r = bureau_by.get(inst)
    return firsturl(r['Source_links']) if r else None

def ftv(bank, col):
    return ft.cell(bank_row[bank], col).value

# Tab 1 row spec: (name, ftkey, why, nearest, bureau_txt, bureau_url, chex, chex_url, ews, ews_url,
#                  card, card_url, limit, limit_url, nodoc, nodoc_url, selfuw, apply, assets, action)
T1 = [
 ('Chase Ink', 'Chase (JPMorgan Chase)', 'Best self-UW online; Ink train', 'LB 0.44 mi; Fontana 1.18 mi',
  'EX-lean; CA dual EX+EQ seen [COMMUNITY-DP]', 'https://www.doctorofcredit.com/knowledge-base/which-credit-bureau-does-chase-pull/',
  'No', ews_links('Chase'), 'Yes', ews_links('Chase'),
  'Y', 'https://creditcards.chase.com/business-credit-cards/ink/cash',
  '$5K–$36K (myFICO 6803261)', 'https://creditcards.chase.com/business-credit-cards/ink/cash',
  'Y — PG, no tax returns', 'https://creditcards.chase.com/business-credit-cards',
  'Self (JPMorgan Chase NA)', 'Online', '~$4.9T', 'Apply Ink Cash; mind 5/24'),
 ('Amex (business)', 'American Express', 'EX-only; soft pull for members', 'Online (national)',
  'EX ~95%; CA EX 41/TU 2/EQ 1 [COMMUNITY-DP]', 'https://www.doctorofcredit.com/credit-bureau-american-express-pull/',
  'Yes (consumer chk)', ews_links('American Express') or DOC_CHEX, 'Yes', DOC_EWS,
  'Y', 'https://www.americanexpress.com/us/credit-cards/business/business-credit-cards/',
  '$6K–$20K; charge NPSL to $250K (myFICO 6507116)', 'https://www.americanexpress.com/us/credit-cards/business/business-credit-cards/',
  'Y — PG, no tax returns', 'https://www.americanexpress.com/us/credit-cards/business/business-credit-cards/',
  'Self (American Express NB)', 'Online', '—', 'Unfreeze EX; apply Blue Business'),
 ('U.S. Bank (card + Cash Flow Manager)', 'U.S. Bank (direct)', 'TU biz DP; LOC sidecar', 'LB 0.80 mi; Fontana 1.40 mi',
  'TU (biz Triple Cash DP ~2026-01) [COMMUNITY-DP]; CFM LOC bureau UNKNOWN', bp_url('U.S. Bank (business)'),
  'Yes (not sensitive)', ews_links('U.S. Bank') or DOC_CHEX, 'Yes', DOC_EWS,
  'Y', 'https://www.usbank.com/business-banking/business-credit-cards.html',
  '$3K–$10K typ; $20K–$35K strong (myFICO 6845780); CFM LOC no-doc at standard limits', 'https://www.usbank.com/business-banking/business-credit-cards.html',
  'Y — no tax returns (cards/CFM std)', 'https://www.usbank.com/business-banking/business-lines-of-credit.html',
  'Self (U.S. Bank NA)', 'Online', '~$695B ‖ $683.38B (FDIC 2026-03-31)', 'Freeze SageStream+ARS first; apply Triple Cash'),
 ('FNBO (First National Bank of Omaha)', 'FNBO (First National Bank of Omaha)', '$50K stated max; EX biz DP', 'Online (national)',
  'EX — BUSINESS-specific DP (prequal language) [COMMUNITY-DP]', bp_url('FNBO Business Edition'),
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.fnbo.com/small-business/credit-cards',
  '$50,000 stated MAX (fnbo.com); instants $7.1K–$14K', 'https://www.fnbo.com/small-business/credit-cards',
  'Y — no financials at app (std limits)', 'https://www.fnbo.com/small-business/credit-cards',
  'Self ("Cards are issued by FNBO")', 'Online', '—', 'Run soft-pull pre-qual NOW'),
 ('Bank of America', 'Bank of America', 'EX; no Chex; big instants', 'Fontana 1.50 mi; LB 2.11 mi',
  'EX-primary (82% of 129 DPs; CA EX 15/TU 6/EQ 4) [COMMUNITY-DP]', 'https://www.doctorofcredit.com/credit-bureau-bank-america-pull/',
  'No', ews_links('Bank of America'), 'Yes', ews_links('Bank of America'),
  'Y', 'https://business.bankofamerica.com/en/credit-cards',
  '$8K–$16K typ; to $39.5K instant (myFICO 6510094)', 'https://business.bankofamerica.com/en/credit-cards/business-advantage-unlimited-cash-rewards',
  'Y — PG, no tax returns', 'https://business.bankofamerica.com/en/credit-cards',
  'Self (Bank of America NA)', 'Online', '~$3.41T ‖ $2.67T bank-level (FDIC 2026-03-31)', 'Apply Business Advantage online'),
 ('BMO', 'BMO Bank', 'TU confirmed; 0% 18 mo', 'LB 0.68 mi; Fontana 7.16 mi (RC)',
  'TU — mentor AGREES, strongest DP set ($7.5K biz approval) [COMMUNITY-DP]', bp_url('BMO'),
  'Yes (not inquiry-sensitive)', ews_links('BMO') or DOC_CHEX, 'Reports (no inquiry DP)', DOC_EWS,
  'Y', 'https://www.bmo.com/en-us/main/business-banking/credit-cards/',
  'DPs $1.3K–$7.5K; grows via CLI (myFICO 6773916)', 'https://www.bmo.com/en-us/main/business-banking/credit-cards/platinum/',
  'Y — no income docs typical', 'https://www.bmo.com/en-us/main/business-banking/credit-cards/',
  'Self (BMO Bank NA)', 'Online', '~$265B', 'Open biz deposit acct first, then card'),
 ('Enterprise Bank & Trust', 'Enterprise Bank & Trust', 'Self-UW verified; ~$50K claimed', '8 SoCal branches (Cerritos/DTLA/Anaheim)',
  'NO-DATA (datapoint desert); FIS rails verified [UNKNOWN]', bp_url('Enterprise Bank & Trust'),
  'Unknown (absent DoC lists)', DOC_CHEX, 'Unknown (absent DoC lists)', DOC_EWS,
  'Y', 'https://www.enterprisebank.com/business/credit-cards',
  '~$50K claimed via broker channel (UNVERIFIED)', 'https://www.enterprisebank.com/business/credit-cards',
  'Partial — terms unpublished; RM-flex', 'https://www.enterprisebank.com/business/credit-cards',
  'Self (VERIFIED HIGH — own card loans on call report)', 'RM', '$17.2B', 'Email RM jmorrow@ / (833) 896-2850 for written terms'),
 ('Stanford FCU', 'Stanford Federal Credit Union', '$50K card + $100K LOC, self-UW', 'Online (remote join, Zoom verify)',
  'EX — membership hard pull, reusable ~30-60d [CONSUMER-PROXY]', bp_url('Stanford FCU'),
  'Unknown (site 403s)', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.sfcu.org/business/business-credit-card',
  'Biz Visa to $50K + unsecured LOC to $100K (sfcu.org)', 'https://www.sfcu.org/business/business-credit-card',
  'Partial — PFS-based; no tax returns stated (card)', 'https://www.sfcu.org/business/business-credit-card',
  'Self (own Business Visa disclosure)', 'Online', '$4.8B ‖ NCUA 2026-03-31 fill', 'Start 6-month membership clock NOW ($5)'),
 ('Rize CU (ex-SCE FCU)', None, 'NEW: $50K published, ACC join', 'Ontario 7.88 mi (Fontana); ACC join from both',
  'UNKNOWN — ask at app', "#'8. BUREAU EVIDENCE'!A1",
  'Unknown — verify', "#'3. FUNDING TABLE MASTER'!A238", 'Unknown', "#'3. FUNDING TABLE MASTER'!A238",
  'Y', 'https://rizecu.com/business-visa/',
  '$5K–$50K published (rizecu.com/business-visa)', 'https://rizecu.com/business-visa/',
  'Partial — "speak with a loan consultant"', 'https://rizecu.com/business-visa/',
  'Likely self (no agent tells) — confirm', 'Phone/branch', '$1.3B', 'Call 800-866-6474: issuer + Chex + docs'),
 ('PNC', 'PNC Bank', 'Soft-EQ BusinessOptions path', 'Online (CA footprint thin)',
  'EX hard (std card); BusinessOptions/LOC = SOFT EQ [CONSUMER-PROXY]', bp_url('PNC'),
  'Yes (mixed DPs)', ews_links('PNC') or DOC_CHEX, 'Yes', DOC_EWS,
  'Y', 'https://www.pnc.com/en/small-business/borrowing/business-credit-cards.html',
  'BusinessOptions/LOC $25K–$100K soft-pull (myFICO 6589727)', 'https://www.pnc.com/en/small-business/borrowing/business-credit-cards/pnc-businessoptions-visa-signature-card.html',
  'Y — soft-pull path, no tax returns', 'https://www.pnc.com/en/small-business/borrowing/business-credit-cards.html',
  'Self (PNC Bank NA)', 'Online', '~$574B', 'Unfreeze EX+EQ; target BusinessOptions'),
 ('Ramp', 'Ramp (corporate charge card)', 'No pull, no PG, cash-UW', 'Online (national)',
  'NONE — no personal credit check [VERIFIED]', 'https://ramp.com',
  'N/A (fintech KYC)', 'https://ramp.com', 'N/A', 'https://ramp.com',
  'Y (charge)', 'https://ramp.com',
  '~10–20% of linked business cash; needs ≥$25K balance', 'https://ramp.com',
  'Y — bank-link IS the underwriting', 'https://ramp.com',
  'Issuers: Celtic + Sutton Bank', 'Online', 'issuer Celtic $4.97B', 'Layer now — costless, no pull'),
 ('Mercury IO', 'Mercury (IO charge card)', 'No pull; balance-based', 'Online (national)',
  'NONE — no credit check [VERIFIED]', 'https://mercury.com/credit',
  'N/A (fintech KYC)', 'https://mercury.com/credit', 'N/A', 'https://mercury.com/credit',
  'Y (charge)', 'https://mercury.com/credit',
  '~$5K intro; $15K+ kept at Mercury unlocks more', 'https://mercury.com/credit',
  'Y — no docs, balance-based', 'https://mercury.com/credit',
  'Issuer: Patriot Bank N.A.', 'Online', 'issuer Patriot $1.18B', 'Open Mercury checking; IO in-dashboard'),
 ('BILL Spend & Expense (Divvy)', 'BILL Spend & Expense (Divvy)', '$20–50K lines, no tax docs', 'Online (national)',
  'No hard personal pull reported [COMMUNITY-DP]', 'https://www.bill.com/product/spend-and-expense',
  'Unknown (no traditional Chex)', 'https://www.bill.com/product/spend-and-expense', 'Unknown', 'https://www.bill.com/product/spend-and-expense',
  'Y (charge)', 'https://apply.divvy.co',
  '$20K–$50K common w/ real balance (Nav)', 'https://apply.divvy.co',
  'Y — bank-link; PG status CONFLICTING', 'https://www.bill.com/product/spend-and-expense',
  'Issuers: Cross River + WEX', 'Online', 'issuer Cross River $8.71B', 'Apply after parking balance; document PG ask'),
 ("Sam's Club Business MC (Synchrony)", "Sam's Club Business Mastercard (Synchrony)", 'EIN-only no-PG path', 'Clubs across IE/LA (in-club app)',
  'EIN-only path: NO personal pull; PG path: Synchrony HP [COMMUNITY-DP]', 'https://www.samsclub.com/credit',
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.samsclub.com/credit',
  '$8K–$10K instant; $30K via CLIs (myFICO 6701656)', 'https://www.samsclub.com/credit',
  'Y — EIN-only w/ established D&B file', 'https://www.samsclub.com/credit',
  'Synchrony co-brand', 'In-club', 'Synchrony $114.86B', 'Build D&B file, then in-club EIN-only'),
 ('First Citizens', 'First Citizens Bank', '$20K SL DP; tri-bureau', 'LB 2.82 mi (branch-apply only)',
  'Tri-bureau blend + biz bureaus; EQ shows on checking [COMMUNITY-DP]', bp_url('First Citizens Bank'),
  'Yes (sensitive 32+/12mo only)', ews_links('First Citizens') or DOC_CHEX, 'Reports (no inquiry DP)', DOC_EWS,
  'Y', 'https://www.firstcitizens.com/small-business/credit-financing/credit-cards',
  '$20K SL (Low-Interest, branch + biz checking; myFICO 6704468)', 'https://www.firstcitizens.com/small-business/credit-financing/credit-cards',
  'Partial — no tax returns but branch+relationship gated', 'https://www.firstcitizens.com/small-business/credit-financing/credit-cards',
  'Self (First-Citizens B&T)', 'Branch', '~$236B', 'Open biz checking at LB branch, then card'),
 ('California Bank & Trust (CB&T)', 'California Bank & Trust', 'Mentor: branch verify; $250K ads', 'LB 0.12 mi (Zions branch)',
  'BRANCH VERIFY (mentor) — zero public DPs, AGREES [UNKNOWN]', bp_url('California Bank & Trust'),
  'Yes (Zions deposit agreement)', ews_links('California Bank') or DOC_CHEX, 'Furnishes (no inquiry screen)', DOC_EWS,
  'Y', 'https://www.calbanktrust.com/business/business-credit-cards/',
  'Advertised up to $250K (AmaZing)', 'https://www.calbanktrust.com/business/business-credit-cards/',
  'Y — no tax-return req advertised (cards)', 'https://www.calbanktrust.com/business/business-credit-cards/',
  'Self (Zions Bancorporation NA)', 'Online', '—', 'Ask banker the bureau at 444 W Ocean Blvd'),
 ('Banc of California', 'Banc of California', '0.0 mi LB; best 0% (20 cyc)', 'LB 0.0 mi (WTC); Fontana 7.80 mi',
  'Elan-variable: EQ then EX seen, same applicant (myFICO 6244031) [COMMUNITY-DP]', "#'8. BUREAU EVIDENCE'!A1",
  'Yes (weak source; verify)', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://bancofcal.com/payment-solutions/business-credit-card/',
  '~$3K–$10K start, ~$20K strong (Elan; myFICO 6658960)', 'https://bancofcal.com/payment-solutions/business-credit-card/',
  'Y — PG, no tax returns <$25K', 'https://bancofcal.com/payment-solutions/business-credit-card/',
  'Elan (biz card); self commercial card', 'Online', '$34.64B', '0% 20-cycle card if Elan OK'),
 ('City National Bank', 'City National Bank', '0.09 mi LB; RM commercial', 'LB 0.09 mi; Ontario 7.91 mi',
  'EX (personal Crystal proxy); biz N/M [CONSUMER-PROXY]', "#'8. BUREAU EVIDENCE'!A1",
  'Yes (not sensitive)', ews_links('City National') or DOC_CHEX, 'Unknown', DOC_EWS,
  'Y (commercial only)', 'https://www.cnb.com/business-banking/lending/credit-cards/commercial.html',
  'Commercial Visa — limits relationship-based', 'https://www.cnb.com/business-banking/lending/credit-cards/commercial.html',
  'N — full-doc (PFS + tax returns)', 'https://www.cnb.com/business-banking/lending/credit-cards/commercial.html',
  'Self', 'Branch/RM', '$99.95B', 'Only with RM relationship; full-doc'),
 ('F&M Bank Long Beach', 'Farmers & Merchants Bank of Long Beach', '0.47 mi; LB institution', 'LB 0.47 mi (302 Pine Ave)',
  'Elan-variable (TU-lean, EX/EQ seen) [COMMUNITY-DP]', "#'8. BUREAU EVIDENCE'!A1",
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://creditcardlearnmore.com/11t3/index?ecdma-lc=31463&ecid=OTHE_25940',
  '~$3K–$10K start (Elan platform; myFICO 6658960)', 'https://creditcardlearnmore.com/11t3/index?ecdma-lc=31463&ecid=OTHE_25940',
  'Y — PG, no tax returns <$25K', 'https://www.fmb.com',
  'Elan (card); relationship bank', 'Online', '$11.86B', 'Elan roulette; or branch relationship LOC'),
 ('Comerica → Fifth Third (CA)', 'Comerica Bank', 'Branches now 5/3 (2026-02)', 'LB 0.58 mi; RC 5.29 mi — now Fifth Third',
  'Elan (usually EX, TU seen); new customers in-branch [UNKNOWN]', bp_url('Comerica'),
  'Yes (not sensitive 80+/24mo OK)', ews_links('Comerica') or DOC_CHEX, 'Yes', DOC_EWS,
  'Y', 'https://www.comerica.com/small-business/banking-and-services/business-credit-cards.html',
  '$25K starting DP (Elan; creditboards 632782)', 'https://www.comerica.com/small-business/banking-and-services/business-credit-cards.html',
  'Y — PG, no tax returns', 'https://www.comerica.com/small-business/banking-and-services/business-credit-cards.html',
  'Elan', 'Branch', '~$79B', 'Verify Fifth Third onboarding at ex-Comerica branch'),
 ('UMB Bank', 'UMB Bank N.A.', 'Self-issued; online app', 'Online (dao.umb.com/sbla)',
  'TU (consumer Simply Rewards proxy) [COMMUNITY-DP]', bp_url('UMB'),
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.umb.com/business-banking/business-credit-cards',
  'No published max, no DPs — ask RM', 'https://www.umb.com/business-banking/business-credit-cards',
  'Y — questionnaire-style, no tax returns stated', 'https://www.umb.com/business-banking/business-credit-cards',
  'Self ("issued by UMB Bank n.a.")', 'Online', '~$68B', 'Apply online (15–20 min) or ask limit band first'),
 ('Valley Strong CU', 'Valley Strong Credit Union', 'Open join; self-UW inferred', 'Online join (FFA) from both addresses',
  'UNKNOWN; membership = SOFT pull (DoC) [UNKNOWN]', bp_url('Valley Strong CU'),
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.valleystrong.com/business-platinum-credit-card',
  'Unpublished ("high spending limits") — phone screen', 'https://www.valleystrong.com/business-platinum-credit-card',
  'Partial — docs unpublished', 'https://www.valleystrong.com',
  'Self (INFERRED-strong, CFPB agreement DB)', 'Online', '$4.1B', 'PHONE SCREEN: limits + docs + Chex'),
 ('Southland CU', None, 'NEW: 1.1 mi LB; Chex BLOCKER', 'LB 1.14 mi (1050 Linden Ave)',
  'UNKNOWN', "#'8. BUREAU EVIDENCE'!A1",
  'Yes — CONFIRMED (own FAQ)', 'https://www.southlandcu.org/faq/', 'Unknown', DOC_EWS,
  'N (biz CC unverified)', 'https://www.southlandcu.org/borrow/credit-cards/',
  '$25K max unsecured (consumer cap published)', 'https://www.southlandcu.org/borrow/credit-cards/',
  'Partial — biz account app PDF published', 'https://www.southlandcu.org/southlandcu/media/PDF-Form/account_application_business.pdf',
  'Self ("issued by Southland CU")', 'Branch', '$1.29B', 'Phone-confirm biz CC exists; clean Chex required'),
 ('LAFCU', 'Los Angeles Federal Credit Union (LAFCU)', '$50K max ads; Chex BLOCKER', 'LA County branches (~20 mi LB)',
  'TU (mentor + ~2010 rep DP + Dovly tie-in) [PARTIAL proxy]', bp_url('Los Angeles FCU'),
  'Yes — own site verbatim (BLOCKER)', 'https://www.lafcu.org', 'Unknown', DOC_EWS,
  'Y', 'https://www.lafcu.org/loans-credit/business-loans',
  'Up to $50,000 advertised MAX', 'https://www.lafcu.org/loans-credit/business-loans',
  'Partial — income docs N/M; 2yr TIB gate', 'https://www.lafcu.org/loans-credit/business-loans',
  'Self (LAFCU is creditor)', 'Online', '~$1.28B', 'Only w/ clean Chex + low inquiries + 2yr TIB'),
 ('Arrowhead CU', 'Arrowhead Credit Union', '1.2 mi Fontana; loans only', 'Fontana 1.22 mi; LB 4.75 mi',
  'N/M (free EX/FICO score to cardholders) [UNKNOWN]', "#'8. BUREAU EVIDENCE'!A1",
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'N (no biz card)', 'https://www.arrowheadcu.org/arrowhead-visa',
  'No biz card; business = loans/SBA/CRE', 'https://www.arrowheadcu.org/arrowhead-visa',
  'N — loan path full-doc likely', 'https://www.arrowheadcu.org/arrowhead-visa',
  'Self (consumer cards)', 'Branch', '~$2.31B ‖ NCUA fill', 'Skip for cards; SBA/LOC backup only'),
 ('CU SoCal', 'Credit Union of Southern California (CU SoCal)', '1.8 mi Fontana; no biz card', 'Fontana 1.83 mi',
  'N/M [UNKNOWN]', "#'8. BUREAU EVIDENCE'!A1",
  'Yes — HIGHLY SENSITIVE (denied 3 inq/6mo)', ews_links('CU SoCal') or DOC_CHEX, 'Unknown', DOC_EWS,
  'N (no biz card)', 'https://www.cusocal.org/cards/credit-card/',
  'Consumer Visa only ("may not be used for business")', 'https://www.cusocal.org/cards/credit-card/',
  'N', 'https://www.cusocal.org/cards/credit-card/',
  'Self (consumer)', 'Branch', '~$3.7B', 'Skip — no biz card + Chex-sensitive'),
 ('Premier America CU', 'Premier America Credit Union', 'EX; $50K cap; Chex-sensitive', 'SoCal branches (Chatsworth HQ)',
  'EX (consumer card hard-EX 2022 + EX×4 prior) [COMMUNITY-DP]', bp_url('Premier America CU'),
  'Yes — INQUIRY-SENSITIVE (3/90d deny)', ews_links('Premier America') or DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.premieramerica.com/business/loans/premier-business-mastercard',
  '$5K–$50K ($50K needs 2+yr profitable)', 'https://www.premieramerica.com/business/loans/premier-business-mastercard',
  'Y — PG, no tax returns on app', 'https://www.premieramerica.com/business/loans/premier-business-mastercard',
  'Self (confirmed creditor)', 'Online', '~$3.31B', 'Only if Chex inquiries <3/90d'),
 ('Kinecta FCU', 'Kinecta Federal Credit Union', 'EX; in-house biz MC', 'South Bay branches (~10 mi LB)',
  'EX-only consensus (myFICO 5679332) [CONSUMER-PROXY]', bp_url('Kinecta'),
  'Yes — pulls + reports (own doc)', ews_links('Kinecta') or DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.kinecta.org/credit-card/mypro',
  'N/M — no DP found; 3/2/1% cash back', 'https://www.kinecta.org/credit-card/mypro',
  'Partial — financials on request', 'https://www.kinecta.org/credit-card/mypro',
  'Self (Kinecta FCU)', 'Branch', '~$6.8B', 'By-appointment app; ask limit band'),
 ('SDCCU', 'San Diego County Credit Union (SDCCU)', 'EX membership pull reused', 'Online (SD-based; CA join)',
  'EX membership hard pull, reused (myFICO 6761567, 2024-05-24) [COMMUNITY-DP]', bp_url('SDCCU'),
  'Yes — uses/furnishes (own FAQ)', ews_links('SDCCU') or DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.sdccu.com/business/business-credit-card/',
  '$5K–$20K (myFICO 6761567); 11.74% low APR', 'https://www.sdccu.com/business/business-credit-card/',
  'Partial — income reviewed', 'https://www.sdccu.com/business/business-credit-card/',
  'Self (SDCCU)', 'Online', '~$9.26B', 'One EX pull covers join+card; clean Chex needed'),
 ('Banner Bank', 'Banner Bank', 'Mentor TU; Chex BLOCKER flag', 'RC 5.37 mi (Fontana side)',
  'Mentor: TU — NO public DP; bank will DISCLOSE bureau on request [MENTOR-ONLY]', bp_url('Banner Bank'),
  'Yes — SENSITIVE (BLOCKER; BankOn acct = fallback)', ews_links('Banner') or DOC_CHEX, 'Unknown', DOC_EWS,
  'Y', 'https://www.bannerbank.com/business-solutions/services/small-business-credit-card/business-platinum-mastercard',
  'N/M — no DP; 0% 12mo $0 AF', 'https://www.bannerbank.com/business-solutions/services/small-business-credit-card/business-platinum-mastercard',
  'N — full PG underwrite (Tax ID + revenue + guarantor)', 'https://www.bannerbank.com/business-solutions/services/small-business-credit-card/business-platinum-mastercard',
  'Self (Banner Bank; ampliFI rewards)', 'Branch (PDF app)', '~$16.56B ‖ $16.34B (FDIC 2026-03-31)', 'Clean Chex first; ask bureau post-app (they disclose)'),
 ('Chino Commercial Bank (LOC)', 'Chino Commercial Bank, N.A.', 'Nearby LOC; NO card', 'RC 5.50 mi (Fontana side)',
  'N/M [UNKNOWN]', "#'8. BUREAU EVIDENCE'!A1",
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'N (LOC only)', 'https://www.chinocommercialbank.com',
  'Unsecured biz LOC offered; limits unpublished', 'https://www.chinocommercialbank.com',
  'N — standard commercial UW expected', 'https://www.chinocommercialbank.com',
  'Self (LOC); no card', 'Branch', '$0.47B', 'LOC backup only — not a card play'),
 ('American Plus Bank dba ICB (LB)', 'American Plus Bank, N.A.', '0.56 mi LB; TCM caveat', 'LB 0.56 mi (249 E Ocean Blvd)',
  'TCM program — pull not published (TCM ≈ EX per stacking.capital) [UNKNOWN]', "#'8. BUREAU EVIDENCE'!A1",
  'Unknown', DOC_CHEX, 'Unknown', DOC_EWS,
  'Y (TCM-issued)', 'https://bankaplus.com/Credit-Cards',
  'UNKNOWN (TCM; limits not published)', 'https://bankaplus.com/Credit-Cards',
  'Partial — branch UW expected', 'https://www.mycommunitycc.com/',
  'TCM Bank (cards); self commercial loans', 'Branch/Online', '$891M', 'TCM caveat: decision is TCM\'s — branch-ask first'),
]

ws1 = wb.create_sheet('1. BEST TARGETS')
H1 = ['Rank', 'Institution', "Why it's here", 'Nearest location + miles', 'Bureau pulled (+grade)',
      'ChexSystems', 'EWS', 'Biz card confirmed?', 'Stated/Est. limit', 'No-doc?', 'Self-UW / issuer',
      'Apply', 'Assets', 'Next action']
header_row(ws1, 1, H1)
def chexfill(t):
    t = t.lower()
    if t.startswith('no') or t.startswith('n/a'): return GREEN
    if 'sensitive' in t or 'blocker' in t: return RED
    if t.startswith('yes'): return AMBER
    return AMBER
for i, row in enumerate(T1):
    (name, ftk, why, near, btxt, burl, chex, churl, ews, eurl, card, curl,
     lim, lurl, nodoc, nurl, suw, apply_, assets, action) = row
    r = i + 2
    put(ws1, r, 1, i + 1, font=BOLD)
    site = ftv(ftk, 21) if ftk else {'Rize CU (ex-SCE FCU)': 'https://rizecu.com', 'Southland CU': 'https://www.southlandcu.org'}[name]
    link(ws1, r, 2, name, firsturl(str(site)) or site)
    put(ws1, r, 3, why)
    put(ws1, r, 4, near)
    link(ws1, r, 5, btxt, burl, fill=GREEN if ('NONE' in btxt or 'VERIFIED' in btxt) else (AMBER if ('UNKNOWN' in btxt.upper() or 'N/M' in btxt or 'MENTOR-ONLY' in btxt) else None))
    link(ws1, r, 6, chex, churl, fill=chexfill(chex))
    link(ws1, r, 7, ews, eurl, fill=chexfill(ews))
    link(ws1, r, 8, card, curl, fill=GREEN if card.startswith('Y') else RED)
    link(ws1, r, 9, lim, lurl)
    link(ws1, r, 10, nodoc, nurl, fill=GREEN if nodoc.startswith('Y') else (RED if nodoc.startswith('N ') or nodoc == 'N' else AMBER))
    put(ws1, r, 11, suw)
    put(ws1, r, 12, apply_)
    put(ws1, r, 13, assets)
    put(ws1, r, 14, action, font=BOLD)
ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:N{len(T1)+1}'
widths1 = [5, 26, 24, 26, 34, 22, 18, 18, 30, 26, 24, 12, 18, 32]
for i, wd in enumerate(widths1, 1): ws1.column_dimensions[get_column_letter(i)].width = wd
log('1. BEST TARGETS', f'NEW TAB: {len(T1)} ranked actionable institutions for Fontana 92336 / Long Beach 90802; every row carries clickable evidence links (bureau, Chex, EWS, card, limit, no-doc)', 'synthesis of v6 Funding Table + v7 bureau/geo/w2/icba research')

# =====================================================================
# TAB 2 — MENTOR BUREAU LIST
# =====================================================================
ws2 = wb.create_sheet('2. MENTOR BUREAU LIST')
put(ws2, 1, 1, "MENTOR'S BUREAU LIST — verbatim groups, kept in his ordering — vs our evidence (v7 2026-06-11). Full evidence: tab '8. BUREAU EVIDENCE'.", font=Font(bold=True, size=11))
header_row(ws2, 2, ['Mentor group', 'Institution', 'Mentor claim', 'Our verdict', 'Evidence summary (1 line)', 'Evidence link', 'Caveat / next step'])
mentor_order = [
    ('EXPERIAN', 'PNC'), ('EXPERIAN', 'Sunwest Bank'), ('EXPERIAN', 'Tustin Community Bank'),
    ('EQUIFAX', 'First Citizens Bank'), ('EQUIFAX', 'KeyBank'), ('EQUIFAX', 'California Credit Union'), ('EQUIFAX', 'First Western Trust Bank'),
    ('TRANSUNION', 'Banner Bank'), ('TRANSUNION', 'Valley Bank (Valley National)'), ('TRANSUNION', 'BMO'),
    ('TRANSUNION', 'Western Alliance Bank'), ('TRANSUNION', 'WaFd Bank'), ('TRANSUNION', 'Los Angeles FCU'),
    ('BRANCH VERIFY', 'California Bank & Trust'),
]
def verdict_fill(v):
    v = v.upper()
    if v.startswith('AGREES'): return GREEN
    if v.startswith('PARTIAL'): return AMBER
    if v.startswith('CONFLICTS'): return RED
    return AMBER  # NO-DATA
for i, (grp, inst) in enumerate(mentor_order):
    rec = bureau_by[inst]; r = i + 3
    put(ws2, r, 1, grp, font=BOLD)
    put(ws2, r, 2, inst)
    put(ws2, r, 3, rec['Mentor_claim'])
    v = rec['Mentor_verdict']
    vshort = ('AGREES' if v.upper().startswith('AGREES') else 'PARTIAL' if 'PARTIAL' in v.upper() else
              'CONFLICTS' if v.upper().startswith('CONFLICTS') else 'NO-DATA')
    put(ws2, r, 4, f'{vshort} — {v}' if vshort != v else vshort, fill=verdict_fill(vshort))
    ev = rec['New_evidence']
    put(ws2, r, 5, (ev[:240] + '...') if len(ev) > 240 else ev)
    urls = re.findall(r'https?://[^\s|]+', rec['Source_links'])
    link(ws2, r, 6, urls[0] if urls else rec['Source_links'][:80], urls[0] if urls else None)
    extra = ' | extra: ' + ' '.join(urls[1:3]) if len(urls) > 1 else ''
    put(ws2, r, 7, f"[{rec['Evidence_grade']}] {rec['Action']}{extra}")
ws2.freeze_panes = 'A3'
ws2.auto_filter.ref = f'A2:G{len(mentor_order)+2}'
for i, wd in enumerate([14, 26, 14, 26, 60, 40, 50], 1): ws2.column_dimensions[get_column_letter(i)].width = wd
log('2. MENTOR BUREAU LIST', "NEW TAB: mentor's 14 institutions verbatim in his EX/EQ/TU/branch-verify order × our verdict (AGREES/PARTIAL/CONFLICTS/NO-DATA) + evidence link", 'research/v7/out/bureau_pulls.csv + 5 findings MDs')

# =====================================================================
# TABS 5/6 — GEO
# =====================================================================
def build_geo(title, rows, note):
    ws = wb.create_sheet(title)
    cols = list(rows[0].keys())
    header_row(ws, 1, cols)
    for i, rec in enumerate(rows):
        r = i + 2
        hl = BLUEHL if i < 3 else None
        for j, k in enumerate(cols, 1):
            v = rec[k]
            if k == 'Source_links':
                link(ws, r, j, v, firsturl(v), fill=hl)
            else:
                put(ws, r, j, v, fill=hl)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}{len(rows)+1}'
    for i, wd in enumerate([6, 30, 7, 38, 9, 14, 18, 12, 12, 10, 24, 12, 50, 20][:len(cols)], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    log(title, note, 'research/v7/out/geo_*.csv (FDIC SOD 2025 + Overpass/OSM + web, 2026-06-11)')
build_geo('5. GEO - LONG BEACH', geo_lb, f'NEW TAB: {len(geo_lb)} institutions distance-sorted from 1 World Trade Ctr, Long Beach 90802; top-3 nearest highlighted; source links clickable')
build_geo('6. GEO - FONTANA', geo_fon, f'NEW TAB: {len(geo_fon)} institutions distance-sorted from Fontana 92336; top-3 nearest highlighted; source links clickable')

# =====================================================================
# TAB 7 — ICBA
# =====================================================================
ws7 = wb.create_sheet('7. ICBA COMMUNITY BANKS')
cols7 = list(icba[0].keys())
header_row(ws7, 1, cols7)
for i, rec in enumerate(icba):
    r = i + 2
    verdict = rec['Underwriter_verdict(+evidence)'].upper()
    vf = GREEN if 'SELF-ISSUER' in verdict else (RED if 'NO-GO' in verdict or 'NO CARD' in verdict or 'STALE' in verdict else AMBER)
    for j, k in enumerate(cols7, 1):
        v = rec[k]
        if k == 'Card_page_URL' and v:
            link(ws7, r, j, v, firsturl(v))
        elif k == 'Source_links':
            link(ws7, r, j, v, firsturl(v))
        elif k == 'Underwriter_verdict(+evidence)':
            put(ws7, r, j, v, fill=vf)
        elif k == 'Bank':
            put(ws7, r, j, v, fill=vf, font=BOLD)
        else:
            put(ws7, r, j, v)
ws7.freeze_panes = 'A2'
ws7.auto_filter.ref = f'A1:{get_column_letter(len(cols7))}{len(icba)+1}'
for i, wd in enumerate([24, 8, 22, 22, 18, 20, 14, 14, 40, 50, 40, 36, 10, 44], 1):
    ws7.column_dimensions[get_column_letter(i)].width = wd
log('7. ICBA COMMUNITY BANKS', f'NEW TAB: {len(icba)} ICBA-member banks with issuer verdicts color-coded (self-issuer=green; TCM/Elan/TIB/no-card=red)', 'research/v7/out/icba_community_banks.csv (ICBA locator + FDIC, 2026-06-11)')

# =====================================================================
# TAB 8 — BUREAU EVIDENCE
# =====================================================================
ws8 = wb.create_sheet('8. BUREAU EVIDENCE')
put(ws8, 1, 1, 'BUREAU EVIDENCE (v7 2026-06-11) — NOTE: No CA-specific BUSINESS pull datapoint exists publicly for most institutions — evidence GRADES matter. '
               'Grades: VERIFIED-disclosure > COMMUNITY-DP > CONSUMER-PROXY > MENTOR-ONLY > UNKNOWN/NO-DATA.', font=Font(bold=True, size=11))
cols8 = list(bureau[0].keys())
header_row(ws8, 2, cols8)
for i, rec in enumerate(bureau):
    r = i + 3
    for j, k in enumerate(cols8, 1):
        v = rec[k]
        if k == 'Source_links':
            link(ws8, r, j, v, firsturl(v))
        elif k == 'Evidence_grade':
            put(ws8, r, j, v, fill=GREEN if v.startswith('VERIFIED') else (AMBER if 'MENTOR-ONLY' in v or 'UNKNOWN' in v else None))
        else:
            put(ws8, r, j, v)
ws8.freeze_panes = 'A3'
ws8.auto_filter.ref = f'A2:{get_column_letter(len(cols8))}{len(bureau)+2}'
for i, wd in enumerate([24, 14, 26, 60, 16, 22, 22, 34, 56], 1):
    ws8.column_dimensions[get_column_letter(i)].width = wd
log('8. BUREAU EVIDENCE', f'NEW TAB: bureau_pulls.csv in full ({len(bureau)} rows) with grade colors + clickable primary sources + header caveat', 'research/v7/out/bureau_pulls.csv + bureau_pulls.md')

# =====================================================================
# TAB 12 — IBANKNET fill
# =====================================================================
ib = wb['12. IBANKNET CA']
ib_row = {ib.cell(r, 1).value: r for r in range(2, ib.max_row + 1)}
ib_count = 0
for rec in afill:
    if rec['Sheet'] != 'ibanknet': continue
    r = ib_row[rec['Row_key']]
    for col, val in ((5, rec['Assets_value']), (6, rec['Deposits_value'])):
        if not val: continue
        old = ib.cell(r, col).value
        if old in BLANKISH:
            ib.cell(r, col).value = val
        else:
            ib.cell(r, col).value = f'{old}{MARK}{val} total ({rec["As_of_date"]})'
    note = rec['Note']
    src = f"v7 {TODAY}: {rec['Source']} [{rec['Match_method']}]" + (f' — {note[:150]}' if note else '')
    old = ib.cell(r, 11).value
    ib.cell(r, 11).value = src if old in BLANKISH else f'{old}{MARK}{src}'
    ib_count += 1
log('12. IBANKNET CA cols Assets/Deposits/Source_note', f'FILLED/ENRICHED {ib_count} rows with FDIC/NCUA 2026-03-31 totals + sources (existing SOD-derived values preserved, totals appended)', 'research/v7/out/assets_fill.csv (87 ibanknet rows)')

# =====================================================================
# TAB 0 — README
# =====================================================================
ws0 = wb.create_sheet('0. README', 0)
readme = [
    ('CA BUSINESS CREDIT & FUNDING — v7 (2026-06-11)', Font(bold=True, size=14), None),
    ('One workbook, simplicity-first. Start at tab 1. Everything blue+underlined is a CLICKABLE source link.', SMALL, None),
    ('', None, None),
    ('VERSION HISTORY: v5 → v6 (2026-06-10, EWS/Chex map + censuses + ibanknet) → v7 (2026-06-11, INTEGRATOR-2: numbered tabs, BEST TARGETS sheet, mentor list front-and-center, bureau evidence, geo distances, assets fills). ADD-ONLY: all v6 content carried; nothing deleted.', SMALL, None),
    ('', None, None),
    ('COLOR KEY (4 colors):', BOLD, None),
    ('  GREEN = confirmed favorable (verified evidence)', SMALL, GREEN),
    ('  AMBER = unknown — verify before relying on it', SMALL, AMBER),
    ('  RED = blocker (e.g. Chex pull confirmed, no product, NO-GO issuer)', SMALL, RED),
    ('  BLUE underline = clickable source link', LINKFONT, None),
    ('', None, None),
    ('TABS:', BOLD, None),
    ('  1. BEST TARGETS — THE sheet: 32 ranked actionable institutions for Fontana 92336 / Long Beach 90802; bureau pulled, Chex, EWS, biz card, limit, no-doc — each claim linked.', SMALL, None),
    ("  2. MENTOR BUREAU LIST — mentor's EX/EQ/TU/branch-verify groups verbatim vs our verdict + evidence.", SMALL, None),
    ('  3. FUNDING TABLE MASTER — full v6 master (236 rows) + v7 enrichments (HP bureau, assets/deposits, merger flags, corrections) + 7 new rows (ranks 237–243).', SMALL, None),
    ('  4. DIAMOND-PLATINUM — tier shortlist; v7: FNBO bureau=EX (biz DP), Stanford EX membership pull, Rize CU added as Platinum candidate.', SMALL, None),
    ('  5./6. GEO - LONG BEACH / FONTANA — every branch within radius, distance-sorted, top-3 highlighted.', SMALL, None),
    ('  7. ICBA COMMUNITY BANKS — 21 ICBA members, issuer verdict color-coded.', SMALL, None),
    ('  8. BUREAU EVIDENCE — all 35 bureau-pull research rows + grades + links. No CA-specific BUSINESS pull DP exists publicly for most institutions — grades matter.', SMALL, None),
    ('  9. EWS-CHEX MAP — v6 deposit-screening evidence (carried unchanged).', SMALL, None),
    ('  10./11. FDIC SOD 6CTY / NCUA 6CTY — 6-county censuses (carried unchanged).', SMALL, None),
    ('  12. IBANKNET CA — CA institutions; v7 filled assets/deposits (87 rows).', SMALL, None),
    ('  13. CHANGE LOG — v6 log + one row per v7 change.', SMALL, None),
    ('  14.–16. LEGACY (v5) — Legend + 4-county censuses, byte-identical.', SMALL, None),
    ('', None, None),
    ('RULES: ADD-ONLY (v6 cells identical, appended-to with " ‖ v7 2026-06-11: ", or blank/N-M filled). Every factual claim keeps its date + source. ', SMALL, None),
    ('Evidence grades: VERIFIED-disclosure > COMMUNITY-DP > CONSUMER-PROXY > MENTOR-ONLY > INFERRED > UNKNOWN/NO-DATA. No fabrication: unverified = marked UNKNOWN/N-M.', SMALL, None),
    ('Where a cell has multiple sources, the PRIMARY is hyperlinked; extras live in the Sources column of that row.', SMALL, None),
]
for i, (txt, fnt, fill) in enumerate(readme, 1):
    c = ws0.cell(i, 1, txt)
    if fnt: c.font = fnt
    if fill: c.fill = fill
ws0.column_dimensions['A'].width = 160
log('0. README', 'NEW TAB: one-screen guide — tab map, 4-color key, version history, add-only + grading rules', 'integrator')

# =====================================================================
# TAB 13 — CHANGE LOG append
# =====================================================================
cl = wb['13. CHANGE LOG']
r = cl.max_row + 2
put(cl, r, 1, f'v7 {TODAY} CHANGES (INTEGRATOR-2) — ADD-ONLY vs ca_business_credit_funding_v6.xlsx; appends use " ‖ v7 {TODAY}: ":', font=BOLD)
for loc, change, src in changelog:
    r += 1
    put(cl, r, 1, loc); put(cl, r, 2, change); put(cl, r, 3, src); put(cl, r, 4, TODAY)
r += 1
put(cl, r, 1, 'deliverables/')
put(cl, r, 2, 'EXPORTED: ca_business_credit_funding_v7_best_targets.csv + ca_business_credit_funding_v7_funding_table.csv (values incl. URLs in parentheses)')
put(cl, r, 3, 'integrator'); put(cl, r, 4, TODAY)

# =====================================================================
# REORDER tabs
# =====================================================================
order = ['0. README', '1. BEST TARGETS', '2. MENTOR BUREAU LIST', '3. FUNDING TABLE MASTER',
         '4. DIAMOND-PLATINUM', '5. GEO - LONG BEACH', '6. GEO - FONTANA', '7. ICBA COMMUNITY BANKS',
         '8. BUREAU EVIDENCE', '9. EWS-CHEX MAP', '10. FDIC SOD 6CTY', '11. NCUA 6CTY',
         '12. IBANKNET CA', '13. CHANGE LOG', '14. LEGEND (v5)', '15. CENSUS 4CTY FDIC (v5)',
         '16. CENSUS 4CTY NCUA (v5)']
wb._sheets = [wb[t] for t in order]
wb.active = 1
wb.save(V7)
print('Saved', V7, 'tabs:', wb.sheetnames)

# =====================================================================
# TASK 12 — CSV exports
# =====================================================================
wb2 = openpyxl.load_workbook(V7)
def export(sheet, path):
    ws = wb2[sheet]
    with open(path, 'w', newline='', encoding='utf-8') as f:
        wr = csv.writer(f)
        for row in ws.iter_rows():
            out = []
            for c in row:
                v = '' if c.value is None else str(c.value)
                if c.hyperlink and c.hyperlink.target:
                    v = f'{v} ({c.hyperlink.target})'
                out.append(v)
            wr.writerow(out)
    print('Exported', path)
export('1. BEST TARGETS', os.path.join(OUT_DIR, 'ca_business_credit_funding_v7_best_targets.csv'))
export('3. FUNDING TABLE MASTER', os.path.join(OUT_DIR, 'ca_business_credit_funding_v7_funding_table.csv'))
print('NEW ROWS:', n_new, '| HP bureau enriched:', hp_count, '| assets fills funding:', ad_count, '| ibanknet fills:', ib_count)
