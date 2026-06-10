#!/usr/bin/env python3
"""v6 Part 2: paren cleanup + 22 NEW rows (task 5) + autofilter extension (task 10)."""
import json, openpyxl
from openpyxl.styles import PatternFill, Font

BASE = '/home/user/awake/research/v6'
WB = f'{BASE}/deliverables/ca_business_credit_funding_v6.xlsx'
D = '2026-06-10'

F = lambda c: PatternFill('solid', start_color=c)
MINT, GOLDC, CREAM, GRAY = F('00A9D08E'), F('00FFD966'), F('00FFF2CC'), F('00D9D9D9')
T_NAT, T_REG, T_CU, T_CA = F('00BDD7EE'), F('008EAADB'), F('00CCC0DA'), F('00E2EFDA')
U_SELF, U_ELAN, U_UNK, U_NONE = F('00D9EAD3'), F('00F8CBAD'), F('00EFEFEF'), F('00D9D9D9')
ORANGE = F('00F8CBAD')

wb = openpyxl.load_workbook(WB)
ws = wb['Funding Table']
changelog = json.load(open(f'{BASE}/changelog_part1.json'))

# ---- cleanup: rebalance parens in cells composed in part 1 (cosmetic only, same info)
fixed = 0
for r in range(2, 215):
    for c in (7, 8):
        v = ws.cell(r, c).value
        if isinstance(v, str) and f'; v6 {D})' in v and v.count('(') != v.count(')'):
            diff = v.count('(') - v.count(')')
            if diff > 0:
                idx = v.rfind(' (DOC') if ' (DOC' in v else v.rfind(' (COMMUNITY') if ' (COMMUNITY' in v else v.rfind(' (VERIFIED')
                if idx > 0:
                    ws.cell(r, c).value = v[:idx] + ')' * diff + v[idx:]
                    fixed += 1
print('paren fixes:', fixed)

# ---------------------------------------------------------------- TASK 5: NEW ROWS (rows 215-236, Rec Rank 214-235)
H = {h: i + 1 for i, h in enumerate([ws.cell(1, c).value for c in range(1, 37)])}
NM = 'N/M'

def newrow(rownum, rank, vals, name_fill, type_fill, uw_fill, chex_fill=None, ews_fill=None):
    base = {'Rec Rank': str(rank), 'Rec Pick #': str(rank), 'ChexSystems (deposit)': NM,
            'EWS (deposit)': NM, 'Mentor List': None, 'Deposits': '—', 'Assets': '—',
            'Asset Rank': '—', 'HP Bureau': '—', 'Reports To (biz bureaus)': '—',
            'Other No-Doc Funding': '—', 'Deposit Acct?': '—', 'Biz Card Link': '—'}
    base.update(vals)
    for h, col in H.items():
        cell = ws.cell(rownum, col)
        cell.value = base.get(h, '—')
        if h == 'Bank':
            cell.font = Font(b=True); cell.fill = name_fill
        elif h == 'Meets Gold?':
            cell.fill = name_fill
        elif h == 'Type':
            cell.fill = type_fill
        elif h == 'Underwriter':
            cell.fill = uw_fill
        elif h == 'ChexSystems (deposit)' and chex_fill:
            cell.fill = chex_fill
        elif h == 'EWS (deposit)' and ews_fill:
            cell.fill = ews_fill
    changelog.append(dict(tab='Funding Table', row=rownum, inst=base['Bank'], field='(entire row)',
                          old='', new=f"NEW ROW rank {rank}: {base['Best For'][:120]}",
                          source=base.get('_src', '—'), date=D))

R = 215
rows = []

rows.append((214, dict(_src='w2_banks_screen.csv (E) + gold_leads.csv (#5)',
 **{'Bank': 'FNBO (First National Bank of Omaha)', 'Type': 'National',
 'SoCal Priority': 'National (online — no CA branches needed)',
 'Meets Gold?': '3/4  ✓SUW ✓Onl ✗SoCal ✓Card — 💎 Diamond candidate',
 'Card Category': 'A — Self-UW (online)',
 'Est. Credit Limit': '$50,000 stated MAX (VERIFIED fnbo.com 2026-06-10); instant DPs $7.1K–$14K; prequal accurate + soft-pull  [myFICO #6403479/#6332103/#6690470]',
 'CA Footprint': 'Zero CA branches — true online-national issuer (the textbook directory miss)',
 'CA Presence': 'Online to CA (pure card issuer, no branch needed)', 'Self-UW?': 'Yes',
 'Business Card?': '✅ Evergreen Business Edition MC (flagship); legacy Business Edition Visa via agent portals; Business Edition Secured ($2K–$10K self-set)',
 'Underwriter': 'Self (“Cards are issued by First National Bank of Omaha” — fnbo.com 2026-06-10; also operates agent programs for other banks)',
 'Apply': 'Online', 'How to Apply': 'Online national — fnbo.com Apply Now; run pre-qual (soft pull) FIRST',
 'Website': 'https://www.fnbo.com/small-business/credit-cards',
 'Biz Card Link': 'https://www.card.fnbo.com/mpp/fi/offer/business/pb-visa',
 'HP Bureau': 'N/M — inquiry-sensitive; does NOT combine same-day pulls',
 'Reports To (biz bureaus)': 'Personal: negatives-only. Business: D&B + Experian Business (DoC business-card-reporting list; myFICO #6458865)',
 'Approval Info & Limits (+link)': '$50K max stated on site; DPs: $7.1K instant / $13.9K prequal honored at $13.9K / $15K prequal common ($7K–$14K instant band). CLI: small=soft, large=hard — reps disclose which (DoC), but one instant-HP CLI DP exists (myFICO #6695146): ASK FIRST. Realistic: $10K–$15K start, grow to $50K via CLIs.',
 'No-Doc Fit (/5)': '4', 'PG / Tax Returns': 'PG yes; NO tax returns/financials at application for standard limits',
 'Deposit Acct?': 'No',
 'Best For': '★ Top v6 add — self-issued $50K-max online-national card, personal-bureau-quiet (💎 Diamond candidate; Chex/EWS UNKNOWN)',
 'Card Products': 'Evergreen Business Edition Mastercard; Business Edition Visa (legacy/agent portals); Business Edition Secured',
 'Region / Metro': 'National — online to CA', 'City (HQ)': 'Omaha, NE', 'Asset Rank': 'Natl',
 'Confidence / Date': 'NEW v6 2026-06-10 — HIGH $50K/self-issuer/online (fnbo.com); MED CLI behavior (mixed DPs); COMMUNITY-DP limits 2022–2025',
 'Sources': 'fnbo.com/small-business/credit-cards (issuer + $50K max, 2026-06-10); ficoforums.myfico.com #6403479 #6332103 #6690470 #6458865 #6695146 (accessed 2026-06-10); doctorofcredit.com/which-business-credit-cards-report/'}),
 MINT, T_NAT, U_SELF))

rows.append((215, dict(_src='w2_cus_screen.csv (H) + gold_leads.csv (#8)',
 **{'Bank': 'Stanford Federal Credit Union', 'Type': 'Credit Union',
 'SoCal Priority': 'Other — NorCal HQ, REMOTE join from SoCal (FOPAL/MOAH)',
 'Meets Gold?': '3/4  ✓SUW ✓Onl(remote) ✗SoCal ✓Card — ⭐ start 6-mo membership clock NOW',
 'ChexSystems (deposit)': 'N/M (pull type unconfirmed — DoC checking-bonus pages note it; sfcu.org 403s scrapers)',
 'Card Category': 'B — Self-UW (form app via Virtual Branch/branch)',
 'Est. Credit Limit': 'Business Rewards Visa up to $50,000 + unsecured Business LOC up to $100,000 (both VERIFIED sfcu.org snippets 2026-06-10); personal DP $7.5K @ 672 EX  [myFICO #6036159]',
 'Assets': '$4.8B', 'CA Footprint': 'Palo Alto HQ (NCUA charter 13392); no SoCal branches — Virtual Branch Zoom serves remote members',
 'CA Presence': "CA-HQ'd (NorCal)", 'Self-UW?': 'Yes',
 'Business Card?': '✅ Business Rewards Visa: to $50K; 1pt/$, 2pts dining/delivery; NO AF/BT/CA/FTX fees; 16.00–17.99% var',
 'Underwriter': 'Self (VERIFIED — sfcu.org hosts own “Business Visa Application and Solicitation Disclosure”; no Elan/TCM/FNBO anywhere)',
 'Apply': 'Online (remote)', 'How to Apply': 'Join online sfcu.org/join ($5 share, SFCU pays nonprofit dues); ID-verify via Virtual Branch Zoom; biz card via Business Loan Application form (not instant online)',
 'Website': 'https://www.sfcu.org', 'Biz Card Link': 'https://www.sfcu.org/business/business-credit-card',
 'Approval Info & Limits (+link)': 'GATES (quoted from sfcu.org 2026-06-10): min 640 FICO + 2 years in business + 6 months SFCU membership before the biz card → OPEN THE $5 MEMBERSHIP NOW to start the clock. Docs: Business Loan Application + Personal Financial Statement. No public biz-card DPs (thin DP base). Caveat: remote BUSINESS-account onboarding INFERRED via Virtual Branch, not page-verified (NASA-FCU precedent — call first).',
 'No-Doc Fit (/5)': '3', 'PG / Tax Returns': 'PG implied via PFS; tax returns not mentioned for the card (LOC likely fuller doc); 640 FICO min',
 'Other No-Doc Funding': 'Unsecured Business LOC to $100K (above = collateral); equipment/term; SBA 504',
 'Deposit Acct?': 'Membership: anyone via FOPAL (Friends of the Palo Alto Library) or MOAH; any CA resident joins free (SFCU pays dues — DoC/maximizingmoney 2026-06-10); remote join online',
 'Best For': '⭐ THE v6 CU find — best self-issued $50K biz card + $100K unsecured LOC reachable remotely from SoCal (Platinum)',
 'Card Products': 'Business Rewards Visa (to $50K)', 'Region / Metro': 'NorCal / Palo Alto (remote-serve)',
 'City (HQ)': 'Palo Alto', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — HIGH product/limits/gates (multi-snippet VERIFIED); MED remote biz-account opening (INFERRED — confirm by phone)',
 'Sources': 'sfcu.org/business/business-credit-card + /business-banking/loans/ + /join-us/ (via 2026-06-10 search snapshots; site 403s bots); doctorofcredit.com Stanford FCU checking bonus; maximizingmoney.com; myFICO #6036159; ncuso.org charter 13392'}),
 MINT, T_CU, U_SELF))

rows.append((216, dict(_src='w2_banks_screen.csv (D)',
 **{'Bank': 'UMB Bank N.A.', 'Type': 'Regional',
 'SoCal Priority': 'Other — Central CA branches; SoCal = remote/online only',
 'Meets Gold?': '3/4  ✓SUW ✓Onl ✗SoCal ✓Card',
 'Card Category': 'A — Self-UW (online)',
 'Est. Credit Limit': 'N/M — no published max, no forum DPs found (2026-06-10)',
 'Assets': '~$68B (post-Heartland)', 'CA Footprint': '8 CA branches, ALL Central CA: Fresno, SLO ×2, Paso Robles, Salinas, Mariposa, Groveland, Oakhurst (ex-Premier Valley via Heartland acq closed 2025-01-31; FDIC API 2026-06-10)',
 'CA Presence': 'CA branches (Central) + online to SoCal', 'Self-UW?': 'Yes',
 'Business Card?': '✅ UMB Visa Signature Business Rewards ($200/20K-pt bonus @ $2K/90d); commercial card suite',
 'Underwriter': 'Self (“issued by UMB Bank n.a., Member FDIC” — umb.com 2026-06-10; UMB also runs cobrand/agent programs for OTHER banks)',
 'Apply': 'Online', 'How to Apply': 'Online dao.umb.com/sbla (15–20 min, 2–3 day decision) or branch/RM',
 'Website': 'https://www.umb.com/business-banking/business-credit-cards',
 'Biz Card Link': 'https://www.umb.com/business-banking/business-credit-cards',
 'Approval Info & Limits (+link)': 'Low-doc-ISH, not no-doc: online app asks gross sales / net profit / assets & liabilities + formation docs (operating agreement/bylaws); no tax returns stated (umb.com app FAQ 2026-06-10). No limit DPs found.',
 'No-Doc Fit (/5)': '3', 'PG / Tax Returns': 'PG; financials-questionnaire-style app; no tax returns stated',
 'Deposit Acct?': 'No (not stated as required)',
 'Best For': '★ Self-UW + online national app; Gold-tier (limits unknown — ask)',
 'Card Products': 'UMB Visa Signature Business Rewards; commercial cards',
 'Region / Metro': 'Central CA branches; KC-HQ', 'City (HQ)': 'Kansas City, MO', 'Asset Rank': 'Reg',
 'Confidence / Date': 'NEW v6 2026-06-10 — HIGH issuer/app channel; LOW limits (no DPs)',
 'Sources': 'umb.com/business-banking/business-credit-cards (2026-06-10); FDIC API CERT 8273 locations; SEC 8-K Heartland close 2025-01-31'}),
 MINT, T_REG, U_SELF))

rows.append((217, dict(_src='w2_cus_screen.csv (I)',
 **{'Bank': 'Valley Strong Credit Union', 'Type': 'Credit Union',
 'SoCal Priority': 'Other — Central Valley HQ; OPEN join from Fontana/Long Beach via FFA',
 'Meets Gold?': '3/4  ✓SUW ✓Onl ✗SoCal ✓Card — 🚩 Gold candidate, phone-screen next',
 'Card Category': 'A — Self-UW (online)',
 'Est. Credit Limit': 'N/M — “high spending limits” marketing only; limits/docs unpublished (site 403s scrapers; zero forum DPs) → PHONE SCREEN',
 'Assets': '$4.1B', 'CA Footprint': 'Bakersfield HQ; county FOM (Fresno/Kern/Kings/Madera/Mariposa/Merced/Solano/San Joaquin/SLO/Stanislaus/Tulare) does NOT cover SB/LA — but FFA path does',
 'CA Presence': "CA-HQ'd (Central Valley)", 'Self-UW?': 'Yes (INFERRED-strong)',
 'Business Card?': '✅ Business Platinum (“high spending limits”) + Business Rewards (1% cash back, $0 AF, employee cards)',
 'Underwriter': 'Self (INFERRED-strong — CFPB agreement DB lists “Valley Strong Credit Union” as issuer of its own Visa agreements, checked 2026-06-10; no agent branding)',
 'Apply': 'Online', 'How to Apply': 'Online membership ($5 share or Business Share Savings); card app via site/branch',
 'Website': 'https://www.valleystrong.com', 'Biz Card Link': 'https://www.valleystrong.com/business-platinum-credit-card',
 'Approval Info & Limits (+link)': 'Join: “All members of the Financial Fitness Association are eligible” — Valley Strong opens your FFA membership and covers year 1 ($8/yr after). Bonus DPs: $50 share + $250 Business Share Savings (maximizingmoney live 2026-06-10). Limits/docs/TIB all UNKNOWN — best non-Stanford flag; needs one phone call.',
 'No-Doc Fit (/5)': '—', 'PG / Tax Returns': 'N/M — not published',
 'Deposit Acct?': 'Membership via FFA (open); Business Share Savings $5 open',
 'Best For': '🚩 Gold candidate — open FFA join + 2 self-issued biz cards (CFPB-corroborated); limits unpublished',
 'Card Products': 'Business Platinum CC; Business Rewards CC', 'Region / Metro': 'Central Valley / Bakersfield',
 'City (HQ)': 'Bakersfield', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED join path + card lineup; INFERRED-strong self-UW (CFPB DB); UNKNOWN limits',
 'Sources': 'valleystrong.com/business-platinum-credit-card + /business-rewards-credit-card (2026-06-10); consumerfinance.gov/credit-cards/agreements/issuer/valley-strong-credit-union/; maximizingmoney.com valley-strong'}),
 MINT, T_CU, U_SELF))

rows.append((218, dict(_src='w2_banks_screen.csv (F) + gold_leads.csv (#3)',
 **{'Bank': 'Ramp (corporate charge card)', 'Type': 'National',
 'SoCal Priority': 'National (online fintech — lends in CA)',
 'Meets Gold?': '2/4  ✗SUW ✓Onl ✗SoCal ✓Card — EIN-only / NO-PG class',
 'ChexSystems (deposit)': 'N/A — no consumer deposit screening (fintech KYC; vendors unknown)',
 'EWS (deposit)': 'N/A — no consumer deposit screening',
 'Card Category': 'C — Other agent card (fintech partner-bank)',
 'Est. Credit Limit': 'Dynamic — % of linked business cash (~10–20%); needs ≥$25K in a US business bank account; $50K+ routine with ~$250K+ balance; no fixed max',
 'CA Footprint': 'None — online-national fintech', 'CA Presence': 'Online to CA',
 'Self-UW?': 'No (partner-bank issued)',
 'Business Card?': '✅ Ramp Visa Corporate/Commercial CHARGE card (pay-in-full — not a revolver) + Ramp Flex; free spend-mgmt software',
 'Underwriter': 'Celtic Bank (Corporate Card) + Sutton Bank (Commercial Card) — ramp.com issuing-bank schedule, corroborated NerdWallet/Airwallex (2026-06-10)',
 'Apply': 'Online', 'How to Apply': 'Online (minutes) — connect business bank account (Plaid)',
 'Website': 'https://ramp.com', 'Biz Card Link': 'https://ramp.com',
 'HP Bureau': 'NONE — no personal credit check, no personal pull',
 'Reports To (biz bureaus)': 'No personal reporting (no PG); business reporting UNKNOWN/limited',
 'Approval Info & Limits (+link)': 'EIN-only, NO PG, NO personal credit pull; underwriting = live bank-balance link, not documents. Registered entities only (LLC/C-corp/S-corp/LP) — sole props EXCLUDED. Honest framing: limit scales with real cash — lenient, not magic.',
 'No-Doc Fit (/5)': '4', 'PG / Tax Returns': 'NO PG; NO docs — live Plaid bank link IS the underwriting',
 'Deposit Acct?': 'No Ramp account needed; needs an external US business bank account ≥$25K',
 'Best For': '★ Costless to layer: no pull, no PG, no personal reporting (charge card; limit = % of cash)',
 'Card Products': 'Ramp Visa Corporate Card; Ramp Visa Commercial Card; Ramp Flex',
 'Region / Metro': 'National — online', 'City (HQ)': 'New York, NY', 'Asset Rank': '—',
 'Confidence / Date': 'NEW v6 2026-06-10 — HIGH (vendor terms verified + community comparisons)',
 'Sources': 'ramp.com/legal/issuing-bank-schedule; nerdwallet.com Ramp review; lendingtree.com Ramp review; glencoyne.com corporate-cards guide (all 2026-06-10/2026)'}),
 CREAM, T_NAT, U_UNK))

rows.append((219, dict(_src='w2_banks_screen.csv (J)',
 **{'Bank': 'Mercury (IO charge card)', 'Type': 'National',
 'SoCal Priority': 'National (online fintech — lends in CA)',
 'Meets Gold?': '2/4  ✗SUW ✓Onl ✗SoCal ✓Card — no-pull, cash-collared',
 'ChexSystems (deposit)': 'N/A — fintech KYC, no traditional Chex (COMMUNITY-DP/UNKNOWN formal)',
 'EWS (deposit)': 'N/A — fintech KYC',
 'Card Category': 'C — Other agent card (fintech partner-bank)',
 'Est. Credit Limit': 'Cash-underwritten: intro ~up to $5K with DAILY autopay; $15K+ kept at Mercury unlocks higher limits + monthly payment (mercury.com support 2026-06-10)',
 'CA Footprint': 'None — online-national fintech (SF-HQ company)', 'CA Presence': 'Online to CA',
 'Self-UW?': 'No (partner-bank issued)',
 'Business Card?': '✅ IO Mastercard corporate charge card (1.5% cashback class); no LOC',
 'Underwriter': 'Patriot Bank, N.A. (VERIFIED mercury.com/legal/io-charge-card-agreement 2026-06-10); deposits via Choice Financial Group + Column N.A.',
 'Apply': 'Online', 'How to Apply': 'Open Mercury business checking online; IO offered in-dashboard',
 'Website': 'https://mercury.com/credit', 'Biz Card Link': 'https://mercury.com/credit',
 'HP Bureau': 'NONE — no credit check, 0 personal-credit impact',
 'Reports To (biz bureaus)': 'No personal pull/reporting (VERIFIED); business reporting UNKNOWN/likely none',
 'Approval Info & Limits (+link)': 'NO credit check, NO PG; qualification = Mercury account balance, EIN/entity-based; registered entities (sole props generally excluded). Limit floats with balance — hostage to deposits parked at Mercury.',
 'No-Doc Fit (/5)': '4', 'PG / Tax Returns': 'NO PG; NO docs — balance-based',
 'Deposit Acct?': 'YES — REQUIRES Mercury checking (limit scales with balance kept there)',
 'Best For': '★ Easiest no-PG line if you already bank at Mercury',
 'Card Products': 'IO Mastercard (charge)', 'Region / Metro': 'National — online',
 'City (HQ)': 'San Francisco, CA', 'Asset Rank': '—',
 'Confidence / Date': 'NEW v6 2026-06-10 — HIGH (issuer agreement + support docs verified)',
 'Sources': 'mercury.com/credit; mercury.com/legal/io-charge-card-agreement; support.mercury.com “Qualifying for IO” (all 2026-06-10)'}),
 CREAM, T_NAT, U_UNK))

rows.append((220, dict(_src='w2_banks_screen.csv (G) + gold_leads.csv (#4)',
 **{'Bank': 'BILL Spend & Expense (Divvy)', 'Type': 'National',
 'SoCal Priority': 'National (online fintech — lends in CA)',
 'Meets Gold?': '2/4  ✗SUW ✓Onl ✗SoCal ✓Card — most accessible of the no-PG trio',
 'Card Category': 'C — Other agent card (fintech partner-bank)',
 'Est. Credit Limit': '$1K(–$500)–$5M scaled to revenue/cash; ~$20K+ cash balance typical to approve (Nav 2026-06-10); $20K–$50K DPs common for real operating businesses',
 'CA Footprint': 'None — online-national (BILL HQ San Jose CA; Divvy Draper UT)', 'CA Presence': 'Online to CA',
 'Self-UW?': 'No (partner-bank issued)',
 'Business Card?': '✅ BILL Divvy corporate charge card with assigned “Budget” credit line + free spend software',
 'Underwriter': 'Cross River Bank + WEX Bank (VERIFIED crossriver.com Divvy case study 2026-06-10; WebBank named in some older material)',
 'Apply': 'Online', 'How to Apply': 'Online (apply.divvy.co); link bank account (Plaid/MX)',
 'Website': 'https://www.bill.com/product/spend-and-expense', 'Biz Card Link': 'https://apply.divvy.co',
 'HP Bureau': 'No hard personal pull reported; PG status ambiguous (see PG cell)',
 'Reports To (biz bureaus)': 'SBFE → feeds D&B / Experian-biz / Equifax-biz (VERIFIED Nav 2026-06-10); no routine personal reporting',
 'Approval Info & Limits (+link)': 'No hard cash minimum — holistic cash-flow underwriting via bank link; no tax docs at application. ⚠️ PG CONFLICT: Nav 2026 “no PG required” vs Merchant Maverick “PG now required for some/smaller applicants” — get a fresh DP before counting on no-PG.',
 'No-Doc Fit (/5)': '4', 'PG / Tax Returns': 'PG CONFLICTING (Nav: no; Merchant Maverick: sometimes) — treat as PG-possible; no tax returns',
 'Deposit Acct?': 'No — links existing business bank account',
 'Best For': '★ No-PG trio member for bootstrapped businesses (PG ambiguity = verify first); builds SBFE file',
 'Card Products': 'BILL Divvy Card (weekly/semi-monthly/monthly pay cycles)',
 'Region / Metro': 'National — online', 'City (HQ)': 'San Jose, CA / Draper, UT', 'Asset Rank': '—',
 'Confidence / Date': 'NEW v6 2026-06-10 — MED-HIGH (PG conflict unresolved)',
 'Sources': 'nav.com/business-credit-card/bill-divvy-review/; crossriver.com/divvy-case-study; merchantmaverick.com BILL Divvy review; businesscreditworkshop.me (all accessed 2026-06-10)'}),
 CREAM, T_NAT, U_UNK))

rows.append((221, dict(_src='w2_banks_screen.csv (H) + gold_leads.csv (#7)',
 **{'Bank': "Sam's Club Business Mastercard (Synchrony)", 'Type': 'National',
 'SoCal Priority': '★ Core SoCal access (clubs across LA/OC/IE incl. near Fontana)',
 'Meets Gold?': '2/4  ✗SUW ✓Onl ✓SoCal ✓Card — no-PG ONLY via in-club EIN-only path',
 'Card Category': 'C — Other agent card (Synchrony co-brand)',
 'Est. Credit Limit': 'DPs: $8K SL instant NO-PG (“account type 2”, EIN-based) + $8K→$30K CLI history; $10K instants  [myFICO #6701656/#6716041, 2024–2025]',
 'CA Footprint': "Sam's Club locations throughout San Bernardino/Riverside/LA counties", 'CA Presence': 'In-club + online to CA',
 'Self-UW?': 'No (Synchrony co-brand)',
 'Business Card?': "✅ Sam's Club Business Mastercard: 5% gas, 3% dining, $0 AF w/ club membership (+ private-label store card)",
 'Underwriter': 'Synchrony Bank (VERIFIED synchrony.com/partner/sams-club 2026-06-10)',
 'Apply': 'In-club (for no-PG) / Online (PG)',
 'How to Apply': 'IN-CLUB kiosk/associate and insist on EIN-only → Synchrony checks BUSINESS credit only. ONLINE app FORCES PG + personal hard pull.',
 'Website': 'https://www.samsclub.com/credit', 'Biz Card Link': 'https://www.samsclub.com/credit',
 'HP Bureau': 'EIN-only path: none personal. PG path: personal HP (Synchrony, often TU — unverified this pass); ~700+ FICO',
 'Reports To (biz bureaus)': 'EIN-only path: business bureaus only (COMMUNITY-DP)',
 'Approval Info & Limits (+link)': "No-PG path is DP-backed but YMMV by associate + Synchrony discretion; generally requires an ESTABLISHED business credit file (D&B tradelines/revenue) — newer entities will sign a PG. Sam's Club membership required.",
 'No-Doc Fit (/5)': '3', 'PG / Tax Returns': 'No-PG possible (in-club EIN-only, established biz file); PG otherwise; no financials',
 'Deposit Acct?': "No (Sam's Club membership instead)",
 'Best For': '★ The in-club no-PG play — SoCal clubs everywhere; CLIs reported generous',
 'Card Products': "Sam's Club Business Mastercard; Sam's Club store card",
 'Region / Metro': 'National — clubs in LA/OC/IE/SD', 'City (HQ)': 'Issuer: Draper, UT (Synchrony CERT 27314)', 'Asset Rank': 'Natl',
 'Confidence / Date': 'NEW v6 2026-06-10 — MED-HIGH (no-PG path COMMUNITY-DP; tightens in downturns — re-pull DPs within 60 days of sprint)',
 'Sources': 'ficoforums.myfico.com #6701656 “NO PG … Instant Approved 8k SL” + #6716041 (accessed 2026-06-10); synchrony.com/partner/sams-club; truebuild.com + nav.com + wallethub.com Sam’s Club EIN articles'}),
 CREAM, T_NAT, U_UNK))

rows.append((222, dict(_src='w2_cus_screen.csv (L)',
 **{'Bank': 'Monterra Credit Union (ex-San Mateo CU)', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — Bay Area FOM (no SoCal path)',
 'Meets Gold?': '2/4  ✓SUW ✗Onl ✗SoCal ✓Card — Ref: FOM gated',
 'Card Category': 'B — Self-UW (DocuSign/PDF apps)',
 'Est. Credit Limit': 'Card limit UNKNOWN; Business LOC has a DISTINCT application tier for lines >$50K (implies lighter app below $50K — INFERRED)',
 'Assets': '~$1.4B', 'CA Footprint': 'San Mateo/Redwood City; FOM = San Mateo/Alameda/Contra Costa/Santa Clara/SF counties only (VERIFIED 2026-06-10), no association path',
 'CA Presence': "CA-HQ'd (Bay Area)", 'Self-UW?': 'Yes (INFERRED-strong)',
 'Business Card?': '✅ Business Rewards Visa: $0 AF, 1.99% intro APR 12 mo, then 14.49–18.00% var',
 'Underwriter': 'Self (INFERRED-strong — own intro pricing, branded/serviced by Monterra; no Elan/TCM/FNBO found)',
 'Apply': 'DocuSign/PDF', 'How to Apply': 'DocuSign/PDF applications; branch/online',
 'Website': 'https://www.monterra.org/business', 'Biz Card Link': 'https://www.monterra.org/business/credit-cards/business-rewards-visa-credit-card',
 'Approval Info & Limits (+link)': 'Attractive self-UW structure (rare biz 1.99% intro) but Bay-Area-only FOM blocks Fontana/Long Beach. Revisit only if owner gains a Bay Area nexus.',
 'No-Doc Fit (/5)': '—', 'PG / Tax Returns': 'UNKNOWN — separate apps by line size suggest lighter doc under $50K (INFERRED)',
 'Best For': 'Ref — nice structure, closed door (FOM)',
 'Card Products': 'Business Rewards Visa; Business LOC (>$50K separate tier); SBA',
 'Region / Metro': 'Bay Area', 'City (HQ)': 'Redwood City', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED products/FOM; INFERRED self-UW',
 'Sources': 'monterra.org/business/credit-cards/business-rewards-visa-credit-card + /business (2026-06-10)'}),
 CREAM, T_CU, U_SELF))

rows.append((223, dict(_src='w2_cus_screen.csv (M)',
 **{'Bank': 'Bay Federal Credit Union', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — Santa Cruz/San Benito/Monterey FOM (hard-closed)',
 'Meets Gold?': '2/4  ✓SUW ✗Onl ✗SoCal ✓Card — Ref: FOM hard-closed',
 'Card Category': 'B — Self-UW (contact/branch)',
 'Est. Credit Limit': 'Business Visa up to $50,000 (VERIFIED bayfed.com 2026-06-10)',
 'Assets': '~$1.6B', 'CA Footprint': 'Capitola HQ; members AND businesses must be physically in Santa Cruz/San Benito/Monterey Co; page confirms NO outside association path (VERIFIED 2026-06-10)',
 'CA Presence': "CA-HQ'd (Central Coast)", 'Self-UW?': 'Yes (INFERRED-strong)',
 'Business Card?': '✅ Business Visa to $50K: $0 AF, BayREWARDS, 2,500-pt signup, employee cards',
 'Underwriter': 'Self (INFERRED-strong — issued directly by Bay Federal per product page; no agent named, 2026-06-10)',
 'Apply': 'Contact/branch', 'How to Apply': 'Contact/branch (no online biz-card app posted)',
 'Website': 'https://www.bayfed.com/business/visa-credit', 'Biz Card Link': 'https://www.bayfed.com/business/visa-credit',
 'Approval Info & Limits (+link)': '$50K self-issued card but FOM hard-closed; note “certain LLCs and LLPs may not qualify” caveat on membership page.',
 'No-Doc Fit (/5)': '—', 'PG / Tax Returns': 'UNKNOWN — docs not published',
 'Best For': 'Ref — $50K self-issued card, closed FOM',
 'Card Products': 'Business Visa (to $50K)', 'Region / Metro': 'Central Coast', 'City (HQ)': 'Capitola', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED limit + FOM',
 'Sources': 'bayfed.com/business/visa-credit + /about-us/membership-eligibility (2026-06-10)'}),
 CREAM, T_CU, U_SELF))

rows.append((224, dict(_src='w2_cus_screen.csv (A) + ncua_gap_report.md',
 **{'Bank': 'Ventura County Credit Union (VCCU)', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — Ventura/Santa Barbara FOM only',
 'Meets Gold?': '2/4  ✓SUW ✗Onl ✗SoCal ✓Card — Ref: FOM gated',
 'Card Category': 'B — Self-UW (relationship/branch)',
 'Est. Credit Limit': 'N/M — no published limits; commercial side relationship-based',
 'Assets': '$1.47B (NCUA 2026Q1)', 'CA Footprint': 'Ventura HQ, 6 Ventura-Co branches (NCUA census); strongest of the v6 outlier-county gaps',
 'CA Presence': "CA-HQ'd (Ventura)", 'Self-UW?': 'Yes (INFERRED)',
 'Business Card?': '✅ Business Mastercard (employee cards) + Business LOC (revolving) + term/CRE/SBA',
 'Underwriter': 'Self (INFERRED — own-branded “Business Credit Cards from Mastercard”; no Elan/TCM/FNBO branding; agreement not posted)',
 'Apply': 'Branch/phone', 'How to Apply': 'Branch/phone 805.477.4000 via Commercial Lending Specialists; no online biz-card app found',
 'Website': 'https://www.vccuonline.net/Business', 'Biz Card Link': 'https://www.vccuonline.net/Business',
 'Approval Info & Limits (+link)': 'FOM = live/work/school Ventura or Santa Barbara County ONLY (VERIFIED 2026-06-10); no association path → un-joinable from Fontana/Long Beach. Park as reference.',
 'No-Doc Fit (/5)': '—', 'PG / Tax Returns': 'UNKNOWN — relationship-based, no published doc list',
 'Best For': 'Ref — real biz card+LOC, un-joinable from target addresses',
 'Card Products': 'Business Mastercard; Business LOC; term/CRE/SBA',
 'Region / Metro': 'Ventura County', 'City (HQ)': 'Ventura', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED products + FOM; INFERRED self-UW',
 'Sources': 'vccuonline.net/Business + /Business/Commercial-Loans (2026-06-10); NCUA census charter 68458'}),
 CREAM, T_CU, U_SELF))

rows.append((225, dict(_src='w2_cus_screen.csv (C) + ncua_gap_report.md',
 **{'Bank': 'MyPoint Credit Union (ex-PLCU)', 'Type': 'Credit Union',
 'SoCal Priority': 'Other — San Diego (FOM: SD or Riverside Co)',
 'Meets Gold?': '3/4  ✓SUW ✗Onl ✓SoCal ✓Card — Ref: FOM gated (Fontana + Long Beach both FAIL)',
 'Card Category': 'B — Self-UW (email/appointment)',
 'Est. Credit Limit': 'N/M — “flexible credit limits” only; no DPs',
 'Assets': '$701M (NCUA 2026Q1)', 'CA Footprint': 'San Diego HQ, 5 SD branches; FOM = live/work/worship San Diego OR Riverside County (VERIFIED 2026-06-10)',
 'CA Presence': "CA-HQ'd (San Diego)", 'Self-UW?': 'Yes (INFERRED)',
 'Business Card?': '✅ MyPoint Business Visa (flexible limits, $0 AF) + working-capital business loans + CRE (SD/Riv property only)',
 'Underwriter': 'Self (INFERRED — consumer Visa agreement creditor line = “MyPoint Credit Union”, VERIFIED 2026-06-10; zero Elan/TCM/FNBO/PSCU; biz agreement not posted)',
 'Apply': 'Email/phone/appt', 'How to Apply': 'businesslending@mypointcu.com / 1-888-495-3400 / appointment; biz accounts by appointment',
 'Website': 'https://www.mypointcu.com/borrow/business/business-loans', 'Biz Card Link': 'https://www.mypointcu.com/borrow/business/business-loans',
 'Approval Info & Limits (+link)': 'Fontana (San Bernardino Co) and Long Beach (LA Co) BOTH fail the FOM; only a Riverside-County WORK nexus rescues it. Revisit if owner has Riverside operations.',
 'No-Doc Fit (/5)': '—', 'PG / Tax Returns': 'UNKNOWN — not published',
 'Best For': 'Ref — decent self-issued biz Visa; FOM blocks both target addresses',
 'Card Products': 'MyPoint Business Visa', 'Region / Metro': 'San Diego', 'City (HQ)': 'San Diego', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED card + FOM; INFERRED self-UW',
 'Sources': 'mypointcu.com/borrow/business/business-loans + consumer-credit-card-agreement (2026-06-10); NCUA charter 68377'}),
 MINT, T_CU, U_SELF))

rows.append((226, dict(_src='w2_banks_screen.csv (C)',
 **{'Bank': 'Citizens Bank N.A. (Providence RI)', 'Type': 'National',
 'SoCal Priority': 'Ref — NOT CA-accessible (Private Bank channel only)',
 'Meets Gold?': '2/4  ✓SUW ✗Onl(CA gate) ✗SoCal ✓Card — Ref: NOT CA-accessible',
 'Card Category': 'A — Self-UW (online, footprint-gated)',
 'Est. Credit Limit': 'UNKNOWN — no solid public limit DPs (site bot-blocks; myFICO thin)',
 'CA Footprint': 'CA = PRIVATE BANK offices ONLY (SF 2024; LA/Beverly Hills/South Bay teams announced 2025-10-29) targeting HNW/UHNW — the $3.7B CA SOD deposits are Private Bank, not retail branches',
 'CA Presence': '⚠ NOT retail-CA — business products gated to CT/DE/FL/MD/MA/MI/NH/NJ/NY/OH/PA/RI/VT/VA/DC (footer-verified 2026-06-10)',
 'Self-UW?': 'Yes',
 'Business Card?': '✅ (in footprint only) Everyday Points Business MC ($0 AF, 3x pts, 50K-pt/$5K-90d bonus); Business Platinum MC (low intro APR)',
 'Underwriter': 'Self — Citizens Bank N.A. is issuer (VERIFIED WalletHub/citizensbank.com 2026-06-10; no Elan/agent markers)',
 'Apply': 'Online/phone (gated)', 'How to Apply': 'Online or 888-333-5145 — address-gated to 15-state footprint; CA biz likely rejected at address gate unless via Private Bank RM',
 'Website': 'https://www.citizensbank.com/small-business/business-credit-cards/overview.aspx',
 'Approval Info & Limits (+link)': 'RE-SCREEN conclusion (w2_banks 2026-06-10): a normal SoCal small business still cannot get the card. Re-flag if retail CA branches open.',
 'No-Doc Fit (/5)': '—', 'PG / Tax Returns': 'Standard PG app; no no-doc evidence',
 'Best For': 'Ref — self-issued MCs but CA door = Private Bank RM only',
 'Card Products': 'Everyday Points Business Mastercard; Business Platinum Mastercard',
 'Region / Metro': 'East Coast footprint; CA Private Bank only', 'City (HQ)': 'Providence, RI', 'Asset Rank': 'Natl',
 'Confidence / Date': 'NEW v6 2026-06-10 — MED-HIGH (footprint verified; self-issue verified; CA-gate inferred; site 403s)',
 'Sources': 'investor.citizensbank.com 2025-10-29 private-bank LA release; citizensbank.com footer state list (via search 2026-06-10); wallethub.com Citizens biz card answers; FDIC CERT 57957'}),
 CREAM, T_NAT, U_SELF))

rows.append((227, dict(_src='w2_banks_screen.csv (I)',
 **{'Bank': 'Live Oak Bank', 'Type': 'National',
 'SoCal Priority': 'National (online — CA served remotely)',
 'Meets Gold?': 'Ref — NO biz card; SBA/LOC = FULL-DOC (great bank, wrong tool)',
 'Card Category': 'Commercial-only / No card',
 'Est. Credit Limit': 'Loans to $15M (SBA 7(a)/504/USDA/term); LOC limits unpublished',
 'CA Footprint': 'Branchless online-national (Wilmington NC; CERT 58665)', 'CA Presence': 'Online to CA',
 'Self-UW?': 'n/a (no card)',
 'Business Card?': '❌ NO business credit card (VERIFIED NerdWallet 2026 review)',
 'Underwriter': 'None for cards; lending = self (Live Oak Banking Company)',
 'Apply': 'Online + RM call', 'How to Apply': 'Online + dedicated lender/RM call',
 'Website': 'https://www.liveoak.bank/business/',
 'Approval Info & Limits (+link)': 'SBA Preferred Lender, #1 SBA 7(a) class; underwriting = full-doc (financials + tax returns) — ZERO no-doc angle. Actual draw: high-yield business savings + SBA pipeline.',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'FULL-DOC (financials, tax returns)',
 'Deposit Acct?': 'Business checking/savings/CDs (high-yield) — the real reason to use them',
 'Best For': 'Ref — deposit home + SBA pipeline, NOT a card/no-doc target',
 'Card Products': 'None', 'Region / Metro': 'National — online', 'City (HQ)': 'Wilmington, NC', 'Asset Rank': 'Natl',
 'Confidence / Date': 'NEW v6 2026-06-10 — HIGH',
 'Sources': 'liveoak.bank/business/; nerdwallet.com Live Oak review (2026); bankrate.com Live Oak review (2026-06-10)'}),
 GRAY, T_NAT, U_NONE))

rows.append((228, dict(_src='w2_cus_screen.csv (K)',
 **{'Bank': 'Meriwest Credit Union', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — joinable CA-wide (FFA) but Elan card',
 'Meets Gold?': 'Ref — ELAN NO-GO class (joinable, no unique value)',
 'Card Category': 'E — Elan agent card',
 'Est. Credit Limit': 'Elan-standard (typically conservative; not Meriwest-specific)',
 'Assets': '~$2B', 'CA Footprint': 'San Jose HQ; FOM = Bay Area 5-county + Pima AZ; FFA complimentary membership for anyone in CA/AZ (VERIFIED 2026-06-10)',
 'CA Presence': "CA-HQ'd (Bay Area); joinable from SoCal", 'Self-UW?': 'No',
 'Business Card?': '✅ but agent: “issued by Elan Financial Services” (VERIFIED meriwest.com 2026-06-10)',
 'Underwriter': 'Elan (“issued by Elan Financial Services” — meriwest.com business credit cards page, 2026-06-10) → same Elan book as every other Elan CU (myaccountaccess)',
 'Apply': 'Online (Elan app)', 'How to Apply': 'Online via Elan application',
 'Website': 'https://www.meriwest.com/business-services/business-credit-cards',
 'Approval Info & Limits (+link)': 'Joinable from Fontana/Long Beach via FFA, but the card adds NOTHING vs any other Elan CU — deprioritized per project Elan rule.',
 'No-Doc Fit (/5)': '—', 'PG / Tax Returns': 'Elan-standard (PG, no tax returns)',
 'Best For': 'Ref — skip for cards (Elan)', 'Card Products': 'Elan business cards',
 'Region / Metro': 'Bay Area', 'City (HQ)': 'San Jose', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED Elan',
 'Sources': 'meriwest.com/business-services/business-credit-cards + /join-meriwest (2026-06-10)'}),
 CREAM, T_CU, U_ELAN))

rows.append((229, dict(_src='w2_cus_screen.csv (D)',
 **{'Bank': 'Wheelhouse Credit Union', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — San Diego (open join, empty shelf)',
 'Meets Gold?': 'Ref — open join but NO biz products (likely wound down)',
 'Card Category': 'Commercial-only / No card',
 'Est. Credit Limit': 'n/a — no business products',
 'Assets': '$457M (NCUA 2026Q1)', 'CA Footprint': 'San Diego, 4 SD branches; join from ANYWHERE via SD Zoo Wildlife Alliance / SD Automotive Museum / Arbor Day Foundation ~$10 (VERIFIED wheelhousecu.com/join 2026-06-10)',
 'CA Presence': "CA-HQ'd (San Diego)", 'Self-UW?': 'n/a',
 'Business Card?': '❌ none visible on site (2026-06-10); “PPP lender/business loans” claims survive only on chamber listings',
 'Underwriter': 'None (no business products found)',
 'Apply': 'n/a', 'How to Apply': 'n/a (personal join online)',
 'Website': 'https://www.wheelhousecu.com',
 'Approval Info & Limits (+link)': 'Trivially joinable from SoCal but nothing to fund with — skip for business credit.',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'n/a',
 'Best For': 'Ref — open door, empty shelf', 'Card Products': 'None (personal Visa Rewards only)',
 'Region / Metro': 'San Diego', 'City (HQ)': 'San Diego', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED join path + no-biz-products',
 'Sources': 'wheelhousecu.com/join + site nav (2026-06-10); NCUA charter 61003'}),
 GRAY, T_CU, U_NONE))

rows.append((230, dict(_src='w2_cus_screen.csv (B)',
 **{'Bank': 'OceanAir Federal Credit Union (ex-CBC FCU)', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — Oxnard/Ventura (joinable, no card)',
 'Meets Gold?': 'Ref — joinable but NO biz card / NO LOC',
 'Card Category': 'Commercial-only / No card',
 'Est. Credit Limit': 'n/a — business credit limited to CRE loans',
 'Assets': '$871M (NCUA 2026Q1)', 'CA Footprint': 'Oxnard HQ, 6 Ventura-Co branches (charter 5110 INFERRED); join page: “If you’re reading this, you can become an OceanAir member!” (VERIFIED 2026-06-10); nationwide via American Consumer Council (secondary, INFERRED)',
 'CA Presence': "CA-HQ'd (Ventura)", 'Self-UW?': 'n/a',
 'Business Card?': '❌ no biz credit card, no LOC — Epic Business Checking + CRE loans + MONA processing only',
 'Underwriter': 'None (no biz card)',
 'Apply': 'n/a', 'How to Apply': 'Membership online (oceanairfcu.myori.com); CRE via contact form/phone',
 'Website': 'https://oceanair.org',
 'Approval Info & Limits (+link)': 'Joinable but nothing to fund with (cards). Join page lists only ID/SSN/DOB — no Chex language (UNKNOWN).',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'n/a',
 'Best For': 'Ref — possible deposit home; skip for biz credit', 'Card Products': 'None',
 'Region / Metro': 'Ventura County', 'City (HQ)': 'Oxnard', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED no-card; INFERRED ACC join path',
 'Sources': 'oceanair.org/join + /business-overview (2026-06-10); creditunionmatch.com; NCUA charter 7608 (gap report)'}),
 GRAY, T_CU, U_NONE))

rows.append((231, dict(_src='w2_cus_screen.csv (J)',
 **{'Bank': 'Coast Central Credit Union', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — Eureka/Humboldt (FOM closed + full-doc)',
 'Meets Gold?': '2/4  ✓SUW ✗Onl ✗SoCal ✓Card — Ref: FOM closed + FULL-DOC paper app',
 'Card Category': 'B — Self-UW (paper app)',
 'Est. Credit Limit': 'UNKNOWN — “Requested Credit Limit” free-entry; loan officer sets',
 'Assets': '$2.3B', 'CA Footprint': 'Eureka HQ; FOM = Del Norte/Humboldt/Trinity only (+ Cal Poly Humboldt / College of the Redwoods) (VERIFIED 2026-06-10)',
 'CA Presence': "CA-HQ'd (North Coast)", 'Self-UW?': 'Yes (VERIFIED)',
 'Business Card?': '✅ Business Visa Classic/Gold (up to 4 employee cards/app) — but paper app rev 10/04',
 'Underwriter': 'Self (VERIFIED from app PDF: issued “pursuant to Coast Central Credit Union VISA Credit Card Agreement”)',
 'Apply': 'Paper', 'How to Apply': 'Paper application PDF to branch',
 'Website': 'https://www.coastccu.org', 'Biz Card Link': 'https://www.coastccu.org/wp-content/uploads/2017/05/BusinessVisa.pdf',
 'Approval Info & Limits (+link)': 'FULL-DOC verbatim from app: “Remit with current Business Tax Return and Profit/Loss Statement & Balance Sheet” + PFS per owner. Un-joinable from SoCal anyway.',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'FULL-DOC: biz tax return + P&L + balance sheet + PFS',
 'Best For': 'Ref — self-UW but full-doc paper + closed FOM', 'Card Products': 'Business Visa Classic/Gold',
 'Region / Metro': 'North Coast', 'City (HQ)': 'Eureka', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED (primary doc)',
 'Sources': 'coastccu.org BusinessVisa.pdf (parsed 2026-06-10); coastccu.org membership pages'}),
 CREAM, T_CU, U_SELF))

rows.append((232, dict(_src='w2_cus_screen.csv (E)',
 **{'Bank': 'Cabrillo Credit Union', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — San Diego (no biz line)',
 'Meets Gold?': 'Ref — consumer-only; FOM = SD County/SEGs (confirmed cheaply)',
 'Card Category': 'Commercial-only / No card',
 'Est. Credit Limit': 'n/a',
 'Assets': '$572M (NCUA 2026Q1)', 'CA Footprint': 'San Diego, 4 SD branches; FOM = live/work SD County + Border Patrol retirees/federal/Sharp/Carlsbad SEGs (VERIFIED 2026-06-10)',
 'CA Presence': "CA-HQ'd (San Diego)", 'Self-UW?': 'n/a',
 'Business Card?': '❌ none — consumer lineup only (checking, personal cards, personal loans, HELOC)',
 'Underwriter': 'None (no business banking section)',
 'Apply': 'n/a', 'How to Apply': 'n/a',
 'Website': 'https://www.cabrillocu.com',
 'Approval Info & Limits (+link)': 'No business products and no SoCal join path for target addresses — skip.',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'n/a',
 'Best For': 'Ref — skip (confirmed cheap)', 'Card Products': 'None',
 'Region / Metro': 'San Diego', 'City (HQ)': 'San Diego', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED (cheap confirm)',
 'Sources': 'cabrillocu.com/About-Us/Become-a-Member (2026-06-10); NCUA charter 68409'}),
 GRAY, T_CU, U_NONE))

rows.append((233, dict(_src='w2_cus_screen.csv (F)',
 **{'Bank': 'San Diego Firefighters FCU', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — San Diego (occupational FOM)',
 'Meets Gold?': 'Ref — FOM gated (SD-County fire service only); personal products only',
 'Card Category': 'Commercial-only / No card',
 'Est. Credit Limit': 'n/a',
 'Assets': '$152M (NCUA 2026Q1)', 'CA Footprint': 'San Diego, 1 branch; FOM strictly fire-service personnel assigned within SD County (VERIFIED sdffcu.org 2026-06-10)',
 'CA Presence': "CA-HQ'd (San Diego)", 'Self-UW?': 'n/a',
 'Business Card?': '❌ none — personal Visa, personal LOC, auto/home only',
 'Underwriter': 'None (no business services)',
 'Apply': 'n/a', 'How to Apply': 'n/a ($25 savings opens membership — if eligible)',
 'Website': 'https://www.sdffcu.org',
 'Approval Info & Limits (+link)': 'Occupational CU, no biz products — skip.',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'n/a',
 'Best For': 'Ref — skip', 'Card Products': 'None',
 'Region / Metro': 'San Diego', 'City (HQ)': 'San Diego', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED',
 'Sources': 'sdffcu.org/membership + /member-services (2026-06-10); NCUA charter 24110'}),
 GRAY, T_CU, U_NONE))

rows.append((234, dict(_src='w2_cus_screen.csv (G) + ibanknet_gap_report.md',
 **{'Bank': 'Educational Employees CU (Fresno) — EECU', 'Type': 'Credit Union',
 'SoCal Priority': 'Ref — Central Valley FOM',
 'Meets Gold?': 'Ref — FOM gated (Central Valley); no biz card found',
 'Card Category': 'Commercial-only / No card',
 'Est. Credit Limit': 'n/a',
 'Assets': '$5.4B (ibanknet 2026Q1; site says $4.6B+)', 'CA Footprint': 'Fresno HQ; FOM = Fresno/Madera/Tulare/Kings + central-valley counties (VERIFIED 2026-06-10); no association path. ⚠ myeecu.org — NOT eecu.org (EECU Fort Worth TX)',
 'CA Presence': "CA-HQ'd (Central Valley)", 'Self-UW?': 'n/a',
 'Business Card?': '❌ business ACCOUNTS exist (open in branch); biz credit card NOT found on site',
 'Underwriter': 'None found (no biz card)',
 'Apply': 'In branch', 'How to Apply': 'Business accounts opened in branch',
 'Website': 'https://www.myeecu.org',
 'Approval Info & Limits (+link)': 'Big ($5.4B) but consumer-centric and un-joinable from SoCal — reference only.',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'n/a',
 'Best For': 'Ref — largest un-joinable v6 gap; disambiguation flag (myeecu.org)', 'Card Products': 'None (consumer cards only)',
 'Region / Metro': 'Central Valley / Fresno', 'City (HQ)': 'Fresno', 'Asset Rank': 'CU',
 'Confidence / Date': 'NEW v6 2026-06-10 — VERIFIED FOM; card absence INFERRED (not on site)',
 'Sources': 'myeecu.org/about/become-a-member (2026-06-10); ibanknet CA census 2026Q1'}),
 GRAY, T_CU, U_NONE))

rows.append((235, dict(_src='w2_banks_screen.csv (A) + fdic_gap_report.md',
 **{'Bank': 'Monet Bank (ex-Beal Bank SSB)', 'Type': 'Regional',
 'SoCal Priority': 'Ref — San Diego branch (deposit-only)',
 'Meets Gold?': 'Ref — deposit-only shell; FDIC-SOD gap CLOSED',
 'Card Category': 'Commercial-only / No card',
 'Est. Credit Limit': 'n/a — zero credit products',
 'Assets': '~$3.17B', 'CA Footprint': '2 CA points: San Diego (8880 Rio San Diego Dr Ste 103, 92108; $52.5M deposits) + Walnut Creek (FDIC API 2026-06-10). The single NEW_GAP from the FDIC SOD 6-county census.',
 'CA Presence': 'CA branches (deposit-taking only)', 'Self-UW?': 'n/a',
 'Business Card?': '❌ neither card nor LOC — product line = CDs / money market / savings ONLY (VERIFIED monet.bank 2026-06-10)',
 'Underwriter': 'None (no card program)',
 'Apply': 'n/a', 'How to Apply': 'n/a — nothing to apply for (deposits only)',
 'Website': 'https://www.monet.bank',
 'Approval Info & Limits (+link)': 'Renamed Beal Bank SSB (Andy Beal, Plano TX; CERT 32574) — historically a loan-acquisition/CD-funding specialist, not a business-credit issuer. Logged so we never re-chase.',
 'No-Doc Fit (/5)': '0', 'PG / Tax Returns': 'n/a',
 'Best For': 'Ref — gap closed, nothing a business can borrow on', 'Card Products': 'None',
 'Region / Metro': 'San Diego (branch); Plano TX HQ', 'City (HQ)': 'Plano, TX', 'Asset Rank': 'Reg',
 'Confidence / Date': 'NEW v6 2026-06-10 — HIGH (deposit-only confirmed cheaply)',
 'Sources': 'monet.bank (2026-06-10); FDIC API CERT 32574 (locations/history); fdic_gap_report.md'}),
 GRAY, T_REG, U_NONE))

for i, (rank, vals, namef, typef, uwf) in enumerate(rows):
    newrow(R + i, rank, vals, namef, typef, uwf)
print(f'Task 5: {len(rows)} new rows written (rows {R}-{R+len(rows)-1}, Rec Rank 214-{213+len(rows)})')

# ---------------------------------------------------------------- TASK 10: extend autofilter (freeze stays D2)
last = R + len(rows) - 1
ws.auto_filter.ref = f'A1:AJ{last}'
print('autofilter:', ws.auto_filter.ref, '| freeze:', ws.freeze_panes)

wb.save(WB)
json.dump(changelog, open(f'{BASE}/changelog_part2.json', 'w'))
print('changelog entries so far:', len(changelog))
