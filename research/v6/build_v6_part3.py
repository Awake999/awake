#!/usr/bin/env python3
"""v6 Part 3: 💎 Diamond-Platinum tab (task 6), data-dump tabs (task 7),
v6 Change Log tab (task 8), Legend appends (task 9)."""
import csv, json, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE = '/home/user/awake/research/v6'
WB = f'{BASE}/deliverables/ca_business_credit_funding_v6.xlsx'
D = '2026-06-10'
F = lambda c: PatternFill('solid', start_color=c)
HDRFILL, HDRFONT = F('001F4E78'), Font(b=True, color='00FFFFFF')
GOLD, SILVER, CREAM = F('00FFD966'), F('00D9D9D9'), F('00FFF2CC')
WRAP = Alignment(wrap_text=True, vertical='top')

wb = openpyxl.load_workbook(WB)
changelog = json.load(open(f'{BASE}/changelog_part2.json'))

# ---------------------------------------------------------------- TASK 6: Diamond-Platinum tab (right after Funding Table)
dp = wb.create_sheet('💎 Diamond-Platinum Tiers', index=1)
dp.sheet_properties.tabColor = 'FFD966'
widths = [6, 30, 22, 26, 30, 30, 26, 46, 40, 26, 46]
for i, w in enumerate(widths, 1):
    dp.column_dimensions[get_column_letter(i)].width = w

COLS = ['Rank', 'Institution', 'Tier', 'Self-UW?', '$25–50K no-doc?', 'No Chex/EWS?', 'CA / SoCal access?',
        "What's missing / to confirm", 'Next action', 'Est. limit', 'Sources']

def title(r, text, fill=None, font=None):
    c = dp.cell(r, 1, text)
    c.font = font or Font(b=True, sz=12)
    if fill: c.fill = fill
    return r + 1

def hdr(r):
    for i, h in enumerate(COLS, 1):
        c = dp.cell(r, i, h); c.fill = HDRFILL; c.font = HDRFONT; c.alignment = WRAP
    return r + 1

def row(r, vals, fill=None):
    for i, v in enumerate(vals, 1):
        c = dp.cell(r, i, v); c.alignment = WRAP
        if fill: c.fill = fill
        if i == 2: c.font = Font(b=True)
    return r + 1

r = 1
r = title(r, f'💎 DIAMOND / 🥈 PLATINUM TIERS — v6 {D} (verified-evidence-only; companions: ews_chex_report.md, enterprise_bt_dossier.md, lafcu_dossier.md, w2 screens)')
r += 1
r = title(r, 'SECTION 1 — 💎 DIAMOND criteria: self-underwritten + $25–50K no-doc EVIDENCE + NO Chex/EWS (no pull AND no report) + CA-accessible', GOLD)
r = title(r, f'STRICT DIAMOND QUALIFIERS TODAY: 0 (zero). No institution in the 235-row table has positive no-Chex AND no-EWS evidence on BOTH axes (ews_chex_report.md §2, {D}). Absence of data is NOT evidence of no-pull — manual-onboarding banks often run QualiFile quietly. Each candidate below needs one banker question: “Do you screen deposit applicants through ChexSystems or Early Warning?”', None, Font(b=True, color='009C0006'))
r = title(r, 'CLOSEST CANDIDATES (ranked by distance to Diamond):', None, Font(b=True))
r = hdr(r)
r = row(r, ['D1', 'FNBO (First National Bank of Omaha)', '💎 Diamond candidate',
            '✅ VERIFIED — “Cards are issued by FNBO” (fnbo.com 2026-06-10)',
            '✅ $50K stated MAX on site; no financials at app for standard limits (DPs start $7–15K)',
            '❓ UNKNOWN — zero Chex/EWS DPs either way',
            '✅ online-national, no branch needed',
            'Chex/EWS behavior; realistic path to the $50K ceiling is CLI-driven',
            'Run prequal (soft pull); ask Chex/EWS question if opening any FNBO deposit product',
            '$50K max; $10–15K realistic start',
            'fnbo.com/small-business/credit-cards; myFICO #6403479 #6332103 #6690470; w2_banks_screen (E)'], GOLD)
r = row(r, ['D2', 'Enterprise Bank & Trust', '💎 Diamond candidate (Platinum verdict)',
            '✅ VERIFIED HIGH — $3.155M card loans on own call report LNCRCD 2026-03-31 (agent banks hold $0) + FIS rails + 0 Elan tells',
            '⚠️ UNVERIFIED — ~$50K broker-channel claim only; zero public DPs; no published app-only tier (SBA Express ≤$500K = plausible vehicle)',
            '❓ N/M — datapoint desert; absent from DoC Chex+EWS lists (re-checked 2026-06-10)',
            '✅ VERIFIED — 8 SoCal branches LA/OC/SD (FDIC SOD 2025, CERT 27237)',
            'Written program terms: limit, app-only docs, pull type, PG, deposit requirement',
            'Call/email Jessica Morrow, AVP Business Banker — jmorrow@enterprisebank.com / (833) 896-2850; get the box IN WRITING',
            '~$50K (claimed, UNVERIFIED)',
            'enterprise_bt_dossier.md (2026-06-10); api.fdic.gov LNCRCD CERT 27237; theorg.com Jessica Morrow'], GOLD)
r = row(r, ['D3', 'Baxter Credit Union (BCU) — table rank 71', '💎 Diamond candidate (EWS-clean)',
            '(see Funding Table row — not the binding question)',
            '(limits per table row)',
            '✅/❓ EWS-CLEAN VERIFIED: no report + no inquiry (DoC comment 2026-05-22 #2326911); Chex UNVERIFIED',
            '(per table row)',
            'ONE verification call (Chex at membership) from Diamond-tier status',
            'Call membership desk: “Do you pull or report ChexSystems?” — then re-tier',
            '—',
            'ews_chex_report.md §2; doctorofcredit.com EWS comments #2326911 (2026-05-22)'], GOLD)
r = row(r, ['D4', 'Skyla Credit Union — table rank 79', '💎 Diamond candidate (EWS-clean)',
            '(see Funding Table row)', '(per table row)',
            '✅/❓ EWS-CLEAN VERIFIED: no report + no inquiry (same 2026-05-22 DP); Chex UNVERIFIED',
            '(per table row)',
            'Same single Chex verification call as BCU',
            'Call membership desk re Chex; then re-tier', '—',
            'ews_chex_report.md §2; DoC EWS comments #2326911 (2026-05-22)'], GOLD)
r = row(r, ['D5', 'Ramp / Mercury IO (no-pull fintech class)', '💎-adjacent — DIFFERENT ANIMAL',
            '✗ partner-bank issued (Celtic/Sutton; Patriot)',
            '⚠️ limits are cash-underwritten percentages of balance, not unsecured $25–50K credit',
            '✅ no personal pull, no PG, no personal reporting (verified vendor terms 2026-06-10)',
            '✅ online-national',
            'Nothing to confirm — they simply solve a different problem (charge cards collared to cash)',
            'Layer them anyway: costless (no pull, no PG); not a substitute for a $50K revolver',
            '% of business cash (Ramp ~10–20%; Mercury $5K intro)',
            'w2_banks_screen (F, J); ramp.com; mercury.com (2026-06-10)'], GOLD)
r += 1
r = title(r, 'SECTION 2 — 🥈 PLATINUM criteria (user): self-UW + $20K+ business card + no Chex/EWS + SoCal-accessible (near Fontana 92335 / Long Beach 90802)', SILVER)
r = title(r, 'Note: as with Diamond, NO institution passes the no-Chex/EWS leg with positive evidence — ranked by closeness; the Chex/EWS column shows the honest state.', None, Font(i=True))
r = hdr(r)
r = row(r, ['P1', 'Stanford Federal Credit Union', '🥈 Platinum',
            '✅ VERIFIED (own Business Visa disclosure on sfcu.org)',
            '✅ $50K biz Visa + $100K unsecured LOC (verified 2026-06-10); docs = Business Loan App + PFS (light for the limit, not no-doc)',
            '❓ UNKNOWN — pull type unconfirmed (site 403s)',
            '✅ remote: FOPAL/MOAH join, Virtual Branch Zoom; no SoCal branch needed',
            'Remote BUSINESS-account onboarding (INFERRED only — NASA-FCU precedent); Chex/EWS at join; 640 FICO / 2-yr TIB / 6-mo membership seasoning gates',
            '⏱ START THE 6-MONTH MEMBERSHIP CLOCK NOW ($5, SFCU pays dues); call to confirm remote biz onboarding for a Fontana/Long Beach LLC',
            '$50K card / $100K LOC',
            'w2_cus_screen (H); sfcu.org snapshots 2026-06-10'], SILVER)
r = row(r, ['P2', 'Enterprise Bank & Trust', '🥈 Platinum',
            '✅ VERIFIED HIGH (balance-sheet card loans)',
            '⚠️ $50K UNVERIFIED (broker claim; RM channel can flex — they control their own credit box)',
            '❓ N/M — no DPs either way',
            '✅ 8 SoCal branches (Cerritos/DTLA/Pasadena/Alhambra/Anaheim/SJC/Chula Vista/Encinitas)',
            'Same as D2: written terms from the banker',
            'RM channel: jmorrow@ / (833) 896-2850 — no online app exists',
            '~$50K (claimed)',
            'enterprise_bt_dossier.md (2026-06-10)'], SILVER)
r = row(r, ['P3', 'UMB Bank N.A.', '🥈 Platinum candidate',
            '✅ VERIFIED (“issued by UMB Bank n.a.”)',
            '⚠️ low-doc-ish (financials questionnaire, no tax returns stated); limits UNPUBLISHED',
            '❓ UNKNOWN',
            '✅ remote online app (dao.umb.com/sbla); CA branches are Central-CA only',
            'Actual limits (no DPs anywhere); Chex/EWS behavior',
            'Apply online or call an RM; ask limit band + bureaus first',
            'UNKNOWN',
            'w2_banks_screen (D); umb.com 2026-06-10'], SILVER)
r = row(r, ['P4', "Sam's Club Business MC — in-club EIN-only path", '🥈 Platinum candidate (no-PG angle)',
            '✗ Synchrony co-brand (not self-UW — included for the no-PG/no-personal-pull angle)',
            '⚠️ DPs $8K–$10K instant, $30K via CLIs — below $20K at approval, grows past it',
            '✅ for the EIN-only path: NO personal pull at all (business file only)',
            '✅ clubs across LA/OC/IE incl. near Fontana',
            'Needs an ESTABLISHED business credit file (D&B); YMMV by associate',
            'Build D&B tradelines first (Ramp/Divvy/Mercury feed SBFE/D&B), then apply IN-CLUB EIN-only',
            '$8K–$10K start → $30K CLIs',
            'w2_banks_screen (H); myFICO #6701656'], SILVER)
r += 1
r = title(r, 'NEAR-MISSES — exact failing leg:', None, Font(b=True))
r = hdr(r)
r = row(r, ['N1', 'LAFCU (table rank 12, GOLD 4/4)', 'Near-miss',
            '✅ VERIFIED (creditor verbatim)', '✅ $50K advertised max (re-verified 2026-06-10; zero approval DPs)',
            '❌ FAILS — “ChexSystems and hard-hit on your credit will be run” at MEMBERSHIP (verbatim, 2 live pages, 2026-06-10); unavoidable',
            '✅ 8 LA-County branches; but business join = IN-BRANCH only + biz must be in SoCal',
            'Nothing fixes the Chex leg; also 2-yr TIB + conservative human UW + inquiry-sensitive',
            'Use only with clean Chex + low inquiries + 2yr+ business',
            '$50K max (advertised)', 'lafcu_dossier.md (2026-06-10)'], CREAM)
r = row(r, ['N2', 'Valley Strong CU', 'Near-miss (Gold candidate)',
            '✅ INFERRED-strong (CFPB agreement DB)', '❓ limits UNPUBLISHED — could be $20K+, nobody knows',
            '❓ UNKNOWN', '✅ open FFA join from Fontana/Long Beach',
            'Limits, docs, TIB gate, Chex/EWS — all one phone screen away',
            'PHONE SCREEN NEXT: card limits + doc stack + Chex at join',
            'UNKNOWN', 'w2_cus_screen (I)'], CREAM)
r = row(r, ['N3', 'BILL Spend & Expense (Divvy)', 'Near-miss (fintech)',
            '✗ Cross River/WEX issued', '✅ $20K–$50K lines common, no tax docs',
            '✅ no traditional Chex/EWS', '✅ online-national',
            'PG status CONFLICTING (Nav: no PG vs Merchant Maverick: sometimes) — needs a fresh DP',
            'Apply after parking real balance; document whether PG requested',
            '$20K–$50K common', 'w2_banks_screen (G)'], CREAM)
r = row(r, ['N4', 'Citizens Bank N.A. (Providence)', 'Near-miss (geo-blocked)',
            '✅ VERIFIED self-issued MCs', '❓ limits unknown',
            '❓ unverified', '❌ FAILS — business products gated to 15 East-Coast states; CA = Private Bank offices only (2026-06-10)',
            'Retail CA expansion (none announced)', 'Re-flag if retail CA branches open',
            'UNKNOWN', 'w2_banks_screen (C)'], CREAM)
for i in range(1, r):
    dp.row_dimensions[i].height = None
changelog.append(dict(tab='💎 Diamond-Platinum Tiers', row='(new tab)', inst='—', field='—', old='',
                      new='NEW TAB: Diamond (0 strict qualifiers + 5 ranked candidates) + Platinum (4) + near-misses (4)',
                      source='ews_chex_report.md; enterprise_bt_dossier.md; lafcu_dossier.md; w2 screens; gold_leads', date=D))
print('Task 6: Diamond-Platinum tab done,', r - 1, 'rows')

# ---------------------------------------------------------------- TASK 7: data-dump tabs
def dump_tab(name, csvpath, src):
    t = wb.create_sheet(name)
    rows = list(csv.reader(open(csvpath)))
    for ri, rw in enumerate(rows, 1):
        for ci, v in enumerate(rw, 1):
            c = t.cell(ri, ci, v)
            if ri == 1:
                c.fill = HDRFILL; c.font = HDRFONT; c.alignment = WRAP
    t.freeze_panes = 'A2'
    t.auto_filter.ref = f'A1:{get_column_letter(len(rows[0]))}{len(rows)}'
    for ci in range(1, len(rows[0]) + 1):
        ml = max(len(str(rows[ri][ci-1])) for ri in range(min(len(rows), 30)))
        t.column_dimensions[get_column_letter(ci)].width = min(max(ml + 2, 8), 42)
    changelog.append(dict(tab=name, row='(new tab)', inst='—', field='—', old='',
                          new=f'NEW TAB: {len(rows)-1} data rows + header dumped verbatim from {csvpath.split("/")[-1]}',
                          source=src, date=D))
    print(f'  dump tab {name!r}: {len(rows)} rows')
    return len(rows)

dump_tab('FDIC SOD 6cty 2025', f'{BASE}/out/fdic_sod_census_6cty_2025.csv', 'FDIC SOD API 2025 (retrieved 2026-06-10)')
dump_tab('NCUA CU 6cty 2026Q1', f'{BASE}/out/ncua_cu_census_6cty.csv', 'NCUA call-report-data-2026-03.zip (retrieved 2026-06-10)')
dump_tab('EWS-Chex Map v6', f'{BASE}/out/ews_chex_mapping.csv', 'EWS-CHEX agent mapping 2026-06-10')
dump_tab('ibanknet CA 2026Q1', f'{BASE}/out/ibanknet_ca_institutions.csv', 'ibanknet.com 2026Q1 (retrieved 2026-06-10)')

# ---------------------------------------------------------------- TASK 9: Legend appends
lg = wb['Legend']
start = lg.max_row + 2  # 53
legend_lines = [
    ('v6 ADDITIONS (2026-06-10) — ADD-ONLY release built on v5; nothing deleted or destructively overwritten.', Font(b=True, sz=12)),
    ('NEW-v6 markers: appended cell text begins " ‖ v6 2026-06-10: …"; cells that were blank/N/M were written directly and carry a (v6 2026-06-10) tag; new rows (Rec Rank 214–235) have Confidence/Date starting "NEW v6 2026-06-10 —". Full audit trail in the "v6 Change Log" tab.', None),
    ('New tabs: 💎 Diamond-Platinum Tiers (tier shortlists) · FDIC SOD 6cty 2025 (122 banks) · NCUA CU 6cty 2026Q1 (161 CUs) · EWS-Chex Map v6 (213-row determination map) · ibanknet CA 2026Q1 (451 institutions) · v6 Change Log.', None),
    ('💎 DIAMOND tier = self-underwritten + $25–50K no-doc evidence + NO Chex/EWS (no pull AND no report) + CA-accessible. Strict qualifiers today: ZERO — see the Diamond-Platinum tab for ranked closest candidates (FNBO, Enterprise B&T, BCU/Baxter, Skyla, Ramp/Mercury-class).', PatternFill('solid', start_color='00FFD966')),
    ('🥈 PLATINUM tier = self-UW + $20K+ business card + no Chex/EWS + SoCal-accessible (Fontana 92335 / Long Beach 90802). Closest: Stanford FCU (start 6-mo membership clock NOW), Enterprise B&T (RM channel), UMB (remote), Sam\'s Club in-club EIN-only path. LAFCU fails ONLY on membership-stage Chex+hard-pull.', PatternFill('solid', start_color='00D9D9D9')),
    ('Chex/EWS sequencing (v6, from ews_chex_report.md §4): open Chex-SENSITIVE accounts FIRST while the file is empty (CU SoCal 3/6mo!, Premier America, Bank of Hope, MidFirst, Fremont…); inquiry-tolerant Chex-pullers mid-sequence (Comerica, Hanmi, Axos, Tech CU…); EWS-only big banks LAST — except open PNC BEFORE stacking Chase/BofA/Wells/Citi EWS inquiries (EWS cannot be frozen). Chex freeze blocks Chex-only banks (Citi approves frozen); EWS-clean parking: BCU/Baxter, Skyla, CommunityAmerica (+ Grasshopper biz).', None),
]
for i, (txt, fnt) in enumerate(legend_lines):
    c = lg.cell(start + i, 1, txt)
    if isinstance(fnt, Font):
        c.font = fnt
    elif isinstance(fnt, PatternFill):
        c.fill = fnt
changelog.append(dict(tab='Legend', row=f'{start}-{start+len(legend_lines)-1}', inst='—', field='—', old='',
                      new='APPENDED v6 section: marker convention, new-tab list, Diamond/Platinum definitions, Chex/EWS sequencing summary', source='ews_chex_report.md', date=D))
print('Task 9: Legend appended at rows', start, '-', start + len(legend_lines) - 1)

# ---------------------------------------------------------------- TASK 8: v6 Change Log tab
cl = wb.create_sheet('v6 Change Log')
cl.sheet_properties.tabColor = '1F4E78'
W = [5, 22, 10, 34, 22, 50, 70, 40, 12]
for i, w in enumerate(W, 1):
    cl.column_dimensions[get_column_letter(i)].width = w
r = 1
def clline(text, font=None, fill=None):
    global r
    c = cl.cell(r, 1, text)
    c.font = font or Font(b=True)
    if fill: c.fill = fill
    r += 1
clline(f'v6 CHANGE LOG — built {D} by INTEGRATOR agent. ADD-ONLY vs ca_business_credit_funding_v5.xlsx (all v5 tabs/rows/cells preserved; appends use " ‖ v6 {D}: ").', Font(b=True, sz=12))
r += 1
clline('DISCOVERY ENGINES RUN (v6, 2026-06-10):')
clline('• FDIC SOD 2025, 6-county census: 122 distinct banks / 2,939 branch records; all 113 prior 4-county CERTs re-confirmed (zero attrition); 1 NEW_GAP = Monet Bank (ex-Beal Bank SSB, San Diego) — screened, deposit-only. [fdic_gap_report.md]', Font())
clline('• NCUA 2026Q1, 6-county census: 161 unique CUs (140 = exact 4-county match to v5); 18 NEW_GAPs all in outlier counties (Ventura/SD), of which 6 are vendor-address ARTIFACTS (Symitar/Corelation/CO-OP hubs, no real SoCal presence). [ncua_gap_report.md]', Font())
clline('• ibanknet.com BACK UP: 451 CA-relevant institutions captured 2026Q1 (364 CA-HQ + 58 OOS-with-CA-branches + 29 Foreign Banking Organizations); 292 IN_V5, 158 NEW_GAP — gaps mostly NorCal/Central-Valley CUs; the 29 FBOs are EXCLUDED by project rule. [ibanknet_gap_report.md]', Font())
r += 1
clline('EWS / CHEX SEQUENCING STRATEGY (ews_chex_report.md §4, condensed):')
for t in ['1. Open Chex-SENSITIVE accounts FIRST while the Chex file is empty: CU SoCal (3/6mo — first or never), Premier America (3/90d), Bank of Hope (<5/12), MidFirst (7/12), Fremont (10/12), California CU, Patelco, Provident CU, SAFE CU, Wescom, Banner. One window supports ~2–3 of these.',
          '2. Middle: inquiry-TOLERANT Chex-pullers — Comerica (80+/24 OK), Hanmi (10+/12), Axos (50+/12), Tech CU (25+/12), US Bank, BMO, Citi (approves with FROZEN Chex), Tri Counties (soft), SBI, City National.',
          '3. EWS-only big banks LAST or anytime (Chase/BofA/Wells) — EXCEPT open PNC BEFORE stacking their EWS inquiries: PNC is the main EWS-sensitive bank and EWS inquiries (~1 yr) CANNOT be frozen.',
          '4. Chex security freeze blocks Chex-only banks (some auto-deny frozen; Citi approves frozen per 2026 DPs); the freeze does NOTHING at EWS banks. MyIDCare monitors EWS (imperfectly), not Chex.',
          '5. Business vs personal: Chase biz no-pull (×4); Truist biz in-branch skipped Chex; Citizens biz often skips owner Chex; BMO biz reports to separate BUSINESS Chex file — prefer business-account paths where funding needs deposits.',
          '6. EWS-clean parking (no footprint on either bureau, pending Chex confirm): BCU/Baxter, Skyla, CommunityAmerica (+ Grasshopper for biz).']:
    clline(t, Font())
r += 1
clline(f'CELL/ROW/TAB CHANGES ({len(changelog)} entries follow):')
hdr_row = r
for i, h in enumerate(['#', 'Tab', 'Row', 'Institution / Item', 'Field (column)', 'Old value (first 120 ch)', 'New / appended (summary)', 'Source file(s)', 'Date'], 1):
    c = cl.cell(hdr_row, i, h); c.fill = HDRFILL; c.font = HDRFONT
r += 1
for n, e in enumerate(changelog, 1):
    vals = [str(n), e['tab'], str(e['row']), str(e['inst'])[:60], e['field'], e['old'], e['new'], e['source'][:120], e['date']]
    for i, v in enumerate(vals, 1):
        cl.cell(r, i, v).alignment = WRAP
    r += 1
cl.freeze_panes = f'A{hdr_row+1}'
cl.auto_filter.ref = f'A{hdr_row}:I{r-1}'
print('Task 8: Change Log tab with', len(changelog), 'change entries + summary header')

wb.save(WB)
print('Saved. Final tab order:', wb.sheetnames)
