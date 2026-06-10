#!/usr/bin/env python3
"""v6 INTEGRATOR — Part 1: enrich Funding Table (tasks 1-4) and log changes.
ADD-ONLY: substantive cells get ' ‖ v6 2026-06-10: ' appends; blank/N/M cells may be written.
"""
import csv, json, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

BASE = '/home/user/awake/research/v6'
WB = f'{BASE}/deliverables/ca_business_credit_funding_v6.xlsx'
D = '2026-06-10'
MARK = f' ‖ v6 {D}: '   # ' ‖ v6 2026-06-10: '

ORANGE = PatternFill('solid', start_color='00F8CBAD')
GOLD   = PatternFill('solid', start_color='00FFD966')

wb = openpyxl.load_workbook(WB)
ws = wb['Funding Table']

# rank -> row map
rank2row = {}
for r in range(2, ws.max_row + 1):
    rank2row[str(ws.cell(r, 1).value).strip()] = r

changelog = []  # dicts: tab,row,inst,field,old,new,source,date
HDR = [ws.cell(1, c).value for c in range(1, 37)]

def is_nm(v):
    s = ('' if v is None else str(v)).strip()
    return s in ('', 'N/M', '—', '-', '–') or s.startswith('N/M')

def upd(row, col, newtext, source, fill=None, force_append=False):
    """Add-only update: append to substantive cells, write blank/N/M cells."""
    cell = ws.cell(row, col)
    old = cell.value
    inst = ws.cell(row, 3).value
    if is_nm(old) and not force_append:
        cell.value = newtext if f'v6 {D}' in newtext or D in newtext else f'{newtext} (v6 {D})'
        action = f'WRITE (was {repr(old)})'
    else:
        cell.value = f'{old}{MARK}{newtext}'
        action = 'APPEND'
    if fill is not None:
        cell.fill = fill
    changelog.append(dict(tab='Funding Table', row=row, inst=inst, field=HDR[col-1],
                          old=('' if old is None else str(old))[:120], new=f'{action}: {newtext[:160]}',
                          source=source, date=D))

# ---------------------------------------------------------------- TASK 1: Chex/EWS enrichment
maprows = list(csv.DictReader(open(f'{BASE}/out/ews_chex_mapping.csv')))
FLAGGED = lambda n: ('NEW VS' in n.upper() or 'UPGRADE' in n.upper())
# unflagged ranks that still carry a real new determination vs N/M
EXTRA = {'63', '73', '76', '81', '98', '139', '142', '145'}

def chex_text(r):
    p, rep = r['Chex_Pull'].strip(), r['Chex_Report'].strip()
    if p.upper().startswith('UNKNOWN'):
        return None
    if p.startswith('Yes'):
        qual = p[3:].strip(' —-')
        t = 'Yes — pulls'
        if rep.startswith('Yes'):
            t += ' + reports'
            qr = rep[3:].strip(' —-')
            if qr:
                qual = (qual + '; ' if qual else '') + qr.strip('()')
        if qual:
            t += f" — {qual.strip('()')}" if not qual.startswith('(') else f' {qual}'
    elif p.startswith('Mixed'):
        t = 'Mixed — ' + p[5:].strip(' —-()')
    elif p.startswith('No'):
        t = 'No — does not pull' + (' ' + p[2:].strip() if p[2:].strip() else '')
    else:
        return None
    return f"{t} ({r['Evidence_grade']}; DP {r['DP_date']}; v6 {D})"

def ews_text(r):
    p, rep = r['EWS_Pull'].strip(), r['EWS_Report'].strip()
    pu, ru = p.upper().startswith('UNKNOWN'), rep.upper().startswith('UNKNOWN')
    if pu and ru:
        return None, None
    if p.startswith('No') and rep.startswith('No'):
        t, fill = 'No — no EWS inquiry + does not report (EWS-CLEAN)', GOLD
    elif p.startswith('No'):
        t, fill = f"No EWS inquiry on opening DP{(' ' + p[2:].strip()) if p[2:].strip() else ''}", None
    elif p.startswith('Yes') and rep.startswith('Yes'):
        t, fill = 'Yes — pulls + reports', ORANGE
    elif p.startswith('Yes'):
        t, fill = 'Yes — pulls', ORANGE
    elif rep.startswith('Yes'):
        t, fill = 'Reports to EWS (no inquiry DP)', ORANGE
    else:
        return None, None
    return f"{t} ({r['Evidence_grade']}; DP {r['DP_date']}; v6 {D})", fill

n_chex = n_ews = 0
enriched_rows = set()
for r in maprows:
    note = r['Note']
    if not (FLAGGED(note) or r['Rec_Rank'] in EXTRA):
        continue
    # skip flagged rows whose note says nothing new actually found
    if r['Rec_Rank'] in ('10', '39', '136'):  # 'existing cell stands' rows
        continue
    row = rank2row.get(r['Rec_Rank'])
    if row is None:
        print('!! no row for rank', r['Rec_Rank'], r['Institution']); continue
    bank = str(ws.cell(row, 3).value)
    src = f"ews_chex_mapping.csv; {r['Source_links'][:200]}"
    nshort = note.replace('NEW vs N/M: ', '').replace('UPGRADE vs existing EWS=N/M: ', '')[:180]
    did = False
    # Chex (col 7)
    ct = chex_text(r)
    if ct and (is_nm(ws.cell(row, 7).value) or 'UPGRADE' in note.upper() or r['Rec_Rank'] in ('9', '18', '83', '38')):
        existing_sub = not is_nm(ws.cell(row, 7).value)
        fill = ORANGE if (ct.startswith('Yes') or ct.startswith('Mixed')) else None
        if existing_sub:
            # only append genuinely-new chex intel for the EWS-upgrade rows that also carry chex news
            chex_news = {'9':  'biz accounts reported to BUSINESS Chex (separate file; DP 2024-05-05 #1840439); sens page re-confirms NOT sensitive (approved 13+)',
                         '18': 'Chex sens page: declined only at 32+/12mo — tolerant of moderate inquiry counts (fetched 2026-06-10)',
                         '38': 'SENSITIVE per sens page: declined 7/12, denied 12/6 (fetched 2026-06-10)',
                         '83': 'sens page: TIAA Direct NOT inquiry-sensitive (50+/12 OK; fetched 2026-06-10)'}
            t = chex_news.get(r['Rec_Rank'])
            if t:
                upd(row, 7, t, src, fill=None, force_append=True); n_chex += 1; did = True
        else:
            upd(row, 7, ct, src, fill=fill); n_chex += 1; did = True
    # EWS (col 8)
    et, efill = ews_text(r)
    if et and is_nm(ws.cell(row, 8).value):
        # age-caveat for old single DPs
        if r['Rec_Rank'] in ('36', '150'):
            et = et.replace(f'; v6 {D})', f'; OLD single DP — re-verify; v6 {D})')
            efill = None
        upd(row, 8, et, src, fill=efill); n_ews += 1; did = True
    if did:
        enriched_rows.add(row)
        upd(row, 36, f"Chex/EWS v6: {nshort} | {r['Source_links'][:230]}", 'ews_chex_mapping.csv', force_append=True)

print(f'Task 1: {n_chex} Chex cells + {n_ews} EWS cells enriched across {len(enriched_rows)} rows')

# ---------------------------------------------------------------- TASK 2: Enterprise Bank & Trust (rank 26)
r = rank2row['26']
assert 'Enterprise' in ws.cell(r, 3).value
SRC_EBT = 'enterprise_bt_dossier.md'
upd(r, 16, 'VERIFIED (HIGH) — LNCRCD $3.155M credit-card loans on OWN call report 2026-03-31 (agent banks hold $0); upgraded from MED-HIGH inference', SRC_EBT)
upd(r, 18, 'VERIFIED HIGH self-issuer: $3.155M own-balance-sheet card loans (FDIC call rpt LNCRCD, 2026-03-31); absent from CFPB agreement DB (800 issuers) = de-minimis self-issuer exemption; 0 Elan mentions in raw HTML; Elan tells traced to name-twin enterprisebankpgh.com; agreement PDF still unpublished (CFPB-exempt) — issuer-of-record proven via balance sheet instead', SRC_EBT)
upd(r, 11, f'still UNVERIFIED as of {D} — zero public DPs across Reddit/myFICO/CreditBoards/DoC (datapoint desert); claim lives only inside paid broker communities', SRC_EBT)
upd(r, 9, f'Banker: Jessica Morrow, AVP Business Banker (theorg.com verified {D}), jmorrow@enterprisebank.com — NOT “Jordan” (name in earlier brief was an error) (v6 {D})', SRC_EBT)
upd(r, 25, f'banker name corrected: JESSICA Morrow AVP (theorg.com {D}; “Jordan” appears nowhere); CFPB agreement DB has NO EBT entry (de-minimis exemption — consistent with self-issuance, not proof of agent); SBA Express ≤$500K remains the plausible app-only vehicle; BBB complaints vs Freedom Funders stand — get terms in writing from jmorrow@/(833) 896-2850', SRC_EBT)
upd(r, 14, f'8 SoCal branches re-verified vs FDIC SOD 2025 (CERT 27237) {D}; Rowland Heights (legacy First Choice) CLOSED/absent from SOD 2025', SRC_EBT)
upd(r, 23, f'confirmed datapoint desert as of {D}: absent from DoC/CardRight/UponArriving pull lists', SRC_EBT)
upd(r, 24, f'N/M — SBFE membership not publicly checkable (anonymized; sbfe.org members page 404s); ask banker (v6 {D})', SRC_EBT)
upd(r, 7, f'N/M re-confirmed {D}: absent from DoC Chex+EWS lists (docs_D3/D5 re-checked); no web DPs either way', SRC_EBT)
upd(r, 8, f'N/M — absent from DoC Chex+EWS lists (docs_D3 re-checked {D}); no DPs either way (v6 {D})', SRC_EBT)
upd(r, 20, f'phone (833) 896-2850 / email / branch RM — confirmed NO online application (only tel: link in page HTML, {D})', SRC_EBT)
upd(r, 35, f'HIGH self-issuer (VERIFIED via balance sheet) + products + footprint ({D}); $50K claim still UNVERIFIED; HP bureau + biz-bureau reporting = first-party banker questions', SRC_EBT)
upd(r, 36, f'api.fdic.gov financials LNCRCD CERT 27237 ({D}); CFPB cfpbIssuers list (800 issuers, no EBT entry, {D}); theorg.com/org/enterprise-bank-trust/org-chart/jessica-morrow; FDIC SOD 2025 sod_2025_*.json', SRC_EBT)
print('Task 2: Enterprise B&T enriched (13 cells)')

# ---------------------------------------------------------------- TASK 3: LAFCU (rank 12)
r = rank2row['12']
assert 'LAFCU' in ws.cell(r, 3).value
SRC_L = 'lafcu_dossier.md'
upd(r, 7, f'Chex + hard credit pull re-verified VERBATIM on TWO live pages (/accounts/credit-union-membership AND /accounts/checking, {D}); NO second-chance product found anywhere on live site — prior “2nd-chance if Chex balances paid” claim DOWNGRADED to UNVERIFIED (search summary may conflate LAPFCU)', SRC_L)
upd(r, 23, f'card-pull bureau CONFLICTING: 2022 denial pull rendered EQ in one {D} search vs EX in v5 (thread #6596415 now 403s, could not re-read); 2023 approval = TU — treat as multi-bureau risk', SRC_L)
upd(r, 11, f'$50K max / 15.75% var / $0 AF re-verified {D} (lafcu.org/rates eff 6-10-26 + /business-loans); still zero public business DPs', SRC_L)
upd(r, 17, f're-verified {D} (rates page eff 6-10-26)', SRC_L)
upd(r, 25, f'2-yr TIB gate re-confirmed VERBATIM {D}: “The business must be established and operating for a minimum of two full years”; 48-hr prequal + no app fee + local decisioning all verbatim same page', SRC_L)
upd(r, 29, f're-verified {D}: business accounts in-branch ONLY via LACA; verbatim “Your business must be located in Southern California”; LAFCU declines high-volume-activity business accounts (own site)', SRC_L)
upd(r, 14, f'NCUA charter #1207 confirmed (chartered 1936-03-31; v6 {D})', SRC_L)
upd(r, 18, f'consumer Visa agreement names LAFCU as creditor verbatim (fetched {D}); business-card agreement not publicly posted (inferred same issuer — same in-house LoansPQ portal, no agent branding); Business LOC APR unpublished', SRC_L)
upd(r, 35, f'v6 re-verify {D}: $50K/15.75%/$0AF + 2-yr gate verbatim re-confirmed; Chex+hard-pull verified on 2 live pages; 2nd-chance claim DOWNGRADED to unverified; card bureaus conflicting (EQ-or-EX 2022 vs TU 2023)', SRC_L)
upd(r, 36, f'lafcu.org/disclosures/visa-variable-credit-cards-agreement-and-disclosure (creditor verbatim, {D}); /accounts/checking (Chex verbatim 2nd page); creditunions.org/c/los-angeles-federal-credit-union-1207', SRC_L)
print('Task 3: LAFCU enriched (10 cells)')

# ---------------------------------------------------------------- TASK 4: gold-lead intel on existing rows
# U.S. Bank (rank 4)
r = rank2row['4']
assert 'U.S. Bank' in ws.cell(r, 3).value
upd(r, 28, f'GOLD INTEL VERIFIED usbank.com {D}: Cash Flow Manager lines UNDER $50K — “financial statements are not normally required”; >$50K = last FY tax return; decision up to $250K; $150/yr fee ≤$50K ($0 above); PG required; online or branch → apply at $49K to stay in the no-financials tier', 'gold_leads.csv')
upd(r, 35, f'v6 {D}: Cash Flow Manager <$50K no-financials tier VERIFIED on bank site', 'gold_leads.csv')
upd(r, 36, f'usbank.com/business-banking/business-lending/business-lines-of-credit/cash-flow-manager.html + KB0209629 (VERIFIED {D})', 'gold_leads.csv')
# Amex (rank 6)
r = rank2row['6']
assert 'American Express' in ws.cell(r, 3).value
upd(r, 28, f're-verified on issuer site {D}: Business Blueprint BLOC $2K–$250K, real-time bank-account-link underwriting — NO tax returns/financial statements for most line sizes (>$150K needs existing Amex relationship); 660+ FICO, 12 mo TIB, ~$3K/mo rev; monthly-fee pricing (costlier than bank LOC)', 'gold_leads.csv')
upd(r, 36, f'americanexpress.com/en-us/business/blueprint/business-line-of-credit/ + help-center applying FAQ (VERIFIED {D})', 'gold_leads.csv')
# Community West Bank (rank 54)
r = rank2row['54']
assert 'Community West' in ws.cell(r, 3).value
upd(r, 14, f'absorbed United Security Bank (Fresno, $1.2B) effective 2026-04-01 (all-stock ~$191.9M; USB CERT 27132 inactive 2026-04-03) — closes the v5 “lone CA-HQ gap”; USB itself had NO biz card (sitemap verified {D})', 'w2_banks_screen.csv')
upd(r, 18, f'v6 screen attributes the cardaccount.net business-app rail to TCM Bank N.A. (ICBA) (INFERRED from domain, {D}; Elan markers absent) — CONFLICTS with v5 TIB attribution; verify issuer with bank before applying', 'w2_banks_screen.csv')
upd(r, 36, f'businesswire.com/news/home/20260401074081/en/ (merger close 2026-04-01); communitywestbank.com/business-credit-cards (checked {D})', 'w2_banks_screen.csv')
print('Task 4: U.S. Bank, Amex, Community West enriched (Citizens N.A. has no v5 row -> new reference row in part 2)')

wb.save(WB)
json.dump(changelog, open(f'{BASE}/changelog_part1.json', 'w'))
print('cells changed (incl. Sources appends):', len(changelog))
